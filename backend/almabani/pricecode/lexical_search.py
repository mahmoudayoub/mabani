"""
Lexical Search Engine for Price Code candidate retrieval.

Replaces embedding + vector search with a SQLite-backed TF-IDF/BM25-style
matching engine.  Domain-aware normalization, synonym groups, engineering
spec extraction, discipline routing and hard spec filters produce high-
precision candidates for the downstream LLM judge.

Public API
----------
build_index(db_path, ref_paths, rebuild)   – create / refresh SQLite index
LexicalMatcher(db_path, source_files, …)   – search candidates for a BOQ item
"""

from __future__ import annotations

import asyncio
import hashlib
import io
import logging
import math
import os
import re
import sqlite3

import zipfile
from collections import Counter, defaultdict
from rapidfuzz.fuzz import ratio as _rapidfuzz_ratio
from typing import Any, Dict, Iterator, List, Optional, Sequence, Set, Tuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

SCHEMA_VERSION = "lexical_v4"

# ═════════════════════════════════════════════════════════════════════════
# Constants
# ═════════════════════════════════════════════════════════════════════════

STOPWORDS = {
    "a", "an", "and", "or", "the", "of", "to", "for", "in", "on", "with",
    "by", "at", "from", "as", "is", "are", "be", "been", "being",
    "including", "include", "includes", "complete", "supply", "install",
    "installed", "installation", "all", "type", "only", "item", "items",
    "work", "works", "say", "each", "no", "nr", "per", "into", "up",
    "down", "over", "under", "than", "shall", "contractor", "required",
    "accordance", "drawing", "drawings", "specification", "specifications",
    "approval", "engineer", "documents", "testing", "commissioning",
    "warranty", "allow", "allowance", "provide", "provided", "etc",
    "necessary", "ready", "use", "inclusive", "generally", "comprising",
    "comprise", "materials", "material", "tools", "equipment", "services",
    "accessories", "consumables", "like",
}

# First element is the canonical form.
SYNONYM_GROUPS: List[List[str]] = [
    # Earthwork verbs
    ["excavation", "excavate"],
    ["disposal", "dispose"],
    ["removal", "remove"],
    ["hauling", "carting"],
    ["compaction", "compact", "compacting"],
    ["backfill", "backfilling"],
    ["demolition", "demolish"],
    ["dismantling", "dismantle"],
    ["grading", "grade"],
    ["trimming", "trim"],
    ["preparation", "prepare", "preparing"],
    # Wet trades
    ["waterproofing", "waterproof"],
    ["dewatering", "dewater"],
    ["plastering", "plaster"],
    ["rendering", "render"],
    # Materials
    ["reinforcement", "reinforcing", "rebar"],
    ["bituminous", "asphalt"],
    ["geotextile", "geofabric"],
    # Objects – only clear equivalents
    ["footing", "foundation"],
    ["manhole", "chamber"],
    ["kerb", "curb", "curbstone", "kerbstone"],
    ["gully", "catchpit"],
    # ── NEW: gap-analysis-driven synonym groups ──
    # Piping
    ["pipe", "piping"],
    ["fitting", "fittings"],
    ["valve", "valves"],
    # Structural
    ["beam", "beams"],
    ["column", "columns"],
    ["slab", "slabs"],
    # MEP
    ["fan", "fcu"],
    ["duct", "ducting"],
    ["socket", "outlet", "point"],
    ["isolator", "isolation"],
    ["sprinkler"],
    # Water
    ["potable"],
    ["cooling"],
    # Surface / finish
    ["aggregate", "granular"],
    ["blinding", "pcc"],
    ["screed", "levelling"],
    # Cable
    ["lsoh", "lszh"],
    ["aluminium", "aluminum"],
    # Misc
    ["intercom", "intercommunication"],
    ["tank", "cistern", "reservoir"],
    ["mesh"],
    ["roof", "roofing"],
    ["insulation", "insulated"],
    # Fire protection
    ["extinguisher"],
    ["hose"],
    ["nozzle"],
    ["detector"],
    ["damper"],
    # HVAC duct
    ["grille", "grilles"],
    ["diffuser", "diffusers"],
    ["attenuator"],
    ["louvre", "louver"],
    # Pumps
    ["pump", "pumps"],
    ["motor"],
    ["compressor"],
    # General MEP
    ["meter"],
    ["gauge"],
    ["sensor"],
    ["thermostat"],
    ["controller"],
    ["strainer"],
    ["expansion"],
    ["hanger", "support"],
]

_SYNONYM_MAP: Dict[str, str] = {}
for _group in SYNONYM_GROUPS:
    _canonical = _group[0]
    for _term in _group:
        _SYNONYM_MAP[_term] = _canonical

DISCIPLINE_HINTS: Dict[str, Set[str]] = {
    "civil": {
        "excavation", "excavate", "concrete", "rebar", "reinforcement",
        "footing", "slab", "foundation", "blockwork", "masonry", "roof",
        "paint", "tile", "door", "window", "finishes", "earthwork",
        "earthworks", "asphalt", "waterproof", "formwork", "demolition",
        "road", "pavement", "curb", "grubbing", "fence", "hoarding",
        "subgrade", "basecourse", "bituminous", "tack", "prime",
        "embankment", "joint", "tie", "bar", "manhole", "chamber",
        "culvert", "ditch", "channel", "geogrid", "geotextile",
        "geostrip", "erosion", "grading", "blinding", "kerb",
        "disposal", "backfill", "backfilling", "trimming",
        "compaction", "fill", "lean_concrete", "subbase",
        "aggregate", "granular", "plaster", "screed",
    },
    "electrical": {
        "lighting", "light", "cctv", "security", "access", "control",
        "cable", "tray", "containment", "socket", "power", "distribution",
        "db", "fire", "alarm", "ups", "communications", "network",
        "telephone", "earthing", "switch", "signage", "emergency", "exit",
        "generator", "lv", "mv", "hv", "transformer", "airfield",
        "runway", "taxiway", "apron", "agl", "ilcms", "sign", "deep",
        "base", "l824", "l831", "l852", "isolation", "smdb", "mdb",
        "sdb", "mcc", "xlpe", "swa",
    },
    "mechanical": {
        "pipe", "piping", "valve", "pump", "tank", "sanitary", "water",
        "sewage", "stormwater", "storm", "hvac", "duct", "ductwork",
        "fan", "chiller", "sprinkler", "firefighting", "plumbing",
        "drain", "drainage", "manhole", "chamber", "irrigation",
        "potable", "fuel", "lpg", "vent", "waste", "network",
        "conduit", "rcp", "hdpe", "upvc", "grp", "gre", "ductile",
        "iron", "tse", "diffuser", "ahu", "fcu", "damper", "grille",
        "louver", "condensate", "refrigerant",
    },
}

AIRFIELD_HINTS: Set[str] = {
    "agl", "airfield", "runway", "taxiway", "apron", "ilcms",
    "isolation", "transformer", "l824", "l831", "l852", "deep", "base",
    "guidance", "sign", "signage", "light",
}

# ── Shared discipline-inference rules ───────────────────────────────────
# Used at *index time* (from the file name) and at *search time* (from
# source_file + sheet_name) so a single consistent list drives both.
# Each entry is (substring, discipline).  The list is evaluated top-down;
# the first matching substring wins.  Put more specific patterns before
# broader ones to avoid false positives (e.g. "firefight" before "fire").
_DISC_INFER_RULES: List[Tuple[str, str]] = [
    # ── Electrical (narrow terms first) ──
    ("electrical",       "electrical"),
    ("communic",         "electrical"),
    ("security",         "electrical"),
    ("transport",        "electrical"),   # a_Transportation sheets
    ("cctv",             "electrical"),
    ("fire alarm",       "electrical"),   # fire-alarm is Div 28 (elec)
    ("low current",      "electrical"),
    # ── Mechanical (check before civil — "fire" is broad) ──
    ("mechanical",       "mechanical"),
    ("plumbing",         "mechanical"),
    ("hvac",             "mechanical"),
    ("firefight",        "mechanical"),   # Mech - Firefighting
    ("fire suppress",    "mechanical"),
    ("fire protect",     "mechanical"),
    ("sprinkler",        "mechanical"),
    ("district cool",    "mechanical"),
    ("chiller",          "mechanical"),
    ("ventilat",         "mechanical"),
    ("cooling",          "mechanical"),
    ("heating",          "mechanical"),
    ("fuel",             "mechanical"),   # fuel systems, LPG
    ("lpg",              "mechanical"),
    ("irrigation",       "mechanical"),
    ("sewage",           "mechanical"),
    ("potable",          "mechanical"),
    ("sanitary",         "mechanical"),
    ("storm network",    "mechanical"),
    ("drainage network", "mechanical"),
    ("piping",           "mechanical"),
    # ── Civil (broadest — checked last as fallback) ──
    ("civil",            "civil"),
    ("concrete",         "civil"),
    ("masonry",          "civil"),
    ("earthwork",        "civil"),
    ("finish",           "civil"),        # F_Finishes
    ("metal",            "civil"),        # S_Metals
    ("thermal",          "civil"),        # W_Thermal And Moisture
    ("moisture",         "civil"),
    ("waterproof",       "civil"),
    ("opening",          "civil"),        # O_Openings
    ("roadwork",         "civil"),        # Y_Roadworks
    ("structural",       "civil"),
    ("demolit",          "civil"),
    ("excavat",          "civil"),
    ("landscape",        "civil"),
    ("paving",           "civil"),
    ("pavement",         "civil"),
    ("asphalt",          "civil"),
    ("formwork",         "civil"),
    ("reinforc",         "civil"),
    ("rebar",            "civil"),
    ("paint",            "civil"),
    ("tile",             "civil"),        # tiling
    ("door",             "civil"),
    ("window",           "civil"),
    ("roof",             "civil"),
    ("fence",            "civil"),
    ("piling",           "civil"),
    ("pile",             "civil"),
    ("foundation",       "civil"),
    ("existing condition", "civil"),     # B_Existing Conditions
    ("utilit",           "civil"),        # Z_Utilities (civil side)
]


def _infer_discipline_from_context(*texts: str) -> str:
    """Return inferred discipline from one or more context strings.

    Scans *texts* (lowercased, concatenated) against ``_DISC_INFER_RULES``.
    Returns the discipline of the first matching keyword, or ``'unknown'``.
    """
    blob = " ".join(clean_text(t).lower().replace("_", " ") for t in texts if t)
    for keyword, disc in _DISC_INFER_RULES:
        if keyword in blob:
            return disc
    return "unknown"


DOMAIN_ALIASES = [
    (re.compile(r"\bagl\b", re.I), "airfield lighting airfield signaling control equipment"),
    (re.compile(r"\bilcms\b", re.I), "airfield lighting control monitoring system"),
    (re.compile(r"\bl-?824\b", re.I), "l824 airfield cable series circuit cable"),
    (re.compile(r"\bl-?831(?:-?1)?\b", re.I), "l831 isolation transformer"),
    (re.compile(r"\bl-?852c?\b", re.I), "l852 airfield sign or airfield light"),
    (re.compile(r"\bdeep base cans?\b", re.I), "deep base can light base"),
    (re.compile(r"\bduct bank\b", re.I), "electrical duct bank conduit bank"),
    (re.compile(r"\brcp\b", re.I), "reinforced concrete pipe rc pipe stormwater pipe"),
    (re.compile(r"\bpcc\b", re.I), "plain cement concrete blinding concrete"),
    (re.compile(r"\brcc\b", re.I), "reinforced cement concrete reinforced concrete"),
    (re.compile(r"\bupvc\b", re.I), "unplasticized pvc pipe"),
    (re.compile(r"\bpvc-u\b", re.I), "unplasticized pvc pipe"),
    (re.compile(r"\bhdpe\b", re.I), "high density polyethylene pipe"),
    (re.compile(r"\bgrp\b", re.I), "glass reinforced plastic pipe"),
    (re.compile(r"\bgre\b", re.I), "glass reinforced epoxy pipe"),
    (re.compile(r"\bxlpe\b", re.I), "cross linked polyethylene cable"),
    (re.compile(r"\bswa\b", re.I), "steel wire armoured armored cable"),
    (re.compile(r"\bsmdb\b", re.I), "sub main distribution board"),
    (re.compile(r"\bmdb\b", re.I), "main distribution board"),
    (re.compile(r"\bsdb\b", re.I), "sub distribution board"),
    (re.compile(r"\bmcc\b", re.I), "motor control center"),
    (re.compile(r"\bwp\b", re.I), "waterproof waterproofing"),
    (re.compile(r"\btse\b", re.I), "treated sewage effluent irrigation water"),
    (re.compile(r"\bblock paving\b", re.I), "paving block pavers interlock paving"),
    # ── NEW: gap-analysis-driven aliases ──
    (re.compile(r"\blsoh\b", re.I), "low smoke zero halogen cable lszh"),
    (re.compile(r"\blszh\b", re.I), "low smoke zero halogen cable lsoh"),
    (re.compile(r"\bfcu\b", re.I), "fan coil unit"),
    (re.compile(r"\bahu\b", re.I), "air handling unit"),
    (re.compile(r"\bcwp\b", re.I), "chilled water pump"),
    (re.compile(r"\bchwp\b", re.I), "chilled water pump"),
    (re.compile(r"\bbtu\b", re.I), "btu meter energy meter"),
    (re.compile(r"\bgi\b", re.I), "galvanized iron galvanised"),
    (re.compile(r"\bms\b", re.I), "mild steel"),
    (re.compile(r"\bss\b", re.I), "stainless steel"),
    (re.compile(r"\bdi\b", re.I), "ductile iron"),
    (re.compile(r"\brc\b", re.I), "reinforced concrete"),
    (re.compile(r"\bppr\b", re.I), "polypropylene random pipe"),
    (re.compile(r"\bcpvc\b", re.I), "chlorinated pvc pipe"),
    (re.compile(r"\bdb\b", re.I), "distribution board"),
    # ── Earthwork / concrete enrichment ──
    (re.compile(r"\bblinding\b", re.I), "lean_concrete blinding_concrete"),
    (re.compile(r"\btrimming\s+(?:excavat|surface)", re.I), "earthwork surface_preparation"),
    (re.compile(r"\bpreparing\s+(?:excavat|surface)", re.I), "earthwork surface_preparation"),
    (re.compile(r"\bdisposal\s+(?:of\s+)?(?:excavat|surplus|material)", re.I), "earthwork disposal_offsite"),
    (re.compile(r"\bformation\s+level", re.I), "earthwork grading subgrade"),
    # ── Multi-word synonym phrases (can't be per-token synonyms) ──
    (re.compile(r"\bfan\s+coil\s+unit", re.I), "fcu"),
    (re.compile(r"\bsocket\s+outlet", re.I), "socket power point"),
    (re.compile(r"\bfire\s+suppression", re.I), "sprinkler firefighting"),
    (re.compile(r"\bdomestic\s+water", re.I), "potable water"),
    (re.compile(r"\bcold\s+water", re.I), "potable chilled water"),
    (re.compile(r"\bchilled\s+water", re.I), "cooling district cooling"),
    (re.compile(r"\bdistrict\s+cooling", re.I), "cooling chilled water"),
    (re.compile(r"\bcrushed\s+stone", re.I), "aggregate granular"),
    (re.compile(r"\blean\s+concrete", re.I), "blinding pcc"),
    (re.compile(r"\bfloor\s+screed", re.I), "screed levelling"),
    (re.compile(r"\blow\s+smoke\s+zero\s+halogen", re.I), "lsoh lszh"),
    (re.compile(r"\bwire\s+mesh", re.I), "mesh welded"),
    (re.compile(r"\bwelded\s+mesh", re.I), "mesh wire"),
    # ── Fire protection / suppression recall ──
    (re.compile(r"\bclean\s+agent", re.I), "fire suppression gas system fm200 novec"),
    (re.compile(r"\bfm-?200\b", re.I), "clean agent fire suppression gas system"),
    (re.compile(r"\bnovec\b", re.I), "clean agent fire suppression gas system"),
    (re.compile(r"\binert\s+gas", re.I), "clean agent fire suppression system"),
    (re.compile(r"\bfire\s+extinguisher", re.I), "portable extinguisher dry chemical co2"),
    (re.compile(r"\bfire\s+blanket", re.I), "fire blanket safety equipment"),
    (re.compile(r"\bfire\s+hose\s+cabinet", re.I), "fhc fire hose cabinet reel"),
    (re.compile(r"\bfhc\b", re.I), "fire hose cabinet"),
    # ── HVAC duct / accessories recall ──
    (re.compile(r"\bflexible\s+duct", re.I), "flex duct connection"),
    (re.compile(r"\bflex\s+duct", re.I), "flexible duct connection"),
    (re.compile(r"\bcircular\s+duct", re.I), "round duct spiral duct"),
    (re.compile(r"\brectangular\s+duct", re.I), "rect duct galvanized sheet metal"),
    (re.compile(r"\bfire\s+damper", re.I), "fire rated damper volume control"),
    (re.compile(r"\bvolume\s+control\s+damper", re.I), "vcd damper"),
    (re.compile(r"\bvcd\b", re.I), "volume control damper"),
    (re.compile(r"\bvav\b", re.I), "variable air volume box"),
    (re.compile(r"\bsound\s+attenuator", re.I), "silencer acoustic attenuator"),
    (re.compile(r"\bsilencer\b", re.I), "sound attenuator acoustic"),
    # ── Pumps / mechanical equipment recall ──
    (re.compile(r"\bsubmersible\s+pump", re.I), "sump pump drainage pump"),
    (re.compile(r"\bbooster\s+pump", re.I), "pressure pump booster set"),
    (re.compile(r"\bcentrifugal\s+pump", re.I), "inline pump circulation pump"),
    (re.compile(r"\bheat\s+exchanger", re.I), "plate heat exchanger phe calorifier"),
    (re.compile(r"\bcalorifier", re.I), "heat exchanger hot water cylinder"),
    (re.compile(r"\bexpansion\s+(?:tank|vessel)", re.I), "expansion vessel pressure tank"),
    # ── Gas / LPG recall ──
    (re.compile(r"\blpg\b", re.I), "liquefied petroleum gas fuel"),
    (re.compile(r"\bnatural\s+gas", re.I), "gas piping fuel"),
    (re.compile(r"\bgas\s+detection", re.I), "gas detector sensor leak"),
]

# ── Unit compatibility map ──────────────────────────────────────────────
# Groups of BOQ unit strings that share the same physical dimension.
# Used to penalise candidates whose implied unit family doesn't match.
_UNIT_FAMILIES: Dict[str, str] = {}
_UNIT_GROUPS: List[Tuple[str, List[str]]] = [
    ("volume", ["m3", "m³", "cm", "cu.m", "cum", "cub.m", "cubic meter", "cubic metre"]),
    ("area", ["m2", "m²", "sq.m", "sqm", "sq m", "square"]),
    ("length", ["m", "lm", "l.m", "rm", "r.m", "lin.m"]),
    ("count", ["nr", "no", "no.", "nos", "pcs", "pc", "set", "pair", "unit", "ea"]),
    ("weight", ["kg", "t", "ton", "tonne", "tonnes"]),
    ("lump", ["ls", "item", "lot", "sum"]),
    ("power", ["kw", "kva", "hp"]),
]
for _family, _aliases in _UNIT_GROUPS:
    for _a in _aliases:
        _UNIT_FAMILIES[_a.lower()] = _family


def _unit_family(unit_str: str) -> Optional[str]:
    """Return the unit family for a BOQ unit string, or None if unknown."""
    u = clean_text(unit_str).lower().strip().rstrip(".")
    if u in _UNIT_FAMILIES:
        return _UNIT_FAMILIES[u]
    # Partial matches for compound units
    for alias, family in _UNIT_FAMILIES.items():
        if alias in u:
            return family
    return None


# Ref descriptions contain implicit unit clues.  This maps keywords in ref
# descriptions to the expected physical dimension.
_REF_UNIT_HINTS: List[Tuple[str, re.Pattern]] = [
    ("volume",  re.compile(r"\bdepth\b|\bthick\b|\bm3\b|\bcubic\b|\bvolume\b|\bbulk excav", re.I)),
    ("area",    re.compile(r"\bm2\b|\bsq\b|\barea\b|\bsurface\b|\bformwork\b|\bplaster\b|\bpaint(?:ing)?\b|\btil(?:e|ing)\b|\bpaving\b|\bwaterproof", re.I)),
    ("length",  re.compile(r"\blinear\b|\blength\b|\bper m\b|\bkerb\b|\bcurb\b|\bcurbstone\b|\bpipe\b|\bcable\b|\bconduit\b|\bduct\b|\btrench\b|\bgutter\b|\bsidewalk\b|\bedge\b", re.I)),
    ("count",   re.compile(r"\beach\b|\bnumber\b|\bunit\b|\bfitting\b|\bvalve\b|\bpump\b|\btransformer\b|\blight\b|\bcamera\b|\bsensor\b|\bmanhole\b|\bchamber\b|\bgully\b", re.I)),
    ("weight",  re.compile(r"\bkg\b|\btonne\b|\bton\b|\bsteel.*weight\b|\brebar\b|\breinforcement\b", re.I)),
]


def _infer_ref_unit_family(ref_text: str) -> Optional[str]:
    """Guess the unit family from a reference description."""
    for family, pattern in _REF_UNIT_HINTS:
        if pattern.search(ref_text):
            return family
    return None


# ═════════════════════════════════════════════════════════════════════════
# Scope detection from hierarchy context
# ═════════════════════════════════════════════════════════════════════════

_RE_SUPPLY_AND_INSTALL = re.compile(
    r"\bsupply\s+(?:and|&)\s+install", re.I
)
_RE_POUR_WITH_LABOUR = re.compile(
    r"\bpour\b", re.I
)
_RE_LABOUR_OR_MATERIAL = re.compile(
    r"\blabour\b|\ball\s+necessary\s+material", re.I
)
_RE_SUPPLY_ONLY = re.compile(
    r"\bsupply\b", re.I
)
_RE_NO_INSTALL = re.compile(
    r"\binstall|\bpour\b|\berect|\bplace\b|\blay(?:ing)?\b", re.I
)
_RE_REINFORCEMENT = re.compile(
    r"\breinforcement\b|\breinforcing\b|\brebar\b|\bsteel\s+bar\b", re.I
)


def _detect_expected_scope(parent: str, grandparent: str) -> Optional[str]:
    """Detect expected scope letter from hierarchy context.

    Only returns a scope when the evidence is strong and universal.
    Focuses on Supply-only (E) vs Supply+Install (F) detection which
    are the most common scope errors and have consistent meaning across
    all disciplines.  Concrete-specific scopes (A/B/C/D) are NOT
    detected here because their meaning is category-dependent.

    Works on the RAW parent/grandparent text (before stopword removal)
    so keywords like 'supply' and 'install' are visible.
    """
    # Use parent primarily; grandparent only as fallback context
    ctx = f"{parent} ; {grandparent}"

    # 1. "Supply and install…" → full scope F
    #    Exception: "Supply and install reinforcing bars/reinforcement" is
    #    specifically about rebar S+I — the right scope can be B or D,
    #    not necessarily F.  Return None to let the LLM decide.
    if _RE_SUPPLY_AND_INSTALL.search(ctx):
        if not _RE_REINFORCEMENT.search(parent):
            return "F"
        return None

    # 2. "Pour concrete…include labour / all necessary material" → F
    if _RE_POUR_WITH_LABOUR.search(parent) and _RE_LABOUR_OR_MATERIAL.search(parent):
        return "F"

    # 3. "Supply" without install/pour → Supply Only = E
    if _RE_SUPPLY_ONLY.search(parent):
        if not _RE_NO_INSTALL.search(parent):
            return "E"

    return None


# ═════════════════════════════════════════════════════════════════════════
# MEP sub-discipline routing
# ═════════════════════════════════════════════════════════════════════════

# Maps hierarchy keywords → expected price-code prefix letter.
# Within the broad "mechanical" discipline, plumbing (p), HVAC (h),
# fire protection (f), and utilities (Z) share identical physical
# materials (pipes, valves, insulation).  These hints let the search
# engine boost candidates from the correct sub-discipline.
# Only unambiguous, standard construction terminology is used to
# avoid overfitting to any specific project.
_MEP_PREFIX_HINTS: List[Tuple[re.Pattern, str]] = [
    # Plumbing (p) — building water / drainage piping
    (re.compile(r"\bplumbing\b", re.I), "p"),
    (re.compile(r"\bsanitary\b", re.I), "p"),
    (re.compile(r"\bdomestic\s+(?:hot\s+|cold\s+)?water\b", re.I), "p"),
    (re.compile(r"\bhot\s+water\b", re.I), "p"),
    (re.compile(r"\bcold\s+water\b", re.I), "p"),
    (re.compile(r"\bwater\s+supply\b", re.I), "p"),
    (re.compile(r"\bpotable\s+water\b", re.I), "p"),
    (re.compile(r"\bsewage\b", re.I), "p"),
    (re.compile(r"\bdrainage\b", re.I), "p"),
    (re.compile(r"\bsoil\s+(?:and|&)\s*waste\b", re.I), "p"),
    (re.compile(r"\bwater\s+heater\b", re.I), "p"),
    # Fire protection (f) — BEFORE HVAC so fire-specific terms match first
    (re.compile(r"\bfire\s+(?:protect|suppress|fight)", re.I), "f"),
    (re.compile(r"\bfire\s+hose\b", re.I), "f"),
    (re.compile(r"\bfire\s+damper\b", re.I), "f"),
    (re.compile(r"\bsprinkler\b", re.I), "f"),
    # HVAC (h) — heating, ventilation, air conditioning
    (re.compile(r"\bhvac\b", re.I), "h"),
    (re.compile(r"\bchilled\s+water\b", re.I), "h"),
    (re.compile(r"\bair\s+condition", re.I), "h"),
    (re.compile(r"\bventilation\b", re.I), "h"),
    (re.compile(r"\bdistrict\s+cooling\b", re.I), "h"),
    (re.compile(r"\bcondensate\b", re.I), "h"),
    (re.compile(r"\brefrigerant\b", re.I), "h"),
    (re.compile(r"\bductwork\b", re.I), "h"),
    (re.compile(r"\bduct\b", re.I), "h"),
    (re.compile(r"\bfan\s+coil\b", re.I), "h"),
    (re.compile(r"\bfcu\b", re.I), "h"),
    (re.compile(r"\bahu\b", re.I), "h"),
    (re.compile(r"\bair\s+handl", re.I), "h"),
    (re.compile(r"\bdiffuser\b", re.I), "h"),
    (re.compile(r"\bdamper\b", re.I), "h"),
    (re.compile(r"\bgrilles?\b", re.I), "h"),
    (re.compile(r"\blouv(?:er|re)\b", re.I), "h"),
    # Utilities / External (Z)
    (re.compile(r"\butilit(?:y|ies)\b", re.I), "Z"),
    (re.compile(r"\birrigation\b", re.I), "Z"),
]


def _detect_mep_prefix(
    parent: str, grandparent: str, category_path: str = ""
) -> Optional[str]:
    """Detect expected MEP price-code prefix letter from hierarchy context.

    Scans parent, grandparent, **and category_path** text for standard
    construction keywords that unambiguously identify the MEP sub-
    discipline.  Returns the prefix letter (p/h/f/Z) if the evidence is
    clear, or None if ambiguous / unknown.  First match wins, so more-
    specific patterns should appear before less-specific ones.
    """
    ctx = f"{grandparent} {parent} {category_path}"
    for pattern, prefix in _MEP_PREFIX_HINTS:
        if pattern.search(ctx):
            return prefix
    return None


# Compact code regex: e.g. p1316ACC, h3713B01, Z1411AAA
_COMPACT_CODE_RE = re.compile(
    r'^[A-Za-z]\d{4}([A-Za-z][A-Za-z0-9]{1,2})$'
)


def _extract_scope_letter(price_code: str) -> Optional[str]:
    """Extract the scope letter (last char of the suffix) from a price code.

    Supports both spaced format ``C 31 13 CGA`` → ``A``
    and compact format ``p1316ACC`` → ``C``.
    """
    code = price_code.strip()
    # Spaced format: [Disc] [Cat] [Subcat] [ElemGradeScope]
    parts = code.split()
    if len(parts) >= 4:
        suffix = parts[3]
        if len(suffix) >= 2:
            last = suffix[-1].upper()
            if last.isalpha():
                return last
    # Compact format: X9999XYZ
    m = _COMPACT_CODE_RE.match(code)
    if m:
        suffix = m.group(1)
        last = suffix[-1].upper()
        if last.isalpha():
            return last
    return None


def _parse_compact_code(price_code: str):
    """Parse a compact code into (disc, cat, subcat, suffix) or None.

    ``p1316ACC`` → ('p', '13', '16', 'ACC')
    ``h3713B01`` → ('h', '37', '13', 'B01')
    """
    code = price_code.strip()
    m = re.match(r'^([A-Za-z])(\d{2})(\d{2})([A-Za-z][A-Za-z0-9]{1,2})$', code)
    if m:
        return m.group(1), m.group(2), m.group(3), m.group(4)
    return None


# ── Code prefix extraction ──────────────────────────────────────────
# The code prefix identifies the "family" of a code: all codes sharing
# the same prefix describe the same type of work but differ in
# variant dimensions (V1/V2/V3 = size, material, rating, scope, etc.).
#
# Spaced:  "C 31 13 CGA"  → prefix "C 31 13"
# Compact: "Z1411ABC"     → prefix "Z1411"
# Compact: "p1316ACC"     → prefix "p1316"

_PREFIX_SPACED_RE = re.compile(
    r'^([A-Za-z]\s+\d+\s+\d+)\s+[A-Za-z]'
)
_PREFIX_COMPACT_RE = re.compile(
    r'^([A-Za-z]\d{2,4})[A-Za-z]'
)


def extract_code_prefix(price_code: str) -> str:
    """Extract the family-level prefix from a price code.

    Returns the prefix string, or the full code if no pattern matches.
    """
    code = clean_text(price_code)
    if not code:
        return ""
    m = _PREFIX_SPACED_RE.match(code)
    if m:
        return m.group(1)
    m = _PREFIX_COMPACT_RE.match(code)
    if m:
        return m.group(1)
    return code


OBJECT_TOKENS: Set[str] = {
    "pipe", "conduit", "duct", "bank", "sign", "cable", "transformer",
    "light", "base", "joint", "slab", "pavement", "concrete", "manhole",
    "chamber", "valve", "pump", "foundation", "footing", "insulation",
    "can", "culvert", "kerb", "grating", "gully", "board", "camera",
    "barrier", "ladder", "fence", "hoarding",
}

GENERIC_ITEM_PATTERNS = [
    re.compile(r"^\s*pipe size[:\s]", re.I),
    re.compile(r"\bsystem\b", re.I),
    re.compile(r"\binstallations?\b", re.I),
    re.compile(r"^\s*size[:\s]", re.I),
    re.compile(r"^\s*depth\b", re.I),
    re.compile(r"^\s*maximum depth\b", re.I),
    re.compile(r"^\s*max\.? depth\b", re.I),
    re.compile(r"^\s*not exceeding\b", re.I),
    re.compile(r"\b(width|height|length|depth|diameter)\b", re.I),
]


# ═════════════════════════════════════════════════════════════════════════
# Text Processing
# ═════════════════════════════════════════════════════════════════════════

def clean_text(value: object) -> str:
    if value is None:
        return ""
    s = str(value).replace("\xa0", " ").strip()
    if s.lower() in {"nan", "none", "null"}:
        return ""
    return re.sub(r"\s+", " ", s).strip()


def apply_domain_aliases(text: str) -> str:
    out = text
    for pat, repl in DOMAIN_ALIASES:
        if pat.search(out):
            out += " " + repl
    return out


def _canon_num(num: str) -> str:
    n = clean_text(num)
    if not n:
        return ""
    if re.fullmatch(r"\d+\.0+", n):
        return str(int(float(n)))
    if "." in n:
        n = n.rstrip("0").rstrip(".")
    return n


def normalize_text(text: str) -> str:
    text = clean_text(text)
    if not text:
        return ""

    text = apply_domain_aliases(text)
    text = text.lower()
    text = text.replace("–", "-").replace("—", "-").replace("−", "-")
    text = text.replace("×", "x")
    text = text.replace("ø", " diameter ").replace("Ø", " diameter ")

    text = re.sub(r"\banti\s*[- ]?termite\b", "anti termite treatment", text, flags=re.I)
    text = re.sub(r"\buv\s*treatment\b", "ultraviolet treatment", text, flags=re.I)
    text = re.sub(r"\brcp\s*d\s*(\d{2,4})\b", r"rcp dn\1", text, flags=re.I)
    text = re.sub(r"\binternal\s+size\b", "size", text, flags=re.I)
    text = re.sub(r"\bnom(?:inal)?\s+diam(?:eter)?\b", "diameter", text, flags=re.I)

    # Keep cable-size patterns intact before spacing slashes.
    text = re.sub(r"\b(\d+)\s*/\s*(\d+)\s*(?:mm2|mm²|sq\.?\s*mm|sqmm|mm\^2)\b",
                  r"\1/\2mm2", text, flags=re.I)
    text = re.sub(r"\b(\d+)\s*c\s*[x*]?\s*(\d+(?:/\d+)?)\s*(?:mm2|mm²|sq\.?\s*mm|sqmm|mm\^2)\b",
                  r"\1c \2mm2", text, flags=re.I)
    text = re.sub(r"\b(\d+)cx\s*(\d+(?:/\d+)?)\s*(?:mm2|mm²|sq\.?\s*mm|sqmm|mm\^2)\b",
                  r"\1c \2mm2", text, flags=re.I)
    text = re.sub(r"\b(\d+)c\s*(\d+(?:/\d+)?)\s*(?:mm2|mm²|sq\.?\s*mm|sqmm|mm\^2)\b",
                  r"\1c \2mm2", text, flags=re.I)
    text = re.sub(r"\b(\d+)\s*core\s*[x*]?\s*(\d+(?:/\d+)?)\s*(?:mm2|mm²|sq\.?\s*mm|sqmm|mm\^2)\b",
                  r"\1c \2mm2", text, flags=re.I)
    text = re.sub(r"\b(\d+)\s*[x*]\s*(\d+(?:/\d+)?)\s*(?:mm2|mm²|sq\.?\s*mm|sqmm|mm\^2)\b",
                  r"\1c \2mm2", text, flags=re.I)

    text = text.replace("/", " / ")

    # Engineering normalization
    text = re.sub(r"\bdn\s*[-:/]?\s*(\d{2,4})\b", lambda m: f"dn{m.group(1)}", text)
    text = re.sub(
        r"\bd\s*[-:/]?\s*(\d{2,4})\b(?=[^\n,;]{0,24}\b(?:pipe|rcp|culvert|storm|sewer|drain))",
        lambda m: f"dn{m.group(1)}", text,
    )
    # ── Fix: DN ↔ DIA equivalence ──────────────────────────────────
    # BOQ says "50 mm diameter" → dia50; Reference says "DN50" → dn50.
    # They mean the same pipe size.  Emit BOTH tokens so TF-IDF matches.
    text = re.sub(r"\b(\d+)\s*mm\s*(?:dia|diameter)\b",
                  lambda m: f"dia{m.group(1)} dn{m.group(1)} {m.group(1)}mm diameter", text)
    text = re.sub(r"\b(?:dia|diameter)\s*[:\-]?\s*(\d+)\s*mm\b",
                  lambda m: f"dia{m.group(1)} dn{m.group(1)} {m.group(1)}mm diameter", text)
    text = re.sub(r"\b(\d+)\s*cm\s*(?:dia|diameter)\b",
                  lambda m: f"dia{int(m.group(1))*10} dn{int(m.group(1))*10} {int(m.group(1))*10}mm diameter", text)
    text = re.sub(r"\b(?:dia|diameter)\s*[:\-]?\s*(\d+)\s*cm\b",
                  lambda m: f"dia{int(m.group(1))*10} dn{int(m.group(1))*10} {int(m.group(1))*10}mm diameter", text)
    # Also: dn{N} → emit dia{N} (for refs that say DN50, match queries that say diameter)
    text = re.sub(r"\bdn(\d{2,4})\b",
                  lambda m: f"dn{m.group(1)} dia{m.group(1)}", text)
    text = re.sub(r"\b(\d+)\s*mm2\b", r"\1mm2", text)
    text = re.sub(r"\b(\d+)\s*mm²\b", r"\1mm2", text)
    text = re.sub(r"\b(\d+)\s*sq\.?\s*mm\b", r"\1mm2", text)
    text = re.sub(r"\b(\d+)\s*sqmm\b", r"\1mm2", text)
    text = re.sub(r"\b(\d+)\s*mm\b", r"\1mm", text)
    text = re.sub(r"\b(\d+)\s*cm\b", r"\1cm", text)
    text = re.sub(r"\b(\d+)\s*mpa\b", r"\1mpa", text)
    text = re.sub(r"\b(\d+(?:\.\d+)?)\s*kv\b",
                  lambda m: f"{_canon_num(m.group(1))}kv", text)

    text = re.sub(r"\bexcavate\b", "excavation", text)
    text = re.sub(r"\bexcavated\b", "excavation", text)
    text = re.sub(r"\bexcavating\b", "excavation", text)
    text = re.sub(r"\bpad foundations?\b", "pad foundation footing", text)
    text = re.sub(r"\bcolumn bases?\b", "column base footing", text)
    text = re.sub(r"\bpipe dia\b", "pipe diameter", text)
    text = re.sub(r"\bsub-base\b", "subbase", text)
    text = re.sub(r"\bbase coarse\b", "base course", text)
    text = re.sub(r"\bconstruction joints?\b", "construction joint", text)
    text = re.sub(r"\bcontraction joints?\b", "contraction joint", text)
    text = re.sub(r"\bcurbstones?\b", "kerb", text)
    text = re.sub(r"\bcurbs?\b", "kerb", text)
    text = re.sub(r"\bditches\b", "ditch", text)
    text = re.sub(r"\bchannels\b", "channel", text)
    text = re.sub(r"\bblinding\b", "concrete blinding", text)

    # ── Fix: Joined dimension tokenization ─────────────────────────
    # BOQ says "1300 x 550 mm" → tokens: "1300", "550mm" (separated)
    # Reference says "1300x550mm" → token: "1300x550" (joined)
    # Generate both joined and separated forms so TF-IDF can match.
    # Pattern: W x H [mm] — generate WxH joined token
    def _emit_joined_dims(m):
        w, h = m.group(1), m.group(2)
        unit_suffix = m.group(3) or ""
        joined = f"{w}x{h}"
        return f"{m.group(0)} {joined}"
    text = re.sub(
        r"\b(\d{2,5})\s*x\s*(\d{2,5})(\s*mm)?\b",
        _emit_joined_dims, text
    )

    text = re.sub(r"[^a-z0-9;/\-\+\.\sx]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _simple_stem(tok: str) -> str:
    """Aggressive-enough suffix stripping for construction English.

    Handles: pipes→pipe, fittings→fitting, valves→valve,
    excavating→excavat (then synonym), surfaces→surface, etc.
    """
    if len(tok) <= 3:
        return tok
    # -ies → -y  (e.g. assemblies→assembly)
    if tok.endswith("ies") and len(tok) > 5:
        return tok[:-3] + "y"
    # -ness, -ment (keep root)
    for suffix in ("ness", "ment"):
        if tok.endswith(suffix) and len(tok) > len(suffix) + 3:
            return tok[: -len(suffix)]
    # -ing → strip (unless too short)
    if tok.endswith("ing") and len(tok) > 5:
        root = tok[:-3]
        # double consonant: e.g. "running" → "run"
        if len(root) >= 3 and root[-1] == root[-2]:
            root = root[:-1]
        return root
    # -tion → strip to root + "t"  (excavation → excavat)
    if tok.endswith("tion") and len(tok) > 6:
        return tok[:-3]  # keeps 't'
    # -es → strip only after sibilants (ch, sh, ss, x, z)
    if tok.endswith("es") and len(tok) > 4:
        root = tok[:-2]
        if root.endswith(("ch", "sh", "ss")) or root[-1] in ("x", "z"):
            return root
        # Not a sibilant – fall through to -s rule
    # -s  → strip (pipes→pipe, beams→beam, valves→valve, surfaces→surface)
    if tok.endswith("s") and len(tok) > 4:
        return tok[:-1]
    # -ed → strip (installed, painted)
    if tok.endswith("ed") and len(tok) > 4:
        root = tok[:-2]
        if len(root) >= 3 and root[-1] == root[-2]:
            root = root[:-1]
        return root
    return tok


# Rebuild synonym map to include stemmed forms so that post-stemming
# lookups still resolve correctly (e.g. "piping" → stem "pip" → map → "pipe").
for _group in SYNONYM_GROUPS:
    _canonical = _group[0]
    _canon_stemmed = _simple_stem(_canonical)
    for _term in _group:
        _stemmed = _simple_stem(_term)
        if _stemmed not in _SYNONYM_MAP:
            _SYNONYM_MAP[_stemmed] = _canonical
    # Ensure the stemmed canonical form maps back
    if _canon_stemmed not in _SYNONYM_MAP:
        _SYNONYM_MAP[_canon_stemmed] = _canonical


def tokenize_normalized(text: str) -> List[str]:
    """Tokenize pre-normalized text with stopword removal, stemming and synonym mapping."""
    if not text:
        return []
    raw = re.findall(
        r"dn\d+|\d+x\d+x\d+|\d+x\d+|\d+(?:\.\d+)?[a-z]+|[a-z]+[a-z0-9]*|\d+(?:\.\d+)?",
        text,
    )
    out: List[str] = []
    for tok in raw:
        if tok in STOPWORDS:
            continue
        if len(tok) == 1 and not tok.isdigit():
            continue
        tok = _simple_stem(tok)
        tok = _SYNONYM_MAP.get(tok, tok)
        out.append(tok)
    return out


def tokenize(text: str) -> List[str]:
    return tokenize_normalized(normalize_text(text))


def split_hierarchy(text: str) -> List[str]:
    text = clean_text(text)
    if not text:
        return []
    return [clean_text(x) for x in text.split(";") if clean_text(x)]


def is_generic_item(text: str) -> bool:
    toks = tokenize(text)
    if len(toks) <= 2:
        return True
    alpha = [t for t in toks if re.search(r"[a-z]", t)]
    numericish = [t for t in toks if re.search(r"\d", t)]
    if len(alpha) <= 2 and len(numericish) >= len(toks) / 2:
        return True
    return any(p.search(clean_text(text)) for p in GENERIC_ITEM_PATTERNS)


def infer_discipline_from_filename(path: str) -> str:
    """Infer discipline from a reference-file path.

    Uses the shared ``_DISC_INFER_RULES`` mapping so index-time and
    search-time inference stay consistent.
    """
    name = os.path.basename(path)
    return _infer_discipline_from_context(name)


def infer_discipline_from_query(*parts: str) -> Optional[str]:
    toks: Set[str] = set()
    for p in parts:
        toks.update(tokenize(p))
    scores = {d: len(toks & hints) for d, hints in DISCIPLINE_HINTS.items()}
    best = max(scores, key=scores.get)  # type: ignore[arg-type]
    return best if scores[best] > 0 else None


# ═════════════════════════════════════════════════════════════════════════
# Spec Extraction
# ═════════════════════════════════════════════════════════════════════════

def _canon_dim(parts: Sequence[str]) -> str:
    nums = [str(int(p)) for p in parts if clean_text(p)]
    if len(nums) == 2:
        nums = [str(x) for x in sorted((int(nums[0]), int(nums[1])))]
    return "x".join(nums)


# ── Categorical spec vocabularies ────────────────────────────────────────
# These map normalised keywords found in BOQ or price-code descriptions
# to canonical tags used for matching.  Order matters: longer patterns
# must come first so greedy matching picks the most specific.

_PIPE_MAT_PATTERNS: List[Tuple[str, str]] = [
    # Plastics
    (r"\bhdpe\b", "HDPE"), (r"\bpe\s*-?\s*x\b|\bpex\b", "PEX"),
    (r"\bpp\s*-?\s*rct\b", "PPRCT"), (r"\bppr\b|\bpp\s*-?\s*r\b", "PPR"),
    (r"\bupvc\b|\bu\s*-?pvc\b", "UPVC"), (r"\bpvc\b", "PVC"),
    (r"\bgrp\b|\bfrp\b", "GRP"), (r"\bcpvc\b", "CPVC"),
    (r"\babs\b", "ABS"),
    # Metals
    (r"\bdi\b|\bductile\s*iron\b", "DI"), (r"\bci\b|\bcast\s*iron\b", "CI"),
    (r"\bgalvan[iz]+ed\s*steel\b|\bgs\b|\bgalv\.?\s*steel\b", "GS"),
    (r"\bblack\s*steel\b|\bbs\b(?=\s+pipe|\s+bend|\s+tee)", "BS"),
    (r"\bcarbon\s*steel\b|\bcs\b(?=\s+pipe|\s+bend|\s+tee|\s+\w*valve)", "CS"),
    (r"\bstainless\s*steel\b|\bss\b(?=\s+pipe)", "SS"),
    (r"\bcopper\b(?=\s+(?:pipe|tube|tubing))|\bcu\b(?=\s+pipe)", "COPPER"),
    # Concrete pipes – require pipe/culvert context to avoid matching
    # generic "reinforced concrete" in building BOQs.
    (r"\brc\b(?=\s+pipe)|\breinforced\s+concrete\s+pipe", "RC"),
    (r"\bvc\b(?=\s+pipe)|\bvitrified\s*clay\b", "VC"),
]
_PIPE_MAT_RE = [(re.compile(p, re.I), tag) for p, tag in _PIPE_MAT_PATTERNS]

_VALVE_TYPE_PATTERNS: List[Tuple[str, str]] = [
    (r"\bgate\s*valve\b", "GATE"), (r"\bcheck\s*valve\b|\bnon[- ]return\s*valve\b", "CHECK"),
    (r"\bbutterfly\s*valve\b", "BUTTERFLY"), (r"\bball\s*valve\b", "BALL"),
    (r"\bglobe\s*valve\b", "GLOBE"),
    (r"\bwashout\s*valve\b|\bwash[- ]?out\b", "WASHOUT"),
    (r"\bair\s*(?:release\s*)?valve\b|\bair\s*valve\b", "AIRVALVE"),
    (r"\bpressure\s*reduc\w*\s*valve\b|\bprv\b", "PRV"),
    (r"\bsolenoid\s*valve\b", "SOLENOID"),
    (r"\bautomatic\s*vent\w*\s*valve\b", "AUTOVENT"),
    (r"\brelief\s*valve\b|\bsafety\s*valve\b", "RELIEF"),
    (r"\bstrainer\b", "STRAINER"),
]
_VALVE_TYPE_RE = [(re.compile(p, re.I), tag) for p, tag in _VALVE_TYPE_PATTERNS]

_FITTING_TYPE_PATTERNS: List[Tuple[str, str]] = [
    (r"\bbends?\b|\belbows?\b", "BEND"), (r"\btees?\b", "TEE"),
    (r"\breducers?\b|\breducing\b", "REDUCER"), (r"\bcouplings?\b", "COUPLING"),
    (r"\bflanges?\b", "FLANGE"), (r"\bunions?\b", "UNION"),
    (r"\badaptors?\b|\badapters?\b", "ADAPTOR"),
]
_FITTING_TYPE_RE = [(re.compile(p, re.I), tag) for p, tag in _FITTING_TYPE_PATTERNS]

_CONCRETE_ELEM_PATTERNS: List[Tuple[str, str]] = [
    (r"\bisolated\s*foot\w*\b", "ISOFOOT"), (r"\bstrip\s*foot\w*\b", "STRIPFOOT"),
    (r"\brafts?\b", "RAFT"), (r"\bpile\s*caps?\b", "PILECAP"), (r"\bpiles?\b", "PILE"),
    (r"\btie\s*beams?\b", "TIEBEAM"), (r"\bdrop\s*(?:beams?|panels?)\b", "DROPBEAM"),
    (r"\bground\s*beams?\b|\bgrade\s*beams?\b", "GRADEBEAM"),
    (r"\bshear\s*walls?\b", "SHEARWALL"), (r"\bretaining\s*walls?\b", "RETWALL"),
    (r"\bslabs?\s*on\s*grade\b|\bsog\b", "SOG"),
    (r"\bsuspended\s*slabs?\b", "SUSPSLAB"), (r"\bflat\s*slabs?\b", "FLATSLAB"),
    (r"\bslabs?\b", "SLAB"), (r"\bcolumns?\s*necks?\b", "COLNECK"),
    (r"\bcolumns?\b", "COLUMN"), (r"\bbeams?\b", "BEAM"),
    (r"\bstairs?\b|\bsteps?\b", "STAIR"), (r"\bparapets?\b", "PARAPET"),
    (r"\bkerbs?\b|\bcurbs?\b", "KERB"), (r"\bblinding\b", "BLINDING"),
    (r"\bpedestals?\b", "PEDESTAL"), (r"\bramps?\b", "RAMP"),
    (r"\bupstand\s*walls?\b", "UPSTAND"), (r"\bbasement\s*walls?\b", "BSMTWALL"),
    (r"\bneck\s*(?:columns?|walls?)\b", "COLNECK"),
    (r"\bwalls?\b", "WALL"), (r"\bfoot\w*\b", "FOOTING"),
    (r"\btransfer\s*beams?\b", "TRANSBEAM"),
]
_CONCRETE_ELEM_RE = [(re.compile(p, re.I), tag) for p, tag in _CONCRETE_ELEM_PATTERNS]

_CONCRETE_SCOPE_PATTERNS: List[Tuple[str, str]] = [
    (r"\bformwork\s*(?:and|&|\+)\s*reinfo?rce?ment\b|\breinfo?rce?ment\s*(?:and|&|\+)\s*formwork\b", "FORM+REBAR"),
    (r"\bconcrete\s*only\b", "CONC_ONLY"),
    (r"\breinfo?rce?ment\s*only\b|\brebar\s*only\b", "REBAR_ONLY"),
    (r"\bformwork\s*only\b|\bshutter\w*\s*only\b", "FORM_ONLY"),
    (r"\bwith\s+reinfo?rce?ment\b|\bincl\w*\s*rebar\b", "WITH_REBAR"),
    (r"\bwith\s+formwork\b|\bincl\w*\s*formwork\b", "WITH_FORM"),
]
_CONCRETE_SCOPE_RE = [(re.compile(p, re.I), tag) for p, tag in _CONCRETE_SCOPE_PATTERNS]

_STEEL_SECTION_PATTERNS: List[Tuple[str, str]] = [
    (r"\bi\s*/\s*h\b|\bih\b|\bi[- ]?section\b|\bh[- ]?section\b|\brolled\s*section\b", "IH"),
    (r"\brhs\b|\brectangular\s*hollow\b", "RHS"),
    (r"\bshs\b|\bsquare\s*hollow\b", "SHS"),
    (r"\bchs\b|\bcircular\s*hollow\b", "CHS"),
    (r"\bangle\b(?!.*\bdeg)", "ANGLE"), (r"\bflat\s*bar\b", "FLATBAR"),
    (r"\bplate\b", "PLATE"), (r"\bchannel\b", "CHANNEL"),
]
_STEEL_SECTION_RE = [(re.compile(p, re.I), tag) for p, tag in _STEEL_SECTION_PATTERNS]

_COATING_PATTERNS: List[Tuple[str, str]] = [
    (r"\bhot[- ]?dip\s*galv\w*\b|\bhdg\b", "GALV"),
    (r"\bgalvan\w*\b", "GALV"),
    (r"\bintumescent\b", "INTUMESCENT"), (r"\bcementitious\s*fire\w*\b", "CEMENTITIOUS"),
    (r"\bshop\s*paint\w*\b", "SHOPPAINT"),
    (r"\bepoxy\s*coat\w*\b|\bepoxy[- ]coated\b", "EPOXYCOAT"),
    (r"\baess\b|\barchitecturally\s*exposed\b", "AESS"),
    (r"\bpowder\s*coat\w*\b", "POWDERCOAT"),
]
_COATING_RE = [(re.compile(p, re.I), tag) for p, tag in _COATING_PATTERNS]

_INSUL_MAT_PATTERNS: List[Tuple[str, str]] = [
    (r"\belastomeric\b|\barnaflex\b|\bkaiflex\b", "ELASTOMERIC"),
    (r"\bfiberg?lass\b|\bglass\s*wool\b", "FIBERGLASS"),
    (r"\bmineral\s*wool\b|\brock\s*wool\b", "MINERALWOOL"),
    (r"\bpir\b|\bpolyisocyanurate\b", "PIR"),
    (r"\bpuf\b|\bpolyurethane\s*foam\b|\bpu\s*foam\b", "PUF"),
    (r"\bxps\b|\bextruded\s*polystyrene\b", "XPS"),
    (r"\beps\b|\bexpanded\s*polystyrene\b", "EPS"),
    (r"\bphenolic\b", "PHENOLIC"),
]
_INSUL_MAT_RE = [(re.compile(p, re.I), tag) for p, tag in _INSUL_MAT_PATTERNS]

_CABLE_INSUL_PATTERNS: List[Tuple[str, str]] = [
    (r"\bfr[- ]?xlpe\b", "FR-XLPE"), (r"\bxlpe\b", "XLPE"),
    (r"\bepr\b", "EPR"), (r"\blszh\b|\bls0h\b|\blsoh\b", "LSZH"),
    (r"\bswa\b", "SWA"), (r"\bawa\b", "AWA"),
    (r"\bpvc\b(?=.*cable|.*\bconductor|.*\bwire)", "PVC"),
    (r"\bmica\b", "MICA"),
]
_CABLE_INSUL_RE = [(re.compile(p, re.I), tag) for p, tag in _CABLE_INSUL_PATTERNS]


def _extract_categorical(text_lower: str) -> Dict[str, Tuple[str, ...]]:
    """Extract categorical (non-numeric) specs from lowered text."""
    def _match_all(patterns, txt):
        found = []
        for rgx, tag in patterns:
            if rgx.search(txt):
                found.append(tag)
        return tuple(sorted(set(found)))

    result: Dict[str, Tuple[str, ...]] = {}
    result["pipe_mat"]       = _match_all(_PIPE_MAT_RE, text_lower)
    result["valve_type"]     = _match_all(_VALVE_TYPE_RE, text_lower)
    result["fitting_type"]   = _match_all(_FITTING_TYPE_RE, text_lower)
    result["concrete_elem"]  = _match_all(_CONCRETE_ELEM_RE, text_lower)
    result["concrete_scope"] = _match_all(_CONCRETE_SCOPE_RE, text_lower)
    result["steel_section"]  = _match_all(_STEEL_SECTION_RE, text_lower)
    result["coating"]        = _match_all(_COATING_RE, text_lower)
    result["insul_mat"]      = _match_all(_INSUL_MAT_RE, text_lower)
    result["cable_insul"]    = _match_all(_CABLE_INSUL_RE, text_lower)

    # Schedule (SCH 40, SCH 80, SDR values)
    sch_vals: Set[str] = set()
    for m in re.finditer(r"\bsch(?:edule)?[. ]*?(\d{2,3})\b", text_lower):
        sch_vals.add(f"SCH{m.group(1)}")
    for m in re.finditer(r"\bsdr\s*[- ]?\s*(\d{1,2}(?:\.\d+)?)\b", text_lower):
        sch_vals.add(f"SDR{m.group(1)}")
    result["schedule"] = tuple(sorted(sch_vals))

    # Pipe grade (PE 80, PE 100, Class C/III/IV/V, K7/K9, SS/ES/EES, ERW/Seamless)
    grade_vals: Set[str] = set()
    for m in re.finditer(r"\bpe\s*[-]?\s*(80|100)\b", text_lower):
        grade_vals.add(f"PE{m.group(1)}")
    for m in re.finditer(r"\bcl(?:ass)?\s+([ivxlc]+|\d{1,3})\b", text_lower):
        grade_vals.add(f"CL{m.group(1).upper()}")
    for m in re.finditer(r"\bk\s*[-]?\s*(\d{1,2})\b", text_lower):
        grade_vals.add(f"K{m.group(1)}")
    for m in re.finditer(r"\b(erw|seamless)\b", text_lower):
        grade_vals.add(m.group(1).upper())
    result["pipe_grade"] = tuple(sorted(grade_vals))

    # Stiffness class (SN2500, SN5000, SN10000)
    sn_vals: Set[str] = set()
    for m in re.finditer(r"\bsn\s*[-]?\s*(\d{3,5})\b", text_lower):
        sn_vals.add(f"SN{m.group(1)}")
    result["stiffness"] = tuple(sorted(sn_vals))

    # Bend angle
    angle_vals: Set[str] = set()
    for m in re.finditer(r"\b(11|22|30|45|60|90)\s*(?:deg|°|degree)\b", text_lower):
        angle_vals.add(m.group(1))
    # Also match "11 deg" pattern in price-code descriptions
    for m in re.finditer(r"\b(11|22|30|45|60|90)\s+deg\b", text_lower):
        angle_vals.add(m.group(1))
    result["bend_angle"] = tuple(sorted(angle_vals))

    # STC sound rating
    stc_vals: Set[str] = set()
    for m in re.finditer(r"\bstc\s*[-]?\s*(\d{2,3})\s*(?:db)?\b", text_lower):
        stc_vals.add(m.group(1))
    result["stc"] = tuple(sorted(stc_vals))

    return result


def extract_specs(text: str) -> Dict[str, Tuple[str, ...]]:
    raw = clean_text(text)
    empty: Dict[str, Tuple[str, ...]] = {
        "dn": (), "dia": (), "mm": (), "mpa": (), "kv": (),
        "dims": (), "mm2": (), "cores": (),
        "thk": (), "fire": (), "pn": (),
        "pipe_mat": (), "valve_type": (), "fitting_type": (),
        "concrete_elem": (), "concrete_scope": (),
        "steel_section": (), "coating": (), "insul_mat": (), "cable_insul": (),
        "schedule": (), "pipe_grade": (), "stiffness": (), "bend_angle": (),
        "stc": (),
    }
    if not raw:
        return empty

    raw_x = raw.replace("×", "x")
    low = raw_x.lower()
    n = normalize_text(raw)

    pipe_ctx = bool(re.search(
        r"\b(rcp|pipe|culvert|storm|stormwater|drain|drainage|sewer|sewage|sanitary|potable|irrigation|manhole)\b", low))
    cable_ctx = bool(
        re.search(r"\b(cable|xlpe|swa|awa|lszh|mv|lv|hv|core|mm2|mm²|sqmm|sq\.?\s*mm)\b", low)
        or re.search(r"\b\d{1,2}c\b|\b\d+mm2\b", n)
    )
    concrete_ctx = bool(re.search(
        r"\b(concrete|blinding|footing|foundation|slab|column|pedestal|grout|mortar|plaster|shotcrete|screed|raft|beam|stair|wall|formwork|shutter|rebar|reinforcement|in[- ]?situ)\b", low))
    steel_ctx = bool(re.search(r"\b(rebar|reinforcement|steel|bar\b|yield|fy)\b", low))

    # DN values
    dn_vals: Set[str] = set(re.findall(r"\bdn(\d{2,4})\b", n))
    for m in re.finditer(r"\brcp\s*d\s*(\d{2,4})\b", low, flags=re.I):
        dn_vals.add(str(int(m.group(1))))
    for m in re.finditer(
        r"\b(?:pipe|culvert|rcp|storm|drain|sewer|sanitary)\b[^\n,;]{0,24}\bd\s*[-:/]?\s*(\d{2,4})\b", low, flags=re.I
    ):
        dn_vals.add(str(int(m.group(1))))
    for m in re.finditer(
        r"\bd\s*[-:/]?\s*(\d{2,4})\b[^\n,;]{0,24}\b(?:pipe|culvert|rcp|storm|drain|sewer|sanitary)\b", low, flags=re.I
    ):
        dn_vals.add(str(int(m.group(1))))

    # Diameter values
    dia_vals: Set[str] = set(re.findall(r"\bdia(\d{2,4})\b", n))
    for m in re.finditer(r"[Øø]\s*(\d{2,4})\b", raw_x):
        v = str(int(m.group(1)))
        dia_vals.add(v)
        if pipe_ctx:
            dn_vals.add(v)
    for m in re.finditer(r"\b(?:dia|diameter)\s*[:\-]?\s*(\d{2,4})\s*(mm|cm)?\b", raw_x, flags=re.I):
        v = int(m.group(1)) * (10 if (m.group(2) or "").lower() == "cm" else 1)
        dia_vals.add(str(v))
        if pipe_ctx:
            dn_vals.add(str(v))
    for m in re.finditer(r"\b(\d{2,4})\s*(mm|cm)?\s*(?:dia|diameter)\b", raw_x, flags=re.I):
        v = int(m.group(1)) * (10 if (m.group(2) or "").lower() == "cm" else 1)
        dia_vals.add(str(v))
        if pipe_ctx:
            dn_vals.add(str(v))

    mm = set(re.findall(r"\b(\d+)mm\b", n))

    # MPa
    mpa: Set[str] = set()
    for m in re.finditer(r"\b(\d+(?:\.\d+)?)mpa\b", n):
        val = _canon_num(m.group(1))
        try:
            f = float(val)
        except ValueError:
            continue
        if steel_ctx and f >= 100:
            continue
        if f < 100 and (concrete_ctx or not steel_ctx):
            mpa.add(val)

    # Concrete grade designation: "C32/40", "C12/15", "C40", "C 40", "Grade C40"
    # The second number in C32/40 is the cylinder strength in MPa.
    # Single-number forms like "C40" also indicate MPa directly.
    for m in re.finditer(r"\bc\s*(\d{1,2})\s*/\s*(\d{1,2})\b", low):
        cyl = int(m.group(2))
        if 10 <= cyl <= 80:
            mpa.add(str(cyl))
    for m in re.finditer(r"(?:\bgrade\s+)?\bc\s*-?\s*(\d{2})\b(?!\s*/\s*\d)", low):
        val = int(m.group(1))
        if 10 <= val <= 80 and concrete_ctx:
            mpa.add(str(val))

    # kV
    kv: Set[str] = set(_canon_num(x) for x in re.findall(r"\b(\d+(?:\.\d+)?)kv\b", n))
    for a, b in re.findall(r"\b(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)\s*kv\b", low, flags=re.I):
        kv.add(_canon_num(a))
        kv.add(_canon_num(b))

    # Dimensions
    dims: Set[str] = set()
    for m in re.finditer(r"\b(\d{2,5})\s*x\s*(\d{2,5})(?:\s*x\s*(\d{2,5}))?\b", low):
        tail = low[m.end(): m.end() + 10]
        if re.search(r"\bmm2\b|\bmm²\b|\bsqmm\b|\bsq\.?\s*mm\b", tail):
            continue
        dims.add(_canon_dim([g for g in m.groups() if g]))

    # Cable cross-section mm²
    mm2_vals: Set[str] = set()
    for raw_size in re.findall(r"\b(\d+(?:/\d+)?)\s*(?:mm2|mm²|sqmm|sq\.?\s*mm|mm\^2)\b", low, flags=re.I):
        for part in raw_size.split("/"):
            mm2_vals.add(str(int(part)))
    for raw_size in re.findall(r"\b\d+\s*c(?:ore)?\s*[x*]?\s*(\d+(?:/\d+)?)\s*(?:mm2|mm²|sqmm|sq\.?\s*mm|mm\^2)\b", low, flags=re.I):
        for part in raw_size.split("/"):
            mm2_vals.add(str(int(part)))
    for raw_size in re.findall(r"\b\d+c\s*(\d+(?:/\d+)?)mm2\b", n, flags=re.I):
        for part in raw_size.split("/"):
            mm2_vals.add(str(int(part)))

    # Core count
    core_vals: Set[str] = set()
    if cable_ctx:
        for c in re.findall(r"\b(\d{1,2})\s*core\b", low, flags=re.I):
            core_vals.add(str(int(c)))
        for c in re.findall(r"\b(\d{1,2})c\b", n, flags=re.I):
            core_vals.add(str(int(c)))
        for c in re.findall(r"\b(\d{1,2})\s*[x*]\s*\d+(?:/\d+)?\s*(?:mm2|mm²|sqmm|sq\.?\s*mm|mm\^2)\b", low, flags=re.I):
            core_vals.add(str(int(c)))

    # ── Thickness (masonry, insulation, cladding) ────────────────────
    thk_vals: Set[str] = set()
    # "Th=100mm", "Th.=200mm", "thickness 150 mm", "150mm thick"
    for m in re.finditer(r"\bth(?:ickness)?\s*[=.:]*\s*(\d{2,4})\s*mm\b", low):
        thk_vals.add(str(int(m.group(1))))
    for m in re.finditer(r"\b(\d{2,4})\s*mm\s*(?:thick|thk)\b", low):
        thk_vals.add(str(int(m.group(1))))

    # ── Fire rating (minutes) ────────────────────────────────────────
    fire_vals: Set[str] = set()
    # "fire rated 60 min", "FRL 120/120/120", "60min fire", "FR 90"
    for m in re.finditer(r"\bfire\s*(?:rated?|rating)?\s*(\d{2,3})\s*min", low):
        fire_vals.add(str(int(m.group(1))))
    for m in re.finditer(r"\b(\d{2,3})\s*min(?:ute)?s?\s*(?:fire|frl|fr)\b", low):
        fire_vals.add(str(int(m.group(1))))
    for m in re.finditer(r"\bfrl\s*(\d{2,3})(?:/\d+)*\b", low):
        fire_vals.add(str(int(m.group(1))))
    for m in re.finditer(r"\bfr\s*[-:]?\s*(\d{2,3})\b", low):
        fire_vals.add(str(int(m.group(1))))

    # ── Pressure class / PN rating ───────────────────────────────────
    pn_vals: Set[str] = set()
    # "PN10", "PN 16", "Class C", "Class 150", "K9"
    for m in re.finditer(r"\bpn\s*[-:]?\s*(\d{1,3})\b", low):
        pn_vals.add(f"PN{m.group(1)}")
    for m in re.finditer(r"\bclass\s+(\w{1,4})\b", low):
        pn_vals.add(f"CL{m.group(1).upper()}")
    for m in re.finditer(r"\bk\s*[-:]?\s*(\d{1,2})\b", low):
        # K-class for ductile iron (K7, K9, K12)
        if 5 <= int(m.group(1)) <= 14:
            pn_vals.add(f"K{m.group(1)}")

    # ── Categorical specs ───────────────────────────────────────
    cat = _extract_categorical(low)

    return {
        "dn":    tuple(sorted(dn_vals,   key=lambda x: (len(x), x))),
        "dia":   tuple(sorted(dia_vals,  key=lambda x: (len(x), x))),
        "mm":    tuple(sorted(mm,        key=lambda x: (len(x), x))),
        "mpa":   tuple(sorted(mpa,       key=lambda x: (float(x), x))),
        "kv":    tuple(sorted(kv,        key=lambda x: float(x))),
        "dims":  tuple(sorted(dims)),
        "mm2":   tuple(sorted(mm2_vals,  key=lambda x: (int(x), x))),
        "cores": tuple(sorted(core_vals, key=lambda x: (int(x), x))),
        "thk":   tuple(sorted(thk_vals,  key=lambda x: (int(x), x))),
        "fire":  tuple(sorted(fire_vals, key=lambda x: (int(x), x))),
        "pn":    tuple(sorted(pn_vals)),
        **cat,
    }


# ═════════════════════════════════════════════════════════════════════════
# Excel Safety  (handles malformed AutoFilter metadata)
# ═════════════════════════════════════════════════════════════════════════

_VALID_CUSTOM_FILTER_OPERATORS = {
    "equal", "greaterThan", "lessThanOrEqual", "notEqual",
    "lessThan", "greaterThanOrEqual",
}
_CUSTOM_FILTER_OPERATOR_RE = re.compile(
    rb'(<customFilter\b[^>]*?)\soperator="([^"]+)"', re.IGNORECASE
)
_AUTOFILTER_BLOCK_RE = re.compile(
    rb'<autoFilter\b[^>]*(?:/>|>.*?</autoFilter>)', re.IGNORECASE | re.DOTALL
)


def _sanitize_filter_xml(xml_bytes: bytes, drop_auto_filters: bool = False) -> bytes:
    def _fix(match: re.Match[bytes]) -> bytes:
        operator = match.group(2).decode("utf-8", "ignore")
        if operator in _VALID_CUSTOM_FILTER_OPERATORS:
            return match.group(0)
        return match.group(1)

    xml_bytes = _CUSTOM_FILTER_OPERATOR_RE.sub(_fix, xml_bytes)
    if drop_auto_filters:
        xml_bytes = _AUTOFILTER_BLOCK_RE.sub(b"", xml_bytes)
    return xml_bytes


def _xlsx_has_invalid_filter(xlsx_path: str) -> bool:
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zin:
            for info in zin.infolist():
                if not info.filename.endswith(".xml"):
                    continue
                data = zin.read(info.filename)
                for m in _CUSTOM_FILTER_OPERATOR_RE.finditer(data):
                    op = m.group(2).decode("utf-8", "ignore")
                    if op not in _VALID_CUSTOM_FILTER_OPERATORS:
                        return True
    except Exception:
        return False
    return False


def _sanitize_xlsx_bytes(xlsx_path: str, drop_auto_filters: bool = False) -> bytes:
    out = io.BytesIO()
    with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename.endswith(".xml"):
                data = _sanitize_filter_xml(data, drop_auto_filters=drop_auto_filters)
            zout.writestr(info, data)
    return out.getvalue()


def safe_load_workbook(path: str, **kwargs):
    """Load an openpyxl workbook, repairing broken AutoFilter metadata if needed."""
    if _xlsx_has_invalid_filter(path):
        logger.warning(f"Repairing invalid filter metadata in {os.path.basename(path)}")
        try:
            repaired = _sanitize_xlsx_bytes(path, drop_auto_filters=False)
            bio = io.BytesIO(repaired)
            wb = load_workbook(bio, **kwargs)
            wb._source_stream = bio  # type: ignore[attr-defined]
            return wb
        except ValueError:
            repaired = _sanitize_xlsx_bytes(path, drop_auto_filters=True)
            bio = io.BytesIO(repaired)
            wb = load_workbook(bio, **kwargs)
            wb._source_stream = bio  # type: ignore[attr-defined]
            return wb
    try:
        return load_workbook(path, **kwargs)
    except ValueError as exc:
        if "Value must be one of" not in str(exc):
            raise
        repaired = _sanitize_xlsx_bytes(path, drop_auto_filters=True)
        bio = io.BytesIO(repaired)
        wb = load_workbook(bio, **kwargs)
        wb._source_stream = bio  # type: ignore[attr-defined]
        return wb


# ═════════════════════════════════════════════════════════════════════════
# Reference File Parsing
# ═════════════════════════════════════════════════════════════════════════

_REF_COL_ALIASES: Dict[str, List[str]] = {
    "price_code": ["price code", "price_code", "code"],
    "price_code_description": ["price code description", "price_code_description"],
}


def _normalized_aliases(key: str) -> List[str]:
    return [normalize_text(x) for x in _REF_COL_ALIASES[key]]


def _score_header_values(values: Sequence[object], alias_keys: Sequence[str]) -> int:
    norm_cells = [normalize_text(v) for v in values]
    score = 0
    for key in alias_keys:
        aliases = _normalized_aliases(key)
        if any(cell in aliases for cell in norm_cells):
            score += 2
        elif any(cell and any(a in cell for a in aliases) for cell in norm_cells):
            score += 1
    return score


def _choose_header_row(buffer: List[Tuple[int, Tuple[object, ...]]], alias_keys: Sequence[str]) -> int:
    best_row, best_score = buffer[0][0], -1
    for row_num, vals in buffer:
        s = _score_header_values(vals, alias_keys)
        if s > best_score:
            best_score = s
            best_row = row_num
    return best_row


def _build_header_map(values: Sequence[object]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for idx, val in enumerate(values, start=1):
        txt = clean_text(val)
        out[idx] = txt if txt else f"Unnamed: {idx}"
    return out


def _find_col_exact(header_map: Dict[int, str], alias_key: str) -> Optional[int]:
    aliases = set(_normalized_aliases(alias_key))
    for idx, name in header_map.items():
        if normalize_text(name) in aliases:
            return idx
    return None


def _find_col_contains(header_map: Dict[int, str], alias_key: str) -> Optional[int]:
    aliases = _normalized_aliases(alias_key)
    for alias in aliases:
        for idx, name in header_map.items():
            nm = normalize_text(name)
            if nm and alias in nm:
                return idx
    return None


def _cell_value(row_vals: Sequence[object], col: Optional[int]) -> str:
    if not col:
        return ""
    if col - 1 >= len(row_vals):
        return ""
    return clean_text(row_vals[col - 1])


def _is_bad_ref_header_row(price_code: str, desc: str) -> bool:
    pc = normalize_text(price_code)
    d = normalize_text(desc)
    if not d:
        return True
    if pc == "price code":
        return True
    if d in {"price code description", "section description", "family description"}:
        return True
    header_phrases = [
        "section description", "family description",
        "variable 1 description", "variable 2 description",
        "variable 3 description", "price code description",
    ]
    if sum(1 for p in header_phrases if p in d) >= 2:
        return True
    if d.startswith("section description ; family description"):
        return True
    return False


def iter_ref_rows(path: str) -> Iterator[
    Tuple[str, str, str, str, str, str, Dict[str, Tuple[str, ...]], List[str]]
]:
    """
    Yield (discipline, source_file, sheet, price_code, prefixed_desc,
           leaf_desc, specs_dict, token_list) for each valid row.
    """
    discipline = infer_discipline_from_filename(path)
    raw_stem = os.path.splitext(os.path.basename(path))[0]
    # Strip the 'ref_' prefix added during S3 download so the stored
    # source_file matches the original S3 key stem used by ALLOCATE filters.
    source_file = raw_stem[4:] if raw_stem.startswith("ref_") else raw_stem
    wb = safe_load_workbook(path, read_only=True, data_only=True)

    # Regex to strip electrical variant letter annotations like "(A)",
    # "(B)", "(a)" from descriptions.  These are redundant with the
    # suffix letters in the price code and clutter LLM input.
    _ELEC_ANNOTATION_RE = re.compile(r"\s*\([A-Za-z]\)")

    for ws in wb.worksheets:
        row_iter = ws.iter_rows(values_only=True)
        buffer: List[Tuple[int, Tuple[object, ...]]] = []
        for idx, vals in enumerate(row_iter, start=1):
            buffer.append((idx, vals))
            if len(buffer) >= min(10, ws.max_row or 10):
                break
        if not buffer:
            continue

        header_row_num = _choose_header_row(buffer, ["price_code", "price_code_description"])
        header_vals = next(vals for rn, vals in buffer if rn == header_row_num)
        header_map = _build_header_map(header_vals)

        pc_col = _find_col_exact(header_map, "price_code") or _find_col_contains(header_map, "price_code")
        desc_col = _find_col_exact(header_map, "price_code_description") or _find_col_contains(header_map, "price_code_description")
        if not desc_col:
            continue

        extra_desc_cols: List[int] = []
        for idx, name in header_map.items():
            nm = normalize_text(name)
            if idx == desc_col:
                continue
            if "description" in nm and "price code description" not in nm:
                extra_desc_cols.append(idx)

        def _process(vals: Sequence[object]):
            desc = _cell_value(vals, desc_col)
            price_code = _cell_value(vals, pc_col)
            if _is_bad_ref_header_row(price_code, desc):
                return None
            prefix_parts = [_cell_value(vals, c) for c in extra_desc_cols[:4]]
            prefix_parts = [x for x in prefix_parts if x]
            prefixed = " ; ".join(prefix_parts + [desc]) if prefix_parts else desc
            if _is_bad_ref_header_row(price_code, prefixed):
                return None
            # Strip electrical variant annotations like "(A)", "(B)"
            # which are redundant with the price code suffix letters.
            if discipline == "E_Electrical":
                prefixed = _ELEC_ANNOTATION_RE.sub("", prefixed)
            segs = split_hierarchy(prefixed)
            leaf = segs[-1] if segs else prefixed
            specs = extract_specs(prefixed)
            toks = sorted(set(tokenize(prefixed)))
            if not toks:
                return None
            return discipline, source_file, ws.title, price_code, prefixed, leaf, specs, toks

        for rn, vals in buffer:
            if rn <= header_row_num:
                continue
            out = _process(vals)
            if out:
                yield out

        consumed = buffer[-1][0]
        for _idx, vals in enumerate(row_iter, start=consumed + 1):
            out = _process(vals)
            if out:
                yield out


# ═════════════════════════════════════════════════════════════════════════
# SQLite Index
# ═════════════════════════════════════════════════════════════════════════

_SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS refs (
    ref_id        INTEGER PRIMARY KEY,
    discipline    TEXT,
    source_file   TEXT,
    sheet_name    TEXT,
    price_code    TEXT,
    prefixed_description TEXT,
    leaf_description     TEXT,
    family_key    TEXT,
    norm_text     TEXT,
    norm_leaf     TEXT,
    dn_csv  TEXT, dia_csv  TEXT, mm_csv  TEXT, mpa_csv TEXT,
    kv_csv  TEXT, dims_csv TEXT, mm2_csv TEXT, core_csv TEXT,
    thk_csv TEXT, fire_csv TEXT, pn_csv TEXT,
    pipe_mat_csv TEXT, valve_type_csv TEXT, concrete_elem_csv TEXT,
    cable_insul_csv TEXT, schedule_csv TEXT, pipe_grade_csv TEXT,
    fitting_type_csv TEXT,
    code_prefix TEXT
);
CREATE TABLE IF NOT EXISTS postings (
    token  TEXT,
    ref_id INTEGER
);
CREATE INDEX IF NOT EXISTS idx_postings_token     ON postings(token);
CREATE INDEX IF NOT EXISTS idx_postings_token_ref ON postings(token, ref_id);
CREATE INDEX IF NOT EXISTS idx_refs_source_file   ON refs(source_file);
CREATE INDEX IF NOT EXISTS idx_refs_code_prefix   ON refs(code_prefix);
CREATE TABLE IF NOT EXISTS df (
    token TEXT PRIMARY KEY,
    df    INTEGER
);
CREATE TABLE IF NOT EXISTS meta (
    key   TEXT PRIMARY KEY,
    value TEXT
);
CREATE TABLE IF NOT EXISTS indexed_files (
    source_file  TEXT PRIMARY KEY,
    file_path    TEXT
);
CREATE TABLE IF NOT EXISTS sheet_tokens (
    sheet_name TEXT,
    token      TEXT,
    score      REAL
);
CREATE INDEX IF NOT EXISTS idx_sheet_tokens_sheet ON sheet_tokens(sheet_name);
"""


def _reset_index(conn: sqlite3.Connection) -> None:
    conn.executescript(
        "DROP TABLE IF EXISTS refs;"
        "DROP TABLE IF EXISTS postings;"
        "DROP TABLE IF EXISTS df;"
        "DROP TABLE IF EXISTS meta;"
        "DROP TABLE IF EXISTS indexed_files;"
        "DROP TABLE IF EXISTS sheet_tokens;"
    )
    conn.executescript(_SCHEMA_SQL)
    conn.commit()


def _get_meta(conn: sqlite3.Connection, key: str) -> Optional[str]:
    row = conn.execute("SELECT value FROM meta WHERE key = ?", (key,)).fetchone()
    return str(row[0]) if row else None


def _ref_signature(ref_paths: Sequence[str]) -> str:
    parts: List[str] = []
    for p in sorted(str(x) for x in ref_paths):
        try:
            st = os.stat(p)
            parts.append(f"{os.path.abspath(p)}|{st.st_size}|{int(st.st_mtime)}")
        except OSError:
            parts.append(f"{os.path.abspath(p)}|missing")
    return hashlib.sha256("\n".join(parts).encode()).hexdigest()


def _index_ready(conn: sqlite3.Connection, ref_paths: Sequence[str]) -> bool:
    cur = conn.cursor()
    required = {"refs", "postings", "df", "meta"}
    cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
    have = {str(r[0]) for r in cur.fetchall()}
    if not required.issubset(have):
        return False
    if _get_meta(conn, "schema_version") != SCHEMA_VERSION:
        return False
    if _get_meta(conn, "build_complete") != "1":
        return False
    if _get_meta(conn, "ref_signature") != _ref_signature(ref_paths):
        return False
    for table in ("refs", "postings", "df"):
        cur.execute(f"SELECT COUNT(*) FROM {table}")
        if int((cur.fetchone() or [0])[0]) <= 0:
            return False
    return True


def build_index(
    db_path: str,
    ref_paths: Sequence[str],
    rebuild: bool = False,
) -> str:
    """
    Build (or reuse) a SQLite lexical index from reference Excel files.

    When *rebuild* is ``False`` and an existing index is found, only reference
    files not already present in the index are processed (append mode).

    Returns *db_path* so callers can pass it straight to ``LexicalMatcher``.
    """
    os.makedirs(os.path.dirname(db_path) or ".", exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")

    if not rebuild and _index_ready(conn, ref_paths):
        logger.info("SQLite index is up-to-date – reusing.")
        conn.close()
        return db_path

    # ── Determine append vs. full rebuild ───────────────────────────────
    append_mode = False
    already_indexed: Set[str] = set()
    start_ref_id = 1
    prior_df: Counter[str] = Counter()

    if not rebuild:
        try:
            tables = {r[0] for r in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()}
            if {"refs", "postings", "df", "meta", "indexed_files"}.issubset(tables):
                for row in conn.execute("SELECT source_file FROM indexed_files"):
                    already_indexed.add(str(row[0]))
                max_id_row = conn.execute(
                    "SELECT COALESCE(MAX(ref_id), 0) FROM refs"
                ).fetchone()
                start_ref_id = int(max_id_row[0]) + 1
                for row in conn.execute("SELECT token, df FROM df"):
                    prior_df[str(row[0])] = int(row[1])
                if already_indexed and start_ref_id > 1:
                    append_mode = True
        except Exception:
            pass  # corrupt or missing – fall through to full build

    paths_to_index: Sequence[str]
    if append_mode:
        def _clean_stem(p: str) -> str:
            s = os.path.splitext(os.path.basename(p))[0]
            return s[4:] if s.startswith("ref_") else s

        paths_to_index = [
            p for p in ref_paths
            if _clean_stem(p) not in already_indexed
        ]
        if not paths_to_index:
            # Check if sheet_tokens needs to be (re)built
            _has_sheet_tokens = False
            try:
                _st_count = conn.execute("SELECT COUNT(*) FROM sheet_tokens").fetchone()[0]
                _has_sheet_tokens = int(_st_count) > 0
            except Exception:
                pass

            if _has_sheet_tokens:
                logger.info(
                    f"All {len(ref_paths)} reference files already indexed – nothing to append."
                )
                # Update signature so _index_ready passes next time
                conn.execute(
                    "INSERT OR REPLACE INTO meta VALUES (?, ?)",
                    ("ref_signature", _ref_signature(ref_paths)),
                )
                conn.commit()
                conn.close()
                return db_path
            else:
                logger.info(
                    "All files indexed but sheet_tokens missing – rebuilding signatures …"
                )
                # Fall through to the sheet_tokens computation below
                # by setting paths_to_index to empty (skip row insertion)
                # but NOT returning early.
                # Ensure the table exists (old index may lack it)
                conn.executescript("""
                    CREATE TABLE IF NOT EXISTS sheet_tokens (
                        sheet_name TEXT,
                        token      TEXT,
                        score      REAL
                    );
                    CREATE INDEX IF NOT EXISTS idx_sheet_tokens_sheet ON sheet_tokens(sheet_name);
                """)
        if paths_to_index:
            logger.info(
                f"Appending {len(paths_to_index)} new file(s) "
                f"({len(already_indexed)} already indexed, next ref_id={start_ref_id})"
            )
    else:
        paths_to_index = list(ref_paths)
        logger.info("Building SQLite lexical index …")
        _reset_index(conn)
        start_ref_id = 1
        prior_df.clear()

    cur = conn.cursor()
    ref_sig = _ref_signature(ref_paths)
    cur.executemany(
        "INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?)",
        [("schema_version", SCHEMA_VERSION), ("ref_signature", ref_sig), ("build_complete", "0")],
    )
    conn.commit()

    insert_buf: List[tuple] = []
    posting_buf: List[tuple] = []
    df_counter: Counter[str] = Counter(prior_df)
    ref_id = start_ref_id

    try:
        for path in paths_to_index:
            raw_stem = os.path.splitext(os.path.basename(path))[0]
            # Strip 'ref_' prefix added by the download step so that the
            # stored source_file matches the original S3 key stem.
            source_stem = raw_stem[4:] if raw_stem.startswith("ref_") else raw_stem
            logger.info(f"Indexing {os.path.basename(path)} …")
            for discipline, source_file, sheet, pc, prefixed, leaf, specs, toks in iter_ref_rows(path):
                segs = split_hierarchy(prefixed)
                family_parts = list(segs[:3]) if segs else [sheet]
                # For compact codes (e.g. p1316ACC), extract section+
                # subsection from the code itself to give a more meaningful
                # family_key than just the sheet name.
                _compact_fk = re.match(r'^[A-Za-z](\d{2})(\d{2})', (pc or "").strip())
                if _compact_fk and len(segs) <= 1:
                    family_parts = [f"sec{_compact_fk.group(1)}", f"sub{_compact_fk.group(2)}"]
                family_key = " | ".join([discipline, sheet] + family_parts)
                insert_buf.append((
                    ref_id, discipline, source_file, sheet, pc,
                    prefixed, leaf, family_key,
                    normalize_text(prefixed), normalize_text(leaf),
                    ",".join(specs["dn"]),  ",".join(specs["dia"]),
                    ",".join(specs["mm"]),  ",".join(specs["mpa"]),
                    ",".join(specs["kv"]),  ",".join(specs["dims"]),
                    ",".join(specs["mm2"]), ",".join(specs["cores"]),
                    ",".join(specs["thk"]), ",".join(specs["fire"]),
                    ",".join(specs["pn"]),
                    ",".join(specs.get("pipe_mat", ())),
                    ",".join(specs.get("valve_type", ())),
                    ",".join(specs.get("concrete_elem", ())),
                    ",".join(specs.get("cable_insul", ())),
                    ",".join(specs.get("schedule", ())),
                    ",".join(specs.get("pipe_grade", ())),
                    ",".join(specs.get("fitting_type", ())),
                    extract_code_prefix(pc),
                ))
                for tok in toks:
                    posting_buf.append((tok, ref_id))
                    df_counter[tok] += 1
                ref_id += 1

                if len(insert_buf) >= 5000:
                    cur.executemany(
                        "INSERT INTO refs VALUES (" + ",".join("?" * 29) + ")",
                        insert_buf,
                    )
                    cur.executemany("INSERT INTO postings VALUES (?,?)", posting_buf)
                    conn.commit()
                    insert_buf.clear()
                    posting_buf.clear()

            # Record this file as indexed
            cur.execute(
                "INSERT OR REPLACE INTO indexed_files (source_file, file_path) VALUES (?, ?)",
                (source_stem, os.path.abspath(path)),
            )

        if insert_buf:
            cur.executemany("INSERT INTO refs VALUES (" + ",".join("?" * 29) + ")", insert_buf)
            cur.executemany("INSERT INTO postings VALUES (?,?)", posting_buf)

        # Full df table replacement (includes prior + new counts)
        cur.execute("DELETE FROM df")
        cur.executemany("INSERT INTO df VALUES (?,?)", list(df_counter.items()))

        # ── Compute sheet-level token signatures ──────────────────────
        # For each sheet, find the top-K most distinctive tokens using
        # a sheet-level TF-IDF score.  Stored once in the DB so that
        # LexicalMatcher can load them instantly at startup.
        logger.info("Computing sheet token signatures …")
        cur.execute("DELETE FROM sheet_tokens")

        # 1. Per-sheet-token counts  (postings × refs GROUP BY)
        stc: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        sheet_sizes: Dict[str, int] = defaultdict(int)
        cur2 = conn.cursor()
        cur2.execute(
            "SELECT r.sheet_name, p.token, COUNT(*) "
            "FROM postings p JOIN refs r ON p.ref_id = r.ref_id "
            "GROUP BY r.sheet_name, p.token"
        )
        for _sheet, _tok, _cnt in cur2:
            stc[_sheet][_tok] = _cnt
            sheet_sizes[_sheet] += _cnt

        num_sheets = max(len(stc), 1)
        # 2. How many sheets each token appears in
        tok_sheet_cnt: Counter[str] = Counter()
        for _sheet, _toks in stc.items():
            for _tok in _toks:
                tok_sheet_cnt[_tok] += 1

        # 3. Sheet-level TF-IDF, keep top-50 per sheet
        _SIG_TOP_K = 50
        sig_buf: List[tuple] = []
        for _sheet, _toks in stc.items():
            _sz = max(sheet_sizes[_sheet], 1)
            scored = []
            for _tok, _cnt in _toks.items():
                _tf = _cnt / _sz
                _idf = math.log(num_sheets / max(tok_sheet_cnt[_tok], 1)) + 1.0
                scored.append((_tok, _tf * _idf))
            scored.sort(key=lambda x: x[1], reverse=True)
            for _tok, _sc in scored[:_SIG_TOP_K]:
                sig_buf.append((_sheet, _tok, _sc))
        cur.executemany("INSERT INTO sheet_tokens VALUES (?,?,?)", sig_buf)
        logger.info(f"Sheet signatures: {len(sig_buf)} entries across {len(stc)} sheets")

        total_count = ref_id - 1
        cur.executemany(
            "INSERT OR REPLACE INTO meta VALUES (?,?)",
            [
                ("ref_count", str(total_count)),
                ("schema_version", SCHEMA_VERSION),
                ("ref_signature", ref_sig),
                ("build_complete", "1"),
            ],
        )
        conn.commit()
        new_count = ref_id - start_ref_id
        logger.info(
            f"Index {'appended' if append_mode else 'built'}: "
            f"{new_count:,} new rows (total {total_count:,}), "
            f"{len(df_counter):,} unique tokens"
        )
    except Exception:
        cur.executemany(
            "INSERT OR REPLACE INTO meta VALUES (?,?)",
            [("schema_version", SCHEMA_VERSION), ("ref_signature", ref_sig), ("build_complete", "0")],
        )
        conn.commit()
        raise
    finally:
        conn.close()

    return db_path


# ═════════════════════════════════════════════════════════════════════════
# Matcher Engine
# ═════════════════════════════════════════════════════════════════════════

class LexicalMatcher:
    """
    Lexical candidate search engine for price-code matching.

    Performs TF-IDF-style scoring with domain-specific boosting,
    hard spec filters and discipline routing.

    Parameters
    ----------
    db_path : str
        Path to the SQLite lexical index built by ``build_index()``.
    source_files : list[str] | None
        Optional list of reference-file stems to restrict search.
    max_candidates : int
        Maximum final candidates returned per search.
    """

    # Tuning knobs (class-level defaults)
    # Tuning knobs (class-level defaults)
    HARD_POSTINGS_LIMIT = 50000
    MAX_QUERY_TERMS = 22
    INITIAL_POOL_LIMIT = 5000
    RELATIVE_CUTOFF = 0.50
    MIN_ABS_SCORE = 1.60
    MIN_OVERLAP_TOKENS = 1

    @classmethod
    async def create(
        cls,
        db_path: str,
        source_files: Optional[List[str]] = None,
        max_candidates: int = 15,
    ) -> "LexicalMatcher":
        """Async factory – loads the full index into memory for zero-IO search.

        After creation the SQLite file is closed; every ``search()`` call
        uses pure Python dict lookups so 200+ items can run concurrently
        without contending on a single DB connection.
        """
        self = cls.__new__(cls)
        self.db_path = db_path
        self.max_candidates = max_candidates

        # Heavy bulk-load runs in a worker thread so we don't block the
        # event loop (takes a few seconds for large indices).
        data = await asyncio.to_thread(cls._load_index, db_path, source_files)

        self.df: Dict[str, int] = data["df"]
        self.ref_count: int = data["ref_count"]
        self.idf: Dict[str, float] = data["idf"]
        self._valid_ref_ids: Optional[Set[int]] = data["valid_ref_ids"]
        self.sheet_sigs: Dict[str, Set[str]] = data["sheet_sigs"]
        self.sheet_discipline: Dict[str, str] = data["sheet_discipline"]
        self._postings: Dict[str, List[int]] = data["postings"]
        self._refs: Dict[int, Dict[str, Any]] = data["refs"]
        self._prefix_groups: Dict[str, List[int]] = data.get("prefix_groups", {})
        self._spec_index: Dict[str, Dict[str, frozenset]] = data.get("spec_index", {})

        logger.info(
            f"LexicalMatcher ready (in-memory): {self.ref_count:,} refs, "
            f"{len(self.df):,} unique tokens, {len(self.sheet_sigs)} sheet sigs, "
            f"{len(self._postings):,} posting lists, "
            f"max_candidates={max_candidates}"
        )
        # Cache for on-the-fly categorical spec extraction from ref descriptions.
        # Populated lazily during reranking; avoids adding columns to the DB.
        self._ref_cat_cache: Dict[int, Dict[str, Tuple[str, ...]]] = {}
        return self

    # ── On-the-fly categorical spec extraction ──────────────────────────
    # Categorical spec columns stored in v4 schema
    _CAT_DB_COLS = {
        "pipe_mat":      "pipe_mat_csv",
        "valve_type":    "valve_type_csv",
        "concrete_elem": "concrete_elem_csv",
        "cable_insul":   "cable_insul_csv",
        "schedule":      "schedule_csv",
        "pipe_grade":    "pipe_grade_csv",
        "fitting_type":  "fitting_type_csv",
    }

    def _get_ref_cat_specs(self, ref_id: int) -> Dict[str, Tuple[str, ...]]:
        """Extract categorical specs from a ref, preferring DB columns (v4)."""
        cached = self._ref_cat_cache.get(ref_id)
        if cached is not None:
            return cached
        ref = self._refs.get(ref_id)
        if ref is None:
            return {}
        # Try DB columns first (v4 schema stores them)
        has_db_cols = ref.get("pipe_mat_csv") is not None
        if has_db_cols:
            cat: Dict[str, Tuple[str, ...]] = {}
            for spec_key, col in self._CAT_DB_COLS.items():
                raw = ref.get(col) or ""
                vals = tuple(v.strip() for v in str(raw).split(",") if v.strip())
                cat[spec_key] = vals
            self._ref_cat_cache[ref_id] = cat
            return cat
        # Fallback: on-the-fly extraction for v3 DBs
        desc = ref.get("prefixed_description", "") or ref.get("description", "") or ""
        low = desc.lower()
        cat = _extract_categorical(low)
        self._ref_cat_cache[ref_id] = cat
        return cat

    # ── bulk loader (runs in a thread) ──────────────────────────────────

    @staticmethod
    def _load_index(
        db_path: str,
        source_files: Optional[List[str]],
    ) -> Dict[str, Any]:
        """Synchronously load the entire SQLite index into Python dicts."""
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")

        # ── DF / IDF ────────────────────────────────────────────────────
        df: Dict[str, int] = {}
        for row in conn.execute("SELECT token, df FROM df"):
            df[str(row["token"])] = int(row["df"])

        ref_count = int(conn.execute("SELECT COUNT(*) FROM refs").fetchone()[0])

        idf: Dict[str, float] = {
            tok: math.log((ref_count + 1) / (d + 1)) + 1.0
            for tok, d in df.items()
        }

        # ── Source-file filter ───────────────────────────────────────────
        valid_ref_ids: Optional[Set[int]] = None
        if source_files:
            ph = ",".join("?" for _ in source_files)
            valid_ref_ids = {
                int(r[0])
                for r in conn.execute(
                    f"SELECT ref_id FROM refs WHERE source_file IN ({ph})",
                    source_files,
                )
            }
            logger.info(
                f"LexicalMatcher: {len(valid_ref_ids):,} refs from "
                f"{source_files} (of {ref_count:,} total)"
            )

        # ── Sheet token signatures ──────────────────────────────────────
        sheet_sigs: Dict[str, Set[str]] = defaultdict(set)
        try:
            for row in conn.execute("SELECT sheet_name, token FROM sheet_tokens"):
                sheet_sigs[str(row["sheet_name"])].add(str(row["token"]))
        except Exception:
            logger.warning("sheet_tokens table not found – sheet affinity disabled")

        # ── Sheet → discipline mapping ──────────────────────────────────
        sheet_discipline: Dict[str, str] = {}
        try:
            for row in conn.execute("SELECT DISTINCT sheet_name, discipline FROM refs"):
                sheet_discipline[str(row["sheet_name"])] = str(row["discipline"])
        except Exception:
            pass

        # ── Postings: token → list[ref_id] ──────────────────────────────
        logger.info("Loading postings into memory …")
        postings: Dict[str, List[int]] = defaultdict(list)
        for row in conn.execute("SELECT token, ref_id FROM postings"):
            postings[row["token"]].append(int(row["ref_id"]))
        postings = dict(postings)  # shed defaultdict overhead

        # ── Refs: ref_id → dict ─────────────────────────────────────────
        logger.info("Loading refs into memory …")
        # Intern frequently-repeated strings to save memory
        _intern_cache: Dict[str, str] = {}
        def _intern(s: str) -> str:
            if s not in _intern_cache:
                _intern_cache[s] = s
            return _intern_cache[s]

        refs: Dict[int, Dict[str, Any]] = {}
        cursor = conn.execute("SELECT * FROM refs")
        col_names = [desc[0] for desc in cursor.description]
        # Columns with few unique values → intern
        intern_cols = {"discipline", "source_file", "sheet_name", "code_prefix"}
        for row in cursor:
            d: Dict[str, Any] = {}
            for col, val in zip(col_names, row):
                if col in intern_cols and isinstance(val, str):
                    d[col] = _intern(val)
                else:
                    d[col] = val
            refs[int(d["ref_id"])] = d

        # ── Prefix groups: code_prefix → [ref_id, …] ───────────────
        # Enables prefix-expansion during search: if TF-IDF finds any
        # code from a prefix group, all siblings can be injected cheaply.
        prefix_groups: Dict[str, List[int]] = defaultdict(list)
        for rid, ref in refs.items():
            pfx = ref.get("code_prefix", "")
            if not pfx:
                # Compute on the fly for legacy v2 indexes lacking the column
                pfx = extract_code_prefix(ref.get("price_code", ""))
            if pfx:
                prefix_groups[pfx].append(rid)
        prefix_groups = dict(prefix_groups)  # shed defaultdict overhead

        # ── Spec inverted indexes ───────────────────────────────────
        # For each numeric AND categorical DB-column spec, build
        # {normalised_value: set(ref_id)} so the search can inject / boost
        # spec-matching refs into the candidate pool *before* the top-N cut.
        _SPEC_COLS = ("mpa_csv", "dn_csv", "dia_csv", "kv_csv",
                      "mm2_csv", "core_csv", "pn_csv",
                      "pipe_mat_csv", "valve_type_csv", "concrete_elem_csv",
                      "cable_insul_csv", "schedule_csv", "pipe_grade_csv")
        spec_index: Dict[str, Dict[str, Set[int]]] = {
            col: defaultdict(set) for col in _SPEC_COLS
        }
        for rid, ref in refs.items():
            for col in _SPEC_COLS:
                raw = ref.get(col)
                if not raw:
                    continue
                s = str(raw).strip()
                if not s:
                    continue
                for v in s.split(","):
                    v = v.strip()
                    if v:
                        spec_index[col][v].add(rid)
        # Convert inner defaultdicts to plain dicts + freeze sets
        spec_index = {
            col: {v: frozenset(rids) for v, rids in vmap.items()}
            for col, vmap in spec_index.items()
        }
        _si_total = sum(len(rids) for vmap in spec_index.values() for rids in vmap.values())
        logger.info(f"Spec inverted indexes: {_si_total:,} entries across {len(_SPEC_COLS)} columns")

        conn.close()
        logger.info(
            f"Index loaded: {len(refs):,} refs, "
            f"{len(postings):,} posting lists, {sum(len(v) for v in postings.values()):,} entries, "
            f"{len(prefix_groups):,} prefix groups"
        )

        return {
            "df": df,
            "ref_count": ref_count,
            "idf": idf,
            "valid_ref_ids": valid_ref_ids,
            "sheet_sigs": dict(sheet_sigs),
            "sheet_discipline": sheet_discipline,
            "postings": postings,
            "refs": refs,
            "prefix_groups": prefix_groups,
            "spec_index": spec_index,
        }

    # ── public API ──────────────────────────────────────────────────────

    async def search(self, item: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Async wrapper – delegates to ``search_sync`` via the event loop.

        For best throughput call ``search_sync`` directly from a thread
        pool (see ``PriceCodePipeline.process_file``).
        """
        return self.search_sync(item)

    def search_sync(self, item: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Find candidate price codes for a BOQ item (synchronous / thread-safe).

        Pure CPU work – no I/O, no awaits.  Designed to be called from a
        ``ThreadPoolExecutor`` so the event loop stays free for LLM I/O.

        Parameters
        ----------
        item : dict
            Keys: description, parent, grandparent, category_path, unit, item_code

        Returns
        -------
        list[dict]
            Each dict has: price_code, description, category, score, metadata
            (compatible with ``PriceCodeMatcher.llm_match``).
        """
        import time as _time
        _t_start = _time.perf_counter()

        qweights, desc_specs, ctx_specs, guessed_disc, distinctive, alpha_dist, short = self._weighted_query(item)
        if not qweights:
            return []

        _t_query = _time.perf_counter()

        is_airfield = bool(alpha_dist & AIRFIELD_HINTS)

        # ── 1. Token-posting lookup ─────────────────────────────────────
        scored_pool: Dict[int, float] = defaultdict(float)
        ordered = sorted(qweights.items(), key=lambda kv: kv[1], reverse=True)[
            : self.MAX_QUERY_TERMS
        ]
        _total_postings_scanned = 0

        # Identify spec-value tokens so we exempt them from the
        # frequency filter.  Numbers like "40" (from C32/40) appear
        # in >8 % of refs and would be dropped, but they carry vital
        # specification information.
        _spec_tok_set: Set[str] = set()
        for _sk in ("mpa", "dn", "dia", "kv", "mm2", "cores", "pn",
                    "dims", "mm"):
            for _sv in desc_specs.get(_sk, ()):
                _spec_tok_set.add(_sv)
                # Also add common suffixed forms (e.g. "dn150" → "150")
                if _sv.replace(".", "", 1).isdigit():
                    _spec_tok_set.add(_sv)

        for _tok_idx, (tok, qscore) in enumerate(ordered):
            df = self.df.get(tok, 0)
            if not df:
                continue
            # Allow spec-value tokens through the frequency filter —
            # numbers like "40" or "150" are high-DF but carry vital
            # specification meaning when extracted by extract_specs().
            _is_spec_tok = tok in _spec_tok_set
            if not _is_spec_tok and self.ref_count and (df / self.ref_count) > 0.08:
                continue
            damp = 1.0 / (1.0 + math.log(df + 1))
            # Give spec tokens a slight extra weight so they contribute
            # more to pool ordering even when their IDF is low.
            if _is_spec_tok:
                damp = max(damp, 0.25)
            posting_ids = self._postings.get(tok, ())
            _scan_len = min(len(posting_ids), self.HARD_POSTINGS_LIMIT)
            _total_postings_scanned += _scan_len
            for rid in posting_ids[:self.HARD_POSTINGS_LIMIT]:
                if self._valid_ref_ids is not None and rid not in self._valid_ref_ids:
                    continue
                scored_pool[rid] += qscore * damp

        if not scored_pool:
            # Relaxed fallback – grab anything that shares a token
            for tok, qscore in ordered:
                posting_ids = self._postings.get(tok, ())
                for rid in posting_ids[:min(self.HARD_POSTINGS_LIMIT, 250)]:
                    if self._valid_ref_ids is not None and rid not in self._valid_ref_ids:
                        continue
                    scored_pool[rid] += qscore * 0.05

        if not scored_pool:
            return []

        _t_postings = _time.perf_counter()

        # ── 1b. Spec-aware pool boosting & injection ────────────────────
        # Refs whose DB spec columns match the query's numeric specs get
        # a score boost in the pool.  Refs that weren't found by TF-IDF
        # at all (complete vocabulary mismatch) get injected with a
        # minimum score so they survive the top-N cut and can be properly
        # reranked later.
        _SPEC_MAP = {
            "mpa":   "mpa_csv",
            "dn":    "dn_csv",
            "dia":   "dia_csv",
            "kv":    "kv_csv",
            "mm2":   "mm2_csv",
            "cores": "core_csv",
            "pn":    "pn_csv",
        }
        # Also boost/inject based on categorical specs from ctx_specs
        _CAT_MAP = {
            "pipe_mat":      "pipe_mat_csv",
            "valve_type":    "valve_type_csv",
            "concrete_elem": "concrete_elem_csv",
            "cable_insul":   "cable_insul_csv",
            "schedule":      "schedule_csv",
            "pipe_grade":    "pipe_grade_csv",
        }
        _pool_min = min(scored_pool.values()) if scored_pool else 0.0
        _pool_median = 0.0
        if scored_pool:
            _sorted_scores = sorted(scored_pool.values(), reverse=True)
            _pool_median = _sorted_scores[len(_sorted_scores) // 2]
        _spec_injected = 0
        _spec_boosted = 0
        for spec_key, col_name in _SPEC_MAP.items():
            qvals = set(desc_specs.get(spec_key, ()))
            if not qvals:
                continue
            col_idx = self._spec_index.get(col_name, {})
            matching_rids: Set[int] = set()
            for qv in qvals:
                matching_rids |= col_idx.get(qv, frozenset())
            # Also cross-reference dia ↔ dn
            if spec_key == "dia":
                dn_idx = self._spec_index.get("dn_csv", {})
                for qv in qvals:
                    matching_rids |= dn_idx.get(qv, frozenset())
            elif spec_key == "dn":
                dia_idx = self._spec_index.get("dia_csv", {})
                for qv in qvals:
                    matching_rids |= dia_idx.get(qv, frozenset())
            if self._valid_ref_ids is not None:
                matching_rids &= self._valid_ref_ids
            for rid in matching_rids:
                if rid in scored_pool:
                    # Boost existing pool entries by 30% per matching spec
                    scored_pool[rid] *= 1.30
                    _spec_boosted += 1
                else:
                    # Inject with median score — enough to survive the cut,
                    # but low enough that pure spec match without text
                    # relevance won't dominate.
                    scored_pool[rid] = max(_pool_median * 0.60, _pool_min)
                    _spec_injected += 1

        # Categorical spec boosting (pipe_mat, valve_type, etc.)
        for spec_key, col_name in _CAT_MAP.items():
            qvals = set(ctx_specs.get(spec_key, ()))
            if not qvals:
                continue
            col_idx = self._spec_index.get(col_name, {})
            matching_rids: Set[int] = set()
            for qv in qvals:
                matching_rids |= col_idx.get(qv, frozenset())
            if self._valid_ref_ids is not None:
                matching_rids &= self._valid_ref_ids
            for rid in matching_rids:
                if rid in scored_pool:
                    scored_pool[rid] *= 1.25
                    _spec_boosted += 1
                else:
                    scored_pool[rid] = max(_pool_median * 0.50, _pool_min)
                    _spec_injected += 1

        if _spec_boosted or _spec_injected:
            logger.debug(
                f"Spec pool: boosted={_spec_boosted}, injected={_spec_injected}"
            )

        # ── 2. Fetch top-N ref rows ─────────────────────────────────────
        prelim = sorted(scored_pool.items(), key=lambda kv: kv[1], reverse=True)[
            : self.INITIAL_POOL_LIMIT
        ]
        ref_rows = self._fetch_ref_rows([rid for rid, _ in prelim])

        _t_fetch = _time.perf_counter()

        # ── 3. Re-rank ──────────────────────────────────────────────────
        core_norm = normalize_text(item.get("description", ""))
        _query_toks_list = tokenize_normalized(core_norm)  # ordered token list for bigram matching
        boq_unit_fam = _unit_family(item.get("unit", "") or "")
        reranked: List[Dict[str, Any]] = []
        _yield_counter = 0

        # ── Pre-compute loop-invariant values ───────────────────────────
        distinctive_objects = distinctive & OBJECT_TOKENS
        parent_str = item.get("parent", "") or ""
        gp_str = item.get("grandparent", "") or ""
        _catpath_str = item.get("category_path", "") or ""
        route_toks = alpha_dist | self._alpha_tokens(
            set(tokenize(" ; ".join([parent_str, gp_str, _catpath_str])))
        )
        # Scope detection from hierarchy context (pre-computed once)
        expected_scope = _detect_expected_scope(parent_str, gp_str)
        # MEP sub-discipline prefix (p/h/f/Z) from hierarchy keywords
        expected_mep_prefix = (
            _detect_mep_prefix(
                parent_str, gp_str, _catpath_str,
            )
            if guessed_disc == "mechanical" else None
        )
        # Pre-compute parent/grandparent/category_path alpha tokens for segment matching
        _parent_alpha = self._alpha_tokens(set(tokenize(parent_str))) if parent_str else set()
        _gp_alpha = self._alpha_tokens(set(tokenize(gp_str))) if gp_str else set()
        _cp_alpha = self._alpha_tokens(set(tokenize(_catpath_str))) if _catpath_str else set()
        _ctx_alpha = _parent_alpha | _gp_alpha | _cp_alpha  # combined hierarchy context tokens
        # Remove description tokens to avoid double-counting with leaf overlap
        _ctx_alpha -= alpha_dist

        # Sheet affinity: compute once, reuse for all candidates
        sheet_aff: Dict[str, float] = {}
        _sa_best_aff = 0.0
        _sa_rank_2 = 0.0
        if self.sheet_sigs and route_toks:
            sheet_aff = self._sheet_affinity(route_toks, qweights)
            if sheet_aff:
                _sa_best_aff = max(sheet_aff.values())
                if _sa_best_aff > 0:
                    sorted_aff = sorted(sheet_aff.values(), reverse=True)
                    _sa_rank_2 = (
                        sorted_aff[1] / _sa_best_aff
                        if len(sorted_aff) > 1 else 0.0
                    )

        for ref_id, lex_score in prelim:
            _yield_counter += 1

            ref = ref_rows.get(ref_id)
            if ref is None:
                continue
            if not self._passes_hard_spec_filters(desc_specs, ref):
                continue

            final = lex_score
            leaf_norm = clean_text(ref["norm_leaf"])
            full_norm = clean_text(ref["norm_text"])
            ref_toks = set(tokenize_normalized(full_norm))
            overlap = distinctive & ref_toks
            alpha_overlap = alpha_dist & ref_toks
            obj_overlap = bool(distinctive_objects & ref_toks)

            # Sequence similarity
            if core_norm and leaf_norm:
                ratio = _rapidfuzz_ratio(core_norm, leaf_norm) / 100.0
                final += 1.35 * ratio
                if core_norm in full_norm or leaf_norm in core_norm:
                    final += 0.8
                # Also check against the full prefixed description —
                # catches matches where query vocabulary aligns with the
                # rate-book hierarchy path.
                if full_norm and full_norm != leaf_norm:
                    full_ratio = _rapidfuzz_ratio(core_norm, full_norm) / 100.0
                    if full_ratio > ratio:
                        final += 0.45 * (full_ratio - ratio)

            # Bigram sequence bonus — consecutive token pairs from query
            # appearing in the ref indicate a phrase match (stronger than
            # individual tokens).  E.g. "ready mix" or "steel bar".
            if len(_query_toks_list) >= 2:
                _ref_norm_str = full_norm
                _bigram_hits = 0
                for _bi in range(len(_query_toks_list) - 1):
                    _bg = _query_toks_list[_bi] + " " + _query_toks_list[_bi + 1]
                    if _bg in _ref_norm_str:
                        _bigram_hits += 1
                if _bigram_hits > 0:
                    final += 0.30 * min(_bigram_hits, 4)

            # Leaf-token overlap  (STRONG signal)
            # Tokens matching the ref's LEAF description are far more
            # meaningful than tokens matching only the hierarchy prefix.
            # Multiplicative to scale with the base TF-IDF score.
            leaf_toks = set(tokenize_normalized(leaf_norm))
            leaf_alpha = self._alpha_tokens(leaf_toks)
            leaf_overlap = alpha_dist & leaf_alpha
            if alpha_dist and leaf_alpha:
                leaf_ratio = len(leaf_overlap) / max(1, len(alpha_dist))
                if leaf_ratio >= 0.5:
                    final *= 1.0 + 0.80 * leaf_ratio   # up to 1.80×
                elif leaf_ratio == 0.0:
                    final *= 0.50                       # heavy penalty

            # ── Segment-level hierarchy matching ────────────────────────
            # The ref's prefixed_description has semicolon-delimited
            # segments: "Ready Mix Concrete ; Cast In Situ ; leaf …".
            # The intermediate (non-leaf) segments are classification
            # context.  When BOQ parent/grandparent tokens match these
            # intermediate segments, boost the candidate — it means the
            # rate-book classification aligns with the BOQ hierarchy.
            if _ctx_alpha:
                _prefix_desc = clean_text(ref["prefixed_description"])
                _seg_parts = [s.strip() for s in _prefix_desc.split(";")]
                if len(_seg_parts) > 1:
                    # All segments except the last (leaf)
                    _intermediate = " ".join(_seg_parts[:-1])
                    _inter_toks = self._alpha_tokens(
                        set(tokenize_normalized(normalize_text(_intermediate)))
                    )
                    if _inter_toks:
                        _seg_overlap = _ctx_alpha & _inter_toks
                        _seg_ratio = len(_seg_overlap) / max(1, len(_ctx_alpha))
                        if _seg_ratio >= 0.3:
                            final *= 1.0 + 0.35 * _seg_ratio  # up to ~1.35×
                        # NOTE: no penalty for _seg_ratio==0 — vocabulary
                        # between BOQ parents and rate-book hierarchy
                        # segments differs across disciplines (e.g. mech
                        # BOQ says "Soil pipe work" but ref intermediate
                        # says "Facility Sanitary Sewerage").  Absence of
                        # match is not evidence of wrong candidate.

            # Token overlap bonuses
            if distinctive:
                final += 1.25 * (len(overlap) / max(1, len(distinctive)))
            if alpha_dist:
                final += 1.1 * (len(alpha_overlap) / max(1, len(alpha_dist)))

            # ── Bigram sequence bonus ───────────────────────────────────
            # Two consecutive query tokens appearing in order in the ref
            # description is a much stronger signal than random unigram
            # matches (e.g. "copper pipe" vs "pipe" + "copper" separately).
            if len(_query_toks_list) >= 2:
                _ref_tok_list = tokenize_normalized(full_norm)
                _ref_tok_set_idx: Dict[str, List[int]] = defaultdict(list)
                for _ri, _rt in enumerate(_ref_tok_list):
                    _ref_tok_set_idx[_rt].append(_ri)
                _bigram_hits = 0
                for _qi in range(len(_query_toks_list) - 1):
                    _t1, _t2 = _query_toks_list[_qi], _query_toks_list[_qi + 1]
                    _positions1 = _ref_tok_set_idx.get(_t1, [])
                    _positions2 = _ref_tok_set_idx.get(_t2, [])
                    if _positions1 and _positions2:
                        for _p1 in _positions1:
                            if (_p1 + 1) in _positions2:
                                _bigram_hits += 1
                                break
                if _bigram_hits > 0:
                    _bigram_ratio = _bigram_hits / max(1, len(_query_toks_list) - 1)
                    final += 0.6 * _bigram_ratio  # up to +0.6 for full bigram match

            # Spec scoring (additive)
            final += self._spec_score(
                ctx_specs, ref, has_object_support=obj_overlap or bool(alpha_overlap)
            )

            # Spec matching (multiplicative) – key variant-differentiating
            # specs like mpa, concrete_elem, pipe_mat scale the entire
            # score so that matching specs dominate the ranking.
            final *= self._spec_multiplier(desc_specs, ctx_specs, ref)

            # Discipline routing  (multiplicative – much stronger signal)
            # Fix 3b: Strengthen cross-discipline penalty when the
            # description is thin / generic.  For thin items, token
            # overlap is very weak so wrong-discipline candidates can
            # easily outscore the correct ones (e.g. Z_Utilities HDPE
            # beating plumbing uPVC for "50 mm diameter").
            ref_disc = clean_text(ref["discipline"])
            ref_sheet = clean_text(ref["sheet_name"])
            ref_sheet_low = ref_sheet.lower()

            # ── Infer discipline for "unknown" refs ──────────────────
            # Specialised source files (e.g. C_Concrete, M_Masonry)
            # may have been indexed without a discipline tag.  Infer
            # from source-file + sheet name via the shared rule table
            # so they participate in discipline routing correctly.
            if ref_disc == "unknown":
                ref_disc = _infer_discipline_from_context(
                    ref["source_file"], ref_sheet
                )

            if guessed_disc:
                if guessed_disc == ref_disc:
                    final *= 1.15            # same discipline boost
                elif short:
                    final *= 0.40            # strong penalty for thin descs
                else:
                    final *= 0.70            # normal cross-discipline penalty

            # ── MEP sub-discipline routing ──────────────────────────────
            # Within "mechanical", the price-code prefix (p/h/f/Z)
            # identifies the sub-discipline.  Identical physical items
            # (pipes, valves, insulation) exist across sub-disciplines,
            # so we boost matching and penalize mismatching prefixes.
            # Penalty is moderate for MEP-vs-MEP (prefix misdetection is
            # common on thin items) but stronger for non-MEP prefixes
            # (Y-insulation, C-civil) which are almost always wrong.
            if expected_mep_prefix:
                _ref_pc_raw = clean_text(ref["price_code"])
                _ref_prefix = _ref_pc_raw[0].upper() if _ref_pc_raw else ""
                if _ref_prefix and _ref_prefix.isalpha():
                    if _ref_prefix == expected_mep_prefix.upper():
                        final *= 1.20       # matching sub-discipline
                    elif _ref_prefix in ("P", "H", "F", "Z"):
                        # Wrong MEP sub-discipline — moderate penalty
                        final *= 0.60
                    else:
                        # Non-MEP prefix (Y, C, E, etc.) — stronger
                        if short:
                            final *= 0.45
                        else:
                            final *= 0.55

            # Airfield routing
            if is_airfield:
                if ref_disc == "electrical":
                    final *= 1.12
                if "transportation" in ref_sheet_low or ref_sheet_low.startswith("a_"):
                    final *= 1.25
                elif ref_disc != "electrical":
                    final *= 0.92

            # ── Data-driven sheet routing (pre-computed above loop) ────
            if sheet_aff and _sa_best_aff > 0:
                aff = sheet_aff.get(ref_sheet, 0.0)
                norm_aff = aff / _sa_best_aff   # 0..1
                if short:
                    if norm_aff >= 0.95:
                        final *= 1.15          # gentle boost
                    elif norm_aff > _sa_rank_2 * 0.9 and norm_aff >= 0.5:
                        final *= 1.05
                    elif norm_aff < 0.15:
                        final *= 0.80
                    else:
                        final *= 0.92
                else:
                    if norm_aff >= 0.95:
                        final *= 1.35          # best sheet
                    elif norm_aff > _sa_rank_2 * 0.9 and norm_aff >= 0.5:
                        final *= 1.10          # close runner-up
                    elif norm_aff < 0.15:
                        final *= 0.65          # unrelated sheet
                    else:
                        final *= 0.85          # weak match

            # Unit compatibility scoring – multiplicative to strongly
            # separate volume (m3=concrete) from area (m2=formwork) from
            # weight (t=reinforcement) etc.
            if boq_unit_fam:
                ref_unit_fam = _infer_ref_unit_family(
                    clean_text(ref["prefixed_description"])
                )
                if ref_unit_fam:
                    if ref_unit_fam == boq_unit_fam:
                        final *= 1.15  # reward matching unit family
                    else:
                        final *= 0.45  # heavy penalty for unit mismatch

            # Penalize numeric-only attraction without semantic support
            if ctx_specs and not alpha_overlap and not obj_overlap:
                final -= 0.55

            # ── Prefer specific subcategory over generic "00" ───────────
            # Rate-book codes: [Disc] [Cat] [Subcat] [Suffix]
            # Subcategory "00" is a generic/template placeholder; codes
            # like C 31 13, C 11 13, F 30 36 are project-specific and
            # should be preferred when both exist.
            _pc = clean_text(ref["price_code"])
            _pc_parts = _pc.split()
            # Subcategory "00" penalty — works for both spaced and compact
            _subcat = None
            if len(_pc_parts) >= 3:
                _subcat = _pc_parts[2]
            else:
                _compact = _parse_compact_code(_pc)
                if _compact:
                    _subcat = _compact[2]  # subcat from compact format
            if _subcat == "00":
                final *= 0.60  # penalise generic subcategory (prefer specific)

            # ── Scope scoring from hierarchy context ────────────────────
            # When parent/grandparent clearly indicates the scope of work
            # (e.g. "Supply ready mix concrete" → Supply Only = E,
            #  "Pour concrete…include labour" → Supply+Install = F),
            # boost candidates with the matching scope letter and
            # penalise those with a different scope.
            # Applied to both spaced and compact codes.
            if expected_scope:
                ref_scope = _extract_scope_letter(_pc)
                if ref_scope:
                    if expected_scope in ("E", "F"):
                        # E (Supply Only) and F (Supply+Install) are
                        # universal across all disciplines.
                        if ref_scope == expected_scope:
                            final *= 1.25      # matching scope boost
                        elif ref_scope in ("E", "F"):
                            final *= 0.80      # wrong supply scope
                    else:
                        # Concrete-specific scopes (A/B/C/D) — Civil only
                        _scope_disc = _pc_parts[0].upper() if _pc_parts else ""
                        if not _scope_disc:
                            _cp = _parse_compact_code(_pc)
                            if _cp:
                                _scope_disc = _cp[0].upper()
                        if _scope_disc == "C":
                            if ref_scope == expected_scope:
                                final *= 1.25  # matching scope boost
                            else:
                                final *= 0.85  # wrong scope (soft)

            reranked.append({
                "ref_id": ref_id,
                "score": round(final, 4),
                "overlap_count": len(overlap),
                "alpha_overlap_count": len(alpha_overlap),
                "discipline": ref_disc,
                "source_file": clean_text(ref["source_file"]),
                "sheet_name": clean_text(ref["sheet_name"]),
                "price_code": clean_text(ref["price_code"]),
                "description": clean_text(ref["prefixed_description"]),
                "leaf_description": clean_text(ref["leaf_description"]),
            })

        _t_rerank = _time.perf_counter()

        if not reranked:
            # Log timing even for empty results
            _desc_short = (item.get('description', '') or '')[:60]
            logger.info(
                f"[search-timing] '{_desc_short}' "
                f"query={(_t_query-_t_start)*1000:.0f}ms "
                f"postings={(_t_postings-_t_query)*1000:.0f}ms({_total_postings_scanned:,}scanned,{len(scored_pool):,}unique) "
                f"fetch={(_t_fetch-_t_postings)*1000:.0f}ms({len(prelim)}prelim) "
                f"rerank={(_t_rerank-_t_fetch)*1000:.0f}ms(0passed) "
                f"total={(_t_rerank-_t_start)*1000:.0f}ms -> 0 results"
            )
            return []

        # ── 3b. Prefix-aware sibling expansion ──────────────────────────
        # The breakdown structure shows that codes sharing the same prefix
        # (Trade+Section+Family) describe the SAME work type but differ
        # in V1/V2/V3 (size, material, rating, scope).  When TF-IDF finds
        # codes from a prefix group, the correct VARIANT might be missing
        # because its spec tokens scored lower than generic tokens.
        #
        # Fix: for each top-scoring prefix group, inject ALL its siblings
        # that weren't in the initial pool, then quick-score them with
        # spec matching + discipline routing so the right variant surfaces.
        _existing_ids = {r["ref_id"] for r in reranked}
        if self._prefix_groups:
            reranked.sort(key=lambda x: float(x["score"]), reverse=True)
            # Collect top prefix groups from the best-scoring candidates
            _seen_prefixes: Dict[str, float] = {}
            MAX_EXPAND_PREFIXES = 12
            for r in reranked:
                pfx = extract_code_prefix(r["price_code"])
                if pfx and pfx not in _seen_prefixes:
                    _seen_prefixes[pfx] = float(r["score"])
                    if len(_seen_prefixes) >= MAX_EXPAND_PREFIXES:
                        break

            _expansion_count = 0
            MAX_SIBLINGS_PER_PREFIX = 200  # cap to avoid blowup on huge groups
            for pfx, best_pfx_score in _seen_prefixes.items():
                sibling_ids = self._prefix_groups.get(pfx, [])
                # Skip overly large prefix groups — they're too generic
                # to benefit from full expansion (e.g. s 13 00 = 96K codes)
                if len(sibling_ids) > 5000:
                    continue
                _pfx_added = 0
                for sid in sibling_ids:
                    if _pfx_added >= MAX_SIBLINGS_PER_PREFIX:
                        break
                    if sid in _existing_ids:
                        continue
                    if self._valid_ref_ids is not None and sid not in self._valid_ref_ids:
                        continue
                    sref = self._refs.get(sid)
                    if sref is None:
                        continue
                    # Hard spec filter
                    if not self._passes_hard_spec_filters(desc_specs, sref):
                        continue

                    # Quick score: start from a fraction of the best sibling
                    # score, then apply spec scoring + discipline routing.
                    sib_score = best_pfx_score * 0.85

                    sib_score += self._spec_score(
                        ctx_specs, sref,
                        has_object_support=True,  # siblings inherit relevance
                    )
                    # Multiplicative spec matching for siblings too
                    sib_score *= self._spec_multiplier(desc_specs, ctx_specs, sref)

                    # Discipline routing
                    sib_disc = clean_text(sref["discipline"])
                    if sib_disc == "unknown":
                        sib_disc = _infer_discipline_from_context(
                            sref["source_file"], sref.get("sheet_name", "")
                        )
                    if guessed_disc:
                        if guessed_disc == sib_disc:
                            sib_score *= 1.15
                        elif short:
                            sib_score *= 0.40
                        else:
                            sib_score *= 0.70

                    # Scope scoring
                    if expected_scope:
                        sib_pc = clean_text(sref["price_code"])
                        sib_scope = _extract_scope_letter(sib_pc)
                        if sib_scope:
                            if expected_scope in ("E", "F"):
                                if sib_scope == expected_scope:
                                    sib_score *= 1.25
                                elif sib_scope in ("E", "F"):
                                    sib_score *= 0.80

                    # Unit compatibility for siblings
                    if boq_unit_fam:
                        sib_unit_fam = _infer_ref_unit_family(
                            clean_text(sref["prefixed_description"])
                        )
                        if sib_unit_fam:
                            if sib_unit_fam == boq_unit_fam:
                                sib_score *= 1.15
                            else:
                                sib_score *= 0.45

                    reranked.append({
                        "ref_id": sid,
                        "score": round(sib_score, 4),
                        "overlap_count": 2,  # minimum to pass filter
                        "alpha_overlap_count": 1,
                        "discipline": sib_disc,
                        "source_file": clean_text(sref["source_file"]),
                        "sheet_name": clean_text(sref.get("sheet_name", "")),
                        "price_code": clean_text(sref["price_code"]),
                        "description": clean_text(sref["prefixed_description"]),
                        "leaf_description": clean_text(sref.get("leaf_description", "")),
                    })
                    _existing_ids.add(sid)
                    _expansion_count += 1
                    _pfx_added += 1

            if _expansion_count:
                logger.debug(
                    f"Prefix expansion: +{_expansion_count} siblings from "
                    f"{len(_seen_prefixes)} prefix groups"
                )

        # ── 4. Filter & cap ─────────────────────────────────────────────
        reranked.sort(key=lambda x: float(x["score"]), reverse=True)
        best = float(reranked[0]["score"])
        floor = max(self.MIN_ABS_SCORE, best * self.RELATIVE_CUTOFF)

        filtered = [
            r
            for r in reranked
            if float(r["score"]) >= floor
            and int(r["overlap_count"]) >= self.MIN_OVERLAP_TOKENS
            and (int(r["alpha_overlap_count"]) >= 1 or int(r["overlap_count"]) >= 2)
        ]
        # If nothing passes the filter, keep the single best if strong enough
        if not filtered and reranked:
            top = reranked[0]
            if float(top["score"]) >= max(self.MIN_ABS_SCORE, 2.8) and (
                int(top["alpha_overlap_count"]) >= 1 or int(top["overlap_count"]) >= 2
            ):
                filtered = [top]

        # ── 4b. Diversity dedup by code stem ────────────────────────────
        # Many rate-book entries share near-identical descriptions and
        # differ only in the last letter of the price-code (scope variant:
        # A=Concrete Only, B=+Reinforcement, C=+Formwork, D=+Both, etc.).
        # Without dedup, a single element type can consume all candidate
        # slots.  We keep the top-scoring representative per "code stem"
        # (price_code with last letter stripped) up to max_candidates
        # groups, then expand each group with all its sibling variants
        # so the LLM can pick the right scope.
        filtered = self._diversity_dedup(filtered)

        _t_filter = _time.perf_counter()

        final_results = self._to_pipeline_format(filtered)

        _t_end = _time.perf_counter()
        _desc_short = (item.get('description', '') or '')[:60]
        logger.info(
            f"[search-timing] '{_desc_short}' "
            f"query={(_t_query-_t_start)*1000:.0f}ms "
            f"postings={(_t_postings-_t_query)*1000:.0f}ms({_total_postings_scanned:,}scanned,{len(scored_pool):,}unique) "
            f"fetch={(_t_fetch-_t_postings)*1000:.0f}ms({len(prelim)}prelim) "
            f"rerank={(_t_rerank-_t_fetch)*1000:.0f}ms({len(reranked)}passed/{len(prelim)}total) "
            f"filter={(_t_filter-_t_rerank)*1000:.0f}ms "
            f"total={(_t_end-_t_start)*1000:.0f}ms -> {len(final_results)} results"
        )

        # Convert to pipeline format
        return final_results

    async def close(self) -> None:
        """No-op – the DB is closed after initial load; kept for API compat."""
        pass

    # ── internal helpers ────────────────────────────────────────────────

    @staticmethod
    def _alpha_tokens(tokens: Set[str]) -> Set[str]:
        """Return tokens that are purely alphabetic (no digits).

        Compound spec tokens like 'dia25', '25mm', 'dn50' are excluded
        because they are numeric specs, not descriptive words.
        """
        return {t for t in tokens if re.fullmatch(r"[a-z]+", t)}

    @staticmethod
    def _csv_to_set(value: object) -> Set[str]:
        s = clean_text(value)
        return {x for x in s.split(",") if x} if s else set()

    def _sheet_affinity(
        self, route_toks: Set[str], qweights: Dict[str, float]
    ) -> Dict[str, float]:
        """Return {sheet_name: affinity_score} for the query tokens.

        For every sheet whose precomputed signature overlaps with
        *route_toks*, accumulate the IDF-weighted overlap.  Fully
        data-driven: adapts automatically when new sheets are added.
        """
        aff: Dict[str, float] = defaultdict(float)
        for sheet, sig_toks in self.sheet_sigs.items():
            overlap = route_toks & sig_toks
            if overlap:
                for tok in overlap:
                    aff[sheet] += qweights.get(tok, 1.0)
        return dict(aff)

    def _weighted_query(
        self, item: Dict[str, Any]
    ) -> Tuple[Dict[str, float], Dict[str, Tuple[str, ...]], Dict[str, Tuple[str, ...]], Optional[str], Set[str], Set[str], bool]:
        """Build weighted token dict, specs, discipline guess, distinctive tokens.

        Returns
        -------
        qweights, desc_specs, ctx_specs, guessed_disc, distinctive, alpha_dist, short
            *desc_specs* are from the description only (used for hard filtering).
            *ctx_specs* are from description + parent + grandparent (used for soft scoring).
            *short* is True when the description is thin/generic (few alpha tokens).
        """
        description = item.get("description", "") or ""
        parent = item.get("parent", "") or ""
        grandparent = item.get("grandparent", "") or ""
        category_path = item.get("category_path", "") or ""

        generic = is_generic_item(description)

        # ── Truncate overly long context ────────────────────────────────
        # BOQ parents sometimes include verbose contractual clauses like
        # "include Pumps, and all necessary equipment and material; as per
        # the specifications and drawings." which add noise tokens that
        # pollute both TF-IDF scoring and sheet routing.  We keep only
        # the first ~120 chars (≈ meaningful first clause).
        MAX_CONTEXT_CHARS = 120
        if len(parent) > MAX_CONTEXT_CHARS:
            # Try to cut at a natural boundary (;, ,, or space)
            cut = parent[:MAX_CONTEXT_CHARS]
            for sep in (";", ",", " "):
                idx = cut.rfind(sep)
                if idx > MAX_CONTEXT_CHARS // 2:
                    cut = cut[:idx]
                    break
            parent = cut.strip()
        if len(grandparent) > MAX_CONTEXT_CHARS:
            cut = grandparent[:MAX_CONTEXT_CHARS]
            for sep in (";", ",", " "):
                idx = cut.rfind(sep)
                if idx > MAX_CONTEXT_CHARS // 2:
                    cut = cut[:idx]
                    break
            grandparent = cut.strip()

        # ── Context injection: enrich short/generic descriptions ────────
        # Fix 3: Use alpha-token count instead of char length to detect
        # "thin" descriptions like "25 mm diameter (Horizontal)" (27 chars
        # but only 1 meaningful alpha token: "diameter").  These need
        # parent context (e.g. "uPVC pipes") to route correctly.
        desc_toks = tokenize(description)
        desc_alpha = self._alpha_tokens(set(desc_toks))
        short = len(desc_alpha) <= 2 or len(desc_toks) <= 2
        enriched_description = description
        if short and parent:
            # Strip continuation markers before injecting
            ctx = re.sub(r"\(cont['’]?d\)", "", parent, flags=re.I).strip()
            if ctx and ctx.lower() != description.lower():
                enriched_description = f"{ctx} ; {description}"
        if short and not parent and grandparent:
            ctx = re.sub(r"\(cont['’]?d\)", "", grandparent, flags=re.I).strip()
            if ctx and ctx.lower() != description.lower():
                enriched_description = f"{ctx} ; {description}"
        # Also enrich when description has no alpha tokens at all
        # (pure dimension like "500 x 200 mm") - always pull context
        if not desc_alpha and parent:
            ctx = re.sub(r"\(cont['’]?d\)", "", parent, flags=re.I).strip()
            if ctx:
                enriched_description = f"{ctx} ; {description}"
                short = True

        parts: List[Tuple[str, float]] = []
        if enriched_description:
            parts.append((enriched_description, 3.4))

        # Upweight context when description is vague
        if generic or short:
            if parent:
                parts.append((parent, 2.5))
            if grandparent:
                parts.append((grandparent, 1.8))
        else:
            if parent:
                parts.append((parent, 1.9))
            if grandparent:
                parts.append((grandparent, 1.2))

        # Broader hierarchy context — category_path contains the full
        # BOQ hierarchy (e.g. "CONCRETE WORK > POURED CONCRETE > In-situ")
        # which carries strong classification signal for sheet routing.
        if category_path:
            parts.append((category_path, 1.2))

        qweights: Dict[str, float] = defaultdict(float)
        for text, w in parts:
            for tok in set(tokenize(text)):
                if tok in self.idf:
                    qweights[tok] += w * self.idf[tok]

        all_text = " ; ".join(x for x, _ in parts if x)
        # Hard spec filter uses ONLY the original description (never
        # the parent/grandparent context).  This prevents a parent like
        # "12 MPa" from rejecting all refs when the rate book only has
        # 10, 15, 20, … MPa grades.
        desc_specs = extract_specs(description)
        ctx_specs = extract_specs(all_text)
        guessed = infer_discipline_from_query(description, parent, grandparent)

        distinctive = set(tokenize(enriched_description))
        alpha_dist = self._alpha_tokens(distinctive)
        # Merge airfield hints from broader context
        all_toks = set(tokenize(all_text))
        alpha_dist = alpha_dist | (AIRFIELD_HINTS & all_toks)

        return dict(qweights), desc_specs, ctx_specs, guessed, distinctive, alpha_dist, short

    def _spec_score(
        self,
        qspecs: Dict[str, Tuple[str, ...]],
        ref: sqlite3.Row,
        *,
        has_object_support: bool,
    ) -> float:
        """Additive spec score for all specs.

        Key differentiating specs are ALSO scored multiplicatively in
        _spec_multiplier.  The additive part provides a baseline signal
        that helps break ties, while the multiplicative part compounds
        across specs for strong differentiation.
        """
        score = 0.0
        scale = 1.0 if has_object_support else 0.25
        pen_scale = 1.0 if has_object_support else 0.55
        # ── Fix 1b: merge dia ↔ dn for soft scoring too ───────────
        dn_vals = self._csv_to_set(ref["dn_csv"])
        dia_vals = self._csv_to_set(ref["dia_csv"])
        merged_pipe_size = dn_vals | dia_vals
        spec_map = {
            "dn":    (3.0, merged_pipe_size),
            "dia":   (2.2, merged_pipe_size),
            "mm":    (1.1, self._csv_to_set(ref["mm_csv"])),
            "mpa":   (2.5, self._csv_to_set(ref["mpa_csv"])),
            "kv":    (2.2, self._csv_to_set(ref["kv_csv"])),
            "dims":  (1.8, self._csv_to_set(ref["dims_csv"])),
            "mm2":   (2.0, self._csv_to_set(ref["mm2_csv"])),
            "cores": (1.2, self._csv_to_set(ref["core_csv"])),
            "thk":   (2.0, self._csv_to_set(ref.get("thk_csv", ""))),
            "fire":  (2.5, self._csv_to_set(ref.get("fire_csv", ""))),
            "pn":    (2.2, self._csv_to_set(ref.get("pn_csv", ""))),
        }
        for key, (bonus, rvals) in spec_map.items():
            qvals = set(qspecs.get(key, ()))
            if qvals and rvals:
                if qvals & rvals:
                    score += bonus * scale
                else:
                    score -= bonus * 0.55 * pen_scale

        # Secondary categorical specs (not in multiplier)
        ref_id = ref.get("ref_id")
        if ref_id is not None:
            rcat = self._get_ref_cat_specs(ref_id)
        else:
            rcat = {}

        # Categorical specs – additive supplement to multiplicative scoring.
        _CAT_WEIGHTS = {
            "pipe_mat":       (3.5, 2.5),
            "valve_type":     (3.5, 2.5),
            "fitting_type":   (2.5, 1.5),
            "concrete_elem":  (3.0, 2.0),
            "concrete_scope": (2.5, 1.5),
            "steel_section":  (2.5, 1.5),
            "coating":        (2.0, 1.0),
            "insul_mat":      (2.0, 1.0),
            "cable_insul":    (2.5, 1.5),
            "schedule":       (2.2, 1.2),
            "pipe_grade":     (2.5, 1.5),
            "stiffness":      (2.0, 1.0),
            "bend_angle":     (2.0, 1.0),
            "stc":            (1.5, 0.8),
        }
        for key, (bonus, penalty) in _CAT_WEIGHTS.items():
            qvals = set(qspecs.get(key, ()))
            rvals = set(rcat.get(key, ()))
            if qvals and rvals:
                if qvals & rvals:
                    score += bonus * scale
                else:
                    score -= penalty * pen_scale

        return score

    def _spec_multiplier(
        self,
        desc_specs: Dict[str, Tuple[str, ...]],
        ctx_specs: Dict[str, Tuple[str, ...]],
        ref: sqlite3.Row,
    ) -> float:
        """Multiplicative spec scoring for key differentiating specs.

        Uses desc_specs (description only) for DB-column numeric specs
        so that a parent-level "C32/40" doesn't over-boost concrete refs
        for formwork/rebar items.  Uses ctx_specs (full context) for
        categorical specs where parent context is reliable (e.g.
        concrete_elem from "Raft slab" description).
        """
        factor = 1.0

        # ── DB-column numeric specs (from desc_specs only) ─────────────
        _DB_SPEC = {
            "mpa":   ("mpa_csv",  1.40, 0.50),
            "dn":    ("dn_csv",   1.35, 0.45),
            "dia":   ("dia_csv",  1.30, 0.50),
            "kv":    ("kv_csv",   1.30, 0.50),
            "mm2":   ("mm2_csv",  1.25, 0.55),
            "cores": ("core_csv", 1.20, 0.60),
            "pn":    ("pn_csv",   1.25, 0.55),
        }
        for key, (col, match_f, mismatch_f) in _DB_SPEC.items():
            qvals = set(desc_specs.get(key, ()))
            if not qvals:
                continue
            rvals = self._csv_to_set(ref.get(col, ""))
            if key == "dia":
                rvals = rvals | self._csv_to_set(ref.get("dn_csv", ""))
            elif key == "dn":
                rvals = rvals | self._csv_to_set(ref.get("dia_csv", ""))
            if not rvals:
                # Generic ref (no spec value) — small penalty so
                # specific-matching refs rank above generics.
                factor *= 0.88
                continue
            if qvals & rvals:
                factor *= match_f
            else:
                factor *= mismatch_f

        # ── Categorical specs (from ctx_specs – parent context OK) ────
        ref_id = ref.get("ref_id")
        rcat = self._get_ref_cat_specs(ref_id) if ref_id is not None else {}

        _CAT_SPEC = {
            "concrete_elem":  (1.45, 0.70, 0.85),
            "pipe_mat":       (1.40, 0.55, 0.85),
            "valve_type":     (1.40, 0.55, 0.85),
            "cable_insul":    (1.30, 0.60, 0.90),
            "schedule":       (1.25, 0.65, 0.92),
            "pipe_grade":     (1.30, 0.60, 0.90),
        }
        for key, (match_f, mismatch_f, missing_f) in _CAT_SPEC.items():
            qvals = set(ctx_specs.get(key, ()))
            if not qvals:
                continue
            rvals = set(rcat.get(key, ()))
            if not rvals:
                factor *= missing_f
            elif qvals & rvals:
                factor *= match_f
            else:
                factor *= mismatch_f

        return factor

    def _passes_hard_spec_filters(
        self, qspecs: Dict[str, Tuple[str, ...]], ref: sqlite3.Row
    ) -> bool:
        # Only apply HARD rejection for unambiguous spec types.
        # dims and mm are too noisy (thickness, spacing, etc.) so they
        # are scored softly in _spec_score instead.
        checks = [
            ("dn",    "dn_csv",   False),
            ("dia",   "dia_csv",  False),
            ("mpa",   "mpa_csv",  False),
            ("kv",    "kv_csv",   False),
            ("mm2",   "mm2_csv",  True),   # subset check
            ("cores", "core_csv", False),
            ("pn",    "pn_csv",   False),   # pressure class
        ]
        for key, col, is_subset in checks:
            qvals = set(qspecs.get(key, ()))
            if not qvals:
                continue
            rvals = self._csv_to_set(ref.get(col, ""))
            # ── Fix 1: dia ↔ dn cross-reference ─────────────────────
            # BOQ "25 mm diameter" → dia=(25,) but rate-book stores
            # "DN25" → dn_csv=[25], dia_csv=[].  They mean the same
            # pipe size, so merge both columns before checking.
            if key == "dia":
                rvals = rvals | self._csv_to_set(ref.get("dn_csv", ""))
            elif key == "dn":
                rvals = rvals | self._csv_to_set(ref.get("dia_csv", ""))
            if not rvals:
                # Ref has no value for this spec — allow through (generic ref).
                # Soft scoring (_spec_multiplier) handles the penalty.
                continue
            if is_subset:
                if not qvals.issubset(rvals):
                    return False
            else:
                if not (qvals & rvals):
                    return False

        # ── Hard categorical spec filters (on-the-fly) ─────────────────
        # Pipe material and valve type are unambiguous differentiators:
        #   BOQ says "HDPE pipe" → must NOT match DI/uPVC/CS refs.
        #   BOQ says "gate valve" → must NOT match butterfly/check refs.
        # We only reject when BOTH query and ref have values (avoids
        # false-rejection when the ref description is too generic).
        ref_id = ref.get("ref_id")
        if ref_id is not None:
            rcat = self._get_ref_cat_specs(ref_id)
            _HARD_CAT_KEYS = ("pipe_mat", "valve_type")
            for key in _HARD_CAT_KEYS:
                qvals = set(qspecs.get(key, ()))
                if not qvals:
                    continue
                rvals = set(rcat.get(key, ()))
                if not rvals:
                    continue  # ref is generic (no material/type mentioned)
                if not (qvals & rvals):
                    return False  # definite mismatch

        return True

    def _fetch_ref_rows(
        self, ref_ids: List[int]
    ) -> Dict[int, Dict[str, Any]]:
        """Look up ref rows from the in-memory index (O(1) per id)."""
        return {rid: self._refs[rid] for rid in ref_ids if rid in self._refs}

    # ── Civil-concrete sections where suffix = Elem/Grade/Scope ────────
    _CONCRETE_SECTIONS: frozenset = frozenset({"31", "21", "11", "10"})

    def _diversity_dedup(
        self, filtered: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """Deduplicate candidates ensuring diversity across variant dimensions.

        Price codes follow ``[Disc][Section][Family] [V1][V2][V3]`` but the
        meaning of V1/V2/V3 CHANGES per section:

        **Civil Concrete** (C prefix, sections 31/21/11/10):
           V1=Element, V2=Grade, V3=Scope — original 3-level dedup is ideal.

        **All other sections** (masonry, MEP, electrical):
           V1/V2/V3 encode physical specs (material, size, thickness, rating).
           Treating V1 alone as "element key" is wrong — e.g. masonry has
           only 3 V1 values (height bands), so grouping by V1 would collapse
           480 codes into 3 buckets.  Instead we group by the code *prefix*
           (disc+section+family) to keep section diversity, then keep
           more siblings per group so spec-relevant variants survive.

        Strategy:
        - Detect concrete vs non-concrete from the discipline letter + section.
        - Concrete: 3-level dedup (element → grade → scope siblings).
        - Non-concrete: 2-level dedup (prefix → top siblings by score).
          A higher sibling cap lets the LLM see diverse spec variants.
        """
        if not filtered:
            return []

        _CODE_RE = re.compile(
            r'^([A-Za-z])\s+(\d+)\s+(\d+)\s+'  # disc, section, family + space
            r'([A-Za-z])'                        # V1
            r'([A-Za-z0-9])'                     # V2
            r'([A-Za-z0-9])$'                    # V3
        )
        _COMPACT_RE = re.compile(
            r'^([A-Za-z])(\d{2})(\d{2})'         # disc, section, family
            r'([A-Za-z])'                         # V1
            r'([A-Za-z0-9])'                      # V2
            r'([A-Za-z0-9])$'                     # V3
        )

        def _parse_full(price_code: str):
            """Return (disc, section, prefix, v1, v2, v3) or None."""
            code = price_code.strip()
            m = _CODE_RE.match(code)
            if m:
                disc, sec, fam, v1, v2, v3 = m.groups()
                prefix = f"{disc} {sec} {fam} "
                return disc.upper(), sec, prefix, v1, v2, v3
            m = _COMPACT_RE.match(code)
            if m:
                disc, sec, fam, v1, v2, v3 = m.groups()
                prefix = f"{disc}{sec}{fam}"
                return disc.upper(), sec, prefix, v1, v2, v3
            return None

        def _is_concrete(disc: str, sec: str) -> bool:
            return disc == "C" and sec in self._CONCRETE_SECTIONS

        # ────────────────────────────────────────────────────────────────
        # Partition candidates into concrete vs non-concrete
        # ────────────────────────────────────────────────────────────────
        concrete_items: List[Dict[str, Any]] = []
        other_items: List[Dict[str, Any]] = []

        for c in filtered:
            parsed = _parse_full(c.get("price_code", ""))
            if parsed and _is_concrete(parsed[0], parsed[1]):
                concrete_items.append(c)
            else:
                other_items.append(c)

        result: List[Dict[str, Any]] = []

        # ════════ CONCRETE PATH: original 3-level dedup ════════════════
        if concrete_items:
            MAX_GRADES_PER_ELEM = 3
            MAX_SIBLINGS_PER_GRADE = 8

            elem_best: Dict[str, float] = {}
            grade_groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

            for c in concrete_items:
                parsed = _parse_full(c.get("price_code", ""))
                if not parsed:
                    continue
                _, _, prefix, v1, v2, v3 = parsed
                elem_key = f"{prefix}{v1}"
                grade_stem = f"{prefix}{v1}{v2}"
                score = float(c["score"])
                if elem_key not in elem_best or score > elem_best[elem_key]:
                    elem_best[elem_key] = score
                grade_groups[grade_stem].append(c)

            # Select top element keys
            seen: set = set()
            selected_elems: List[str] = []
            for c in concrete_items:
                parsed = _parse_full(c.get("price_code", ""))
                if not parsed:
                    continue
                ek = f"{parsed[2]}{parsed[3]}"
                if ek not in seen:
                    seen.add(ek)
                    selected_elems.append(ek)
                    if len(selected_elems) >= self.max_candidates:
                        break

            # Map element → grades
            elem_to_grades: Dict[str, List[str]] = defaultdict(list)
            grade_best: Dict[str, float] = {}
            for gs, members in grade_groups.items():
                parsed = _parse_full(members[0].get("price_code", ""))
                if parsed:
                    ek = f"{parsed[2]}{parsed[3]}"
                    elem_to_grades[ek].append(gs)
                    grade_best[gs] = max(float(m["score"]) for m in members)

            for ek in selected_elems:
                grades = elem_to_grades.get(ek, [])
                grades.sort(key=lambda g: grade_best.get(g, 0), reverse=True)
                for gs in grades[:MAX_GRADES_PER_ELEM]:
                    siblings = grade_groups[gs]
                    siblings.sort(key=lambda x: float(x.get("score", 0)), reverse=True)
                    result.extend(siblings[:MAX_SIBLINGS_PER_GRADE])

        # ════════ NON-CONCRETE PATH: prefix-level dedup ════════════════
        # V1/V2/V3 carry spec info (size, material, rating) so we group
        # only by prefix (disc+section+family).  Within each prefix we
        # keep more siblings so the LLM sees the spec variety it needs
        # to pick the right DN, thickness, cross-section etc.
        if other_items:
            MAX_SIBLINGS_PER_PREFIX = 12
            MAX_PREFIXES = max(4, self.max_candidates // 2)

            prefix_groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
            prefix_best: Dict[str, float] = {}

            for c in other_items:
                parsed = _parse_full(c.get("price_code", ""))
                prefix = parsed[2] if parsed else c.get("price_code", "").strip()
                score = float(c["score"])
                prefix_groups[prefix].append(c)
                if prefix not in prefix_best or score > prefix_best[prefix]:
                    prefix_best[prefix] = score

            # Sheet diversity: cap prefixes per sheet
            MAX_PREFIXES_PER_SHEET = max(3, MAX_PREFIXES // 2)
            sheet_prefix_counts: Dict[str, int] = defaultdict(int)
            prefix_sheet: Dict[str, str] = {}
            for c in other_items:
                parsed = _parse_full(c.get("price_code", ""))
                pfx = parsed[2] if parsed else c.get("price_code", "").strip()
                if pfx not in prefix_sheet:
                    prefix_sheet[pfx] = c.get("sheet_name", "")

            seen_pfx: set = set()
            selected_pfx: List[str] = []
            for c in other_items:  # sorted by score desc
                parsed = _parse_full(c.get("price_code", ""))
                pfx = parsed[2] if parsed else c.get("price_code", "").strip()
                if pfx not in seen_pfx:
                    sheet = prefix_sheet.get(pfx, "")
                    if sheet_prefix_counts[sheet] >= MAX_PREFIXES_PER_SHEET:
                        continue
                    seen_pfx.add(pfx)
                    sheet_prefix_counts[sheet] += 1
                    selected_pfx.append(pfx)
                    if len(selected_pfx) >= MAX_PREFIXES:
                        break

            for pfx in selected_pfx:
                siblings = prefix_groups[pfx]
                siblings.sort(key=lambda x: float(x.get("score", 0)), reverse=True)
                result.extend(siblings[:MAX_SIBLINGS_PER_PREFIX])

        # Final sort by score so best candidates come first for the LLM
        result.sort(key=lambda x: float(x.get("score", 0)), reverse=True)
        return result

    _PIPELINE_SPEC_COLS = (
        ("dn_csv", "DN"), ("dia_csv", "Dia"), ("mpa_csv", "MPa"),
        ("kv_csv", "kV"), ("mm2_csv", "mm²"), ("core_csv", "Cores"),
        ("pn_csv", "PN"), ("thk_csv", "Thk"),
        ("pipe_mat_csv", "Material"), ("valve_type_csv", "Valve"),
        ("concrete_elem_csv", "Element"), ("cable_insul_csv", "Insulation"),
        ("schedule_csv", "Schedule"), ("pipe_grade_csv", "Grade"),
    )

    @staticmethod
    def _to_pipeline_format(candidates: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Convert internal ranking dicts to the format ``llm_match()`` expects."""
        result: List[Dict[str, Any]] = []
        for c in candidates:
            raw_desc = c["description"]
            # Deduplicate "X ; X" repetition from prefixed_description
            if " ; " in raw_desc:
                parts = raw_desc.split(" ; ", 1)
                if parts[0].strip() == parts[1].strip():
                    raw_desc = parts[0].strip()
            # Collect non-empty spec columns for the LLM
            specs: Dict[str, str] = {}
            for col, label in LexicalMatcher._PIPELINE_SPEC_COLS:
                val = c.get(col)
                if val and str(val).strip():
                    specs[label] = str(val).strip()
            entry: Dict[str, Any] = {
                "price_code": c["price_code"],
                "description": raw_desc,
                "leaf_description": c.get("leaf_description", ""),
                "category": c.get("sheet_name", ""),
                "score": c["score"],
                "metadata": {
                    "price_code": c["price_code"],
                    "description": raw_desc,
                    "source_file": c.get("source_file", ""),
                    "reference_sheet": c.get("sheet_name", ""),
                    "reference_category": c.get("sheet_name", ""),
                },
            }
            if specs:
                entry["specs"] = specs
            result.append(entry)
        return result
