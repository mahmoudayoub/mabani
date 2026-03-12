#!/usr/bin/env python3
"""
Pool recall diagnostic: checks whether GT price codes appear in the
initial scored_pool (before reranking) and in the final reranked candidates.

Usage:
    cd backend && .venv/bin/python -u ../scripts/pool_recall_analysis.py civil
    cd backend && .venv/bin/python -u ../scripts/pool_recall_analysis.py mechanical
"""

import asyncio, logging, sys, os, re, math
from pathlib import Path
from collections import defaultdict

BACKEND_DIR = Path(__file__).resolve().parent.parent / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

import boto3, pandas as pd
from openpyxl import load_workbook

logging.basicConfig(level=logging.WARNING, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
logger = logging.getLogger("pool_analysis")
logger.setLevel(logging.INFO)

S3_BUCKET = "pricecodestack-pricecodedata88b02d08-ciqjcb0pjn80"
S3_DB_KEY = "metadata/pricecode_index.db"
LOCAL_DB_PATH = "/tmp/pricecode_index.db"

BASE_DIR = Path(__file__).resolve().parent.parent

_PRESETS = {
    "mechanical": {
        "input":  BASE_DIR / "Astoria-Mechanical Commented.xlsx",
        "gt_col": "estimator",
    },
    "civil": {
        "input":  BASE_DIR / "Astoria trial 1 (Copy).xlsx",
        "gt_col": "agc",
    },
}

_COMPACT_RE = re.compile(r'^([A-Za-z])(\d{2})(\d{2})([A-Za-z][A-Za-z0-9]{1,2})$')

def _norm(code: str) -> str:
    """Normalize to spaced uppercase: p1316ACC -> P 13 16 ACC"""
    c = code.strip()
    m = _COMPACT_RE.match(c)
    if m:
        return f"{m.group(1).upper()} {m.group(2)} {m.group(3)} {m.group(4).upper()}"
    return c.upper()

def _family(code: str) -> str:
    parts = code.split()
    return " ".join(parts[:3]) if len(parts) >= 3 else code

preset_name = sys.argv[1] if len(sys.argv) > 1 else "civil"
if preset_name not in _PRESETS:
    print(f"Unknown preset. Choose from: {', '.join(_PRESETS)}")
    sys.exit(1)
preset = _PRESETS[preset_name]


def download_db():
    if Path(LOCAL_DB_PATH).exists():
        return
    s3 = boto3.client("s3")
    s3.download_file(S3_BUCKET, S3_DB_KEY, LOCAL_DB_PATH)


def load_ground_truth():
    wb = load_workbook(preset["input"], data_only=True)
    ws = wb.active
    hint = preset["gt_col"].lower()
    gt_col = gt_row = None
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=ws.max_column):
        for cell in row:
            if cell.value and hint in str(cell.value).lower():
                gt_col, gt_row = cell.column, cell.row
                break
        if gt_col:
            break
    if not gt_col:
        print("Could not find GT column")
        return {}
    gt = {}
    for row in ws.iter_rows(min_row=gt_row + 1, max_row=ws.max_row, min_col=gt_col, max_col=gt_col):
        cell = row[0]
        if cell.value:
            val = str(cell.value).strip()
            if val and val.lower() not in ("", "nan", "none", "n/a", "-"):
                gt[cell.row] = _norm(val)
    wb.close()
    return gt


async def run():
    download_db()
    from almabani.pricecode.lexical_search import LexicalMatcher

    matcher = await LexicalMatcher.create(
        db_path=LOCAL_DB_PATH,
        source_files=None,
        max_candidates=20,
    )

    # Check: does the GT code exist in the index at all?
    gt = load_ground_truth()
    logger.info(f"Loaded {len(gt)} GT entries")

    # Build a reverse lookup: normalized price_code -> ref_id
    code_to_ids = defaultdict(set)
    for rid, ref in matcher._refs.items():
        pc = _norm(ref["price_code"])
        code_to_ids[pc].add(rid)
    logger.info(f"Built code_to_ids for {len(code_to_ids)} unique codes from {len(matcher._refs)} refs")

    # Parse items
    from almabani.parsers.excel_parser import ExcelParser
    from almabani.parsers.hierarchy_processor import HierarchyProcessor
    from almabani.pricecode.pipeline import PriceCodePipeline

    parser = ExcelParser()
    sheets_data = parser.excel_io.read_excel(str(preset["input"]))
    sheet_name = next(iter(sheets_data.keys()))
    df, header_row_idx = sheets_data[sheet_name]
    columns = parser.excel_io.detect_columns(df)

    dummy = PriceCodePipeline.__new__(PriceCodePipeline)
    dummy.excel_parser = parser
    dummy.hierarchy_processor = HierarchyProcessor()
    parent_map = dummy._build_parent_map(df, header_row_idx, columns)
    items = dummy._extract_items_for_allocation(df, header_row_idx, columns, parent_map)

    # Map row_index+1 -> item (to match GT's Excel 1-based rows)
    row_to_item = {item.get("row_index", -1) + 1: item for item in items}

    # Stats
    gt_not_in_index = 0
    gt_in_index_not_in_pool = 0
    gt_in_pool = defaultdict(int)  # rank bucket
    gt_in_final = defaultdict(int)
    family_in_pool_exact_not = 0
    
    details_not_in_pool = []   # (row, gt_code, desc, reason)
    details_in_pool = []       # (row, gt_code, pool_rank, final_rank)

    # Limit analysis to items that have GT AND are in our parsed items
    count = 0
    total = len(gt)
    
    import time as _time
    
    for row_num, gt_code in sorted(gt.items()):
        item = row_to_item.get(row_num)
        if not item:
            continue
        count += 1
        
        gt_ids = code_to_ids.get(gt_code, set())
        
        if not gt_ids:
            gt_not_in_index += 1
            details_not_in_pool.append((row_num, gt_code, item.get("description", "")[:60], "NOT_IN_INDEX"))
            continue

        # Run the query to get scored_pool (internal)
        item_dict = {
            "description": item.get('description', ''),
            "parent": item.get('parent'),
            "grandparent": item.get('grandparent'),
            "unit": item.get('unit'),
            "item_code": item.get('item_code'),
            "category_path": item.get('category_path'),
        }
        
        qweights, desc_specs, ctx_specs, guessed_disc, distinctive, alpha_dist, short = matcher._weighted_query(item_dict)
        if not qweights:
            details_not_in_pool.append((row_num, gt_code, item.get("description", "")[:60], "NO_QUERY"))
            continue

        # Replay posting lookup to get raw scored_pool
        scored_pool = defaultdict(float)
        ordered = sorted(qweights.items(), key=lambda kv: kv[1], reverse=True)[:matcher.MAX_QUERY_TERMS]
        
        _spec_tok_set = set()
        for _sk in ("mpa", "dn", "dia", "kv", "mm2", "cores", "pn", "dims", "mm"):
            for _sv in desc_specs.get(_sk, ()):
                _spec_tok_set.add(_sv)

        for tok, qscore in ordered:
            df_count = matcher.df.get(tok, 0)
            if not df_count:
                continue
            _is_spec = tok in _spec_tok_set
            if not _is_spec and matcher.ref_count and (df_count / matcher.ref_count) > 0.08:
                continue
            damp = 1.0 / (1.0 + math.log(df_count + 1))
            if _is_spec:
                damp = max(damp, 0.25)
            posting_ids = matcher._postings.get(tok, ())
            for rid in posting_ids[:matcher.HARD_POSTINGS_LIMIT]:
                if matcher._valid_ref_ids is not None and rid not in matcher._valid_ref_ids:
                    continue
                scored_pool[rid] += qscore * damp

        # Check if GT ref_ids are in pool
        pool_rank = None
        if scored_pool:
            sorted_pool = sorted(scored_pool.items(), key=lambda x: x[1], reverse=True)
            pool_id_order = [rid for rid, _ in sorted_pool]
            for gt_rid in gt_ids:
                if gt_rid in scored_pool:
                    rank = pool_id_order.index(gt_rid) + 1
                    if pool_rank is None or rank < pool_rank:
                        pool_rank = rank

        # Also check final candidates
        candidates = matcher.search_sync(item_dict)
        final_rank = None
        for i, cand in enumerate(candidates):
            cand_code = _norm(cand.get("price_code", ""))
            if cand_code == gt_code:
                final_rank = i + 1
                break
        # Also check family match in final
        gt_fam = _family(gt_code)
        final_fam_rank = None
        for i, cand in enumerate(candidates):
            cand_fam = _family(_norm(cand.get("price_code", "")))
            if cand_fam == gt_fam:
                final_fam_rank = i + 1
                break

        if pool_rank is not None:
            bucket = "1" if pool_rank == 1 else "2-10" if pool_rank <= 10 else "11-50" if pool_rank <= 50 else "51-200" if pool_rank <= 200 else "201-1000" if pool_rank <= 1000 else "1000+"
            gt_in_pool[bucket] += 1
            details_in_pool.append((row_num, gt_code, pool_rank, final_rank, final_fam_rank))
        else:
            gt_in_index_not_in_pool += 1
            # Check WHY: which tokens overlap?
            gt_rid = next(iter(gt_ids))
            ref = matcher._refs.get(gt_rid)
            if ref:
                ref_desc = ref.get("prefixed_description", "") or ""
                from almabani.pricecode.lexical_search import tokenize
                ref_tokens = set(tokenize(ref_desc))
                query_tokens = set(qweights.keys())
                overlap = ref_tokens & query_tokens
                details_not_in_pool.append((
                    row_num, gt_code,
                    item.get("description", "")[:60],
                    f"IN_INDEX_NOT_IN_POOL overlap={len(overlap)} ref_toks={len(ref_tokens)} query_toks={len(query_tokens)} shared: {sorted(overlap)[:10]}"
                ))
            else:
                details_not_in_pool.append((row_num, gt_code, item.get("description", "")[:60], "IN_INDEX_NOT_IN_POOL (no ref)"))
        
        if count % 50 == 0:
            logger.info(f"  analyzed {count}/{total}")

    # Check final rank distribution
    for row_num, gt_code, pool_rank, final_rank, final_fam_rank in details_in_pool:
        if final_rank:
            bucket = "1" if final_rank == 1 else "2-5" if final_rank <= 5 else "6-10" if final_rank <= 10 else "11-20" if final_rank <= 20 else "20+"
            gt_in_final[bucket] += 1
        else:
            gt_in_final["NOT_IN_FINAL"] += 1

    # Report
    print("=" * 70)
    print(f"POOL RECALL ANALYSIS — {preset_name.upper()}")
    print("=" * 70)
    print(f"Items with GT analyzed: {count}")
    print()
    print("1. GT code existence in index:")
    print(f"   NOT in index at all:           {gt_not_in_index}")
    print(f"   In index, NOT in raw pool:     {gt_in_index_not_in_pool}")
    in_pool_total = sum(gt_in_pool.values())
    print(f"   In raw pool (scored_pool):     {in_pool_total}")
    print()
    print("2. Raw pool rank distribution (before reranking):")
    for bucket in ["1", "2-10", "11-50", "51-200", "201-1000", "1000+"]:
        c = gt_in_pool.get(bucket, 0)
        print(f"   Rank {bucket:>10s}: {c:>4d}  ({c/max(count,1)*100:.1f}%)")
    print()
    print("3. Final candidate rank (after reranking + filtering):")
    for bucket in ["1", "2-5", "6-10", "11-20", "20+", "NOT_IN_FINAL"]:
        c = gt_in_final.get(bucket, 0)
        print(f"   Rank {bucket:>14s}: {c:>4d}  ({c/max(count,1)*100:.1f}%)")
    print()
    
    # Show items NOT in pool
    if details_not_in_pool:
        print(f"--- Items where GT is NOT in pool ({len(details_not_in_pool)}): ---")
        for row, code, desc, reason in details_not_in_pool[:30]:
            print(f"  Row {row}: {code} | {desc}")
            print(f"    -> {reason}")
        if len(details_not_in_pool) > 30:
            print(f"  ... and {len(details_not_in_pool) - 30} more")
    print()

    # Show items in pool but not in final top-20
    not_in_final = [(r, c, pr, fr, ffr) for r, c, pr, fr, ffr in details_in_pool if fr is None]
    if not_in_final:
        print(f"--- In pool but NOT in final top-20 ({len(not_in_final)}): ---")
        for row, code, pool_rank, _, final_fam_rank in not_in_final[:30]:
            fam_note = f"fam@{final_fam_rank}" if final_fam_rank else "fam=MISS"
            print(f"  Row {row}: {code} | pool_rank={pool_rank} | {fam_note}")
        if len(not_in_final) > 30:
            print(f"  ... and {len(not_in_final) - 30} more")
    print()

    # Show items in pool at rank>1 but final rank=1 (reranking helped)
    rerank_helped = [(r, c, pr, fr) for r, c, pr, fr, _ in details_in_pool if fr == 1 and pr and pr > 1]
    if rerank_helped:
        print(f"--- Reranking promoted to rank 1 (was deeper in pool): {len(rerank_helped)} ---")
    
    # Items where pool has it but reranking pushed it down
    rerank_hurt = [(r, c, pr, fr) for r, c, pr, fr, _ in details_in_pool if fr and fr > 5 and pr and pr <= 5]
    if rerank_hurt:
        print(f"--- Reranking hurt (pool top-5 -> final rank>5): {len(rerank_hurt)} ---")
        for row, code, pr, fr in rerank_hurt[:20]:
            print(f"  Row {row}: {code} | pool={pr} -> final={fr}")

    print("=" * 70)

asyncio.run(run())
