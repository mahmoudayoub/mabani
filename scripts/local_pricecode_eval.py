#!/usr/bin/env python3
"""
Local price-code evaluation script.

Downloads the SQLite lexical index from S3, parses the input Excel file,
runs lexical search (top-K candidates, NO LLM), and writes a filled
output sheet.  Then computes recall stats against the ground-truth
"AGC Comments" column from the previous pipeline output.

Usage:
    cd backend && python -m scripts.local_pricecode_eval   # or just:
    cd /home/ali/Desktop/ammar/Almabani && python scripts/local_pricecode_eval.py
"""

import asyncio
import logging
import sys
import os
from pathlib import Path
from collections import defaultdict

# Ensure backend is on sys.path
BACKEND_DIR = Path(__file__).resolve().parent.parent / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

import boto3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
logger = logging.getLogger("local_eval")

# ─── Config ────────────────────────────────────────────────────────────
S3_BUCKET = "pricecodestack-pricecodedata88b02d08-ciqjcb0pjn80"
S3_DB_KEY = "metadata/pricecode_index.db"
LOCAL_DB_PATH = "/tmp/pricecode_index.db"

TOP_K = 20  # number of candidates to write per item (match production max_candidates)

BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_FILE = BASE_DIR / "Astoria trial 1 (Copy).xlsx"
GT_FILE = BASE_DIR / "Astoria trial 1 (Copy)_pricecode (1).xlsx"
OUTPUT_FILE = BASE_DIR / "Astoria trial 1 (Copy)_search_top5.xlsx"


# ─── Step 1: Download DB from S3 ──────────────────────────────────────
def download_db():
    if Path(LOCAL_DB_PATH).exists():
        sz = Path(LOCAL_DB_PATH).stat().st_size
        logger.info(f"DB already cached locally ({sz:,} bytes): {LOCAL_DB_PATH}")
        return
    logger.info(f"Downloading {S3_DB_KEY} from s3://{S3_BUCKET} ...")
    s3 = boto3.client("s3")
    s3.download_file(S3_BUCKET, S3_DB_KEY, LOCAL_DB_PATH)
    sz = Path(LOCAL_DB_PATH).stat().st_size
    logger.info(f"Downloaded {sz:,} bytes → {LOCAL_DB_PATH}")


# ─── Step 2: Load index ───────────────────────────────────────────────
async def load_matcher():
    from almabani.pricecode.lexical_search import LexicalMatcher
    matcher = await LexicalMatcher.create(
        db_path=LOCAL_DB_PATH,
        source_files=None,  # no filter – search everything
        max_candidates=TOP_K,
    )
    return matcher


# ─── Step 3: Parse the Excel and extract items ────────────────────────
def parse_input():
    """
    Reuse the pipeline's parsing logic to extract items needing allocation.
    Returns (items, sheet_name, df, header_row_idx, columns).
    """
    from almabani.parsers.excel_parser import ExcelParser
    from almabani.parsers.hierarchy_processor import HierarchyProcessor
    from almabani.pricecode.pipeline import PriceCodePipeline

    parser = ExcelParser()
    sheets_data = parser.excel_io.read_excel(str(INPUT_FILE))
    sheet_name = next(iter(sheets_data.keys()))
    df, header_row_idx = sheets_data[sheet_name]
    columns = parser.excel_io.detect_columns(df)
    logger.info(f"Sheet: {sheet_name}, header_row_idx={header_row_idx}, columns={columns}")

    # Detect code column (same logic as pipeline)
    detected_code_col = None
    for col in df.columns:
        col_lower = str(col).lower()
        if 'price code' in col_lower:
            detected_code_col = col
            break
    if not detected_code_col and header_row_idx > 0:
        group_row = df.iloc[header_row_idx - 1]
        code_candidates = []
        for col_idx, col in enumerate(df.columns):
            col_lower = str(col).lower()
            if 'code' in col_lower and 'description' not in col_lower and 'item' not in col_lower:
                item_col_name = columns.get('item')
                if item_col_name and item_col_name == col:
                    continue
                code_candidates.append((col_idx, col))
        for col_idx, col in code_candidates:
            group_val = group_row.iloc[col_idx] if col_idx < len(group_row) else None
            if group_val is not None and not pd.isna(group_val):
                group_str = str(group_val).strip().lower()
                if 'pricing' in group_str or 'price' in group_str:
                    detected_code_col = col
                    break
            for scan_idx in range(col_idx - 1, max(col_idx - 6, -1), -1):
                if scan_idx < 0 or scan_idx >= len(group_row):
                    continue
                scan_val = group_row.iloc[scan_idx]
                if scan_val is not None and not pd.isna(scan_val):
                    scan_str = str(scan_val).strip().lower()
                    if 'pricing' in scan_str or 'price' in scan_str:
                        detected_code_col = col
                    break
            if detected_code_col:
                break
    if not detected_code_col:
        for col in df.columns:
            col_lower = str(col).lower()
            if 'code' in col_lower and 'description' not in col_lower and 'item' not in col_lower:
                item_col_name = columns.get('item')
                if item_col_name and item_col_name == col:
                    continue
                detected_code_col = col
                break
    if detected_code_col:
        columns['code'] = detected_code_col
        columns['code_col_position'] = list(df.columns).index(detected_code_col)
        logger.info(f"Code column: {detected_code_col}")

    # We need a dummy pipeline just for its helper methods
    dummy_pipeline = PriceCodePipeline.__new__(PriceCodePipeline)
    dummy_pipeline.excel_parser = parser
    dummy_pipeline.hierarchy_processor = HierarchyProcessor()
    parent_map = dummy_pipeline._build_parent_map(df, header_row_idx, columns)
    items = dummy_pipeline._extract_items_for_allocation(df, header_row_idx, columns, parent_map)
    logger.info(f"Extracted {len(items)} items for allocation")
    return items, sheet_name, df, header_row_idx, columns


# ─── Step 4: Search all items ─────────────────────────────────────────
def search_all(matcher, items):
    """
    Run search_sync for each item, return list of (item, candidates).
    """
    from almabani.pricecode.pipeline import PriceCodePipeline
    results = []
    for i, item in enumerate(items):
        item_dict = {
            "description": item.get('description', ''),
            "parent": item.get('parent'),
            "grandparent": item.get('grandparent'),
            "unit": item.get('unit'),
            "item_code": item.get('item_code'),
            "category_path": item.get('category_path'),
        }
        candidates = matcher.search_sync(item_dict)
        results.append((item, candidates))
        if (i + 1) % 20 == 0 or i == len(items) - 1:
            logger.info(f"Searched {i+1}/{len(items)} items")
    return results


# ─── Step 5: Load ground truth from previous output ───────────────────
def load_ground_truth():
    """
    Read the AGC Comments column (assumed to contain the correct price codes)
    from the previous pipeline output file, keyed by Excel row number (1-based).
    """
    if not GT_FILE.exists():
        logger.warning(f"Ground truth file not found: {GT_FILE}")
        return {}

    wb = load_workbook(GT_FILE, data_only=True)
    ws = wb.active

    # Find the "AGC Comments" column
    agc_col = None
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=ws.max_column):
        for cell in row:
            if cell.value and 'agc' in str(cell.value).lower() and 'comment' in str(cell.value).lower():
                agc_col = cell.column
                header_row = cell.row
                break
        if agc_col:
            break

    if not agc_col:
        # Try finding any column with "comment" 
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=ws.max_column):
            for cell in row:
                if cell.value and 'comment' in str(cell.value).lower():
                    agc_col = cell.column
                    header_row = cell.row
                    break
            if agc_col:
                break

    if not agc_col:
        logger.warning("Could not find AGC Comments column in ground truth file")
        return {}

    logger.info(f"Found AGC Comments at column {agc_col}, header row {header_row}")

    gt = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=agc_col, max_col=agc_col):
        cell = row[0]
        if cell.value and str(cell.value).strip():
            val = str(cell.value).strip()
            # Normalize: "C 31 13 CGA" format
            gt[cell.row] = val

    logger.info(f"Loaded {len(gt)} ground truth entries from {GT_FILE.name}")
    wb.close()
    return gt


# ─── Step 6: Write output Excel ───────────────────────────────────────
def write_output(items_candidates, sheet_name, header_row_idx):
    """Write the candidates into new columns in the output workbook."""
    wb = load_workbook(INPUT_FILE)
    ws = wb[sheet_name]

    # Find the first free column
    start_col = ws.max_column + 1

    # Write headers for each candidate slot
    header_font = Font(bold=True)
    for k in range(TOP_K):
        code_col = start_col + k * 3
        desc_col = start_col + k * 3 + 1
        score_col = start_col + k * 3 + 2
        ws.cell(row=header_row_idx + 1, column=code_col, value=f"Cand{k+1}_Code").font = header_font
        ws.cell(row=header_row_idx + 1, column=desc_col, value=f"Cand{k+1}_Desc").font = header_font
        ws.cell(row=header_row_idx + 1, column=score_col, value=f"Cand{k+1}_Score").font = header_font

    # Write candidate data
    for item, candidates in items_candidates:
        row_idx = item['row_index'] + 1  # 1-based for openpyxl
        for k in range(min(TOP_K, len(candidates))):
            cand = candidates[k]
            code_col = start_col + k * 3
            desc_col = start_col + k * 3 + 1
            score_col = start_col + k * 3 + 2
            ws.cell(row=row_idx, column=code_col, value=cand.get('price_code', ''))
            ws.cell(row=row_idx, column=desc_col, value=cand.get('description', ''))
            ws.cell(row=row_idx, column=score_col, value=round(cand.get('score', 0), 3))

    wb.save(OUTPUT_FILE)
    wb.close()
    logger.info(f"Output written: {OUTPUT_FILE}")


# ─── Step 7: Compute stats ────────────────────────────────────────────
def compute_stats(items_candidates, gt):
    """
    Compute recall@K – is the ground truth code among the top K candidates?
    Also reports family match rate (first 4 code segments match).
    """
    if not gt:
        logger.warning("No ground truth available, skipping recall computation")
        return

    total_with_gt = 0
    recall_at = {1: 0, 3: 0, 5: 0, 10: 0, 20: 0}
    family_recall_at = {1: 0, 3: 0, 5: 0, 10: 0, 20: 0}  # first 3 segments match (e.g. C 31 13)
    super_family_recall = {1: 0, 3: 0, 5: 0, 10: 0, 20: 0}  # first 2 segments match (e.g. C 31)

    # Map from row_index (0-based df) to Excel row (1-based)
    # In the pipeline, row_index is the df index. Excel row = row_index + 1
    # But the GT file might have different row numbering due to headers.
    # The header_row_idx offset means: Excel row = header_row_idx + 1 (header) + (row_index - header_row_idx)
    # Actually item['row_index'] from pipeline = df row index, and
    # write_results uses row_idx = item['row_index'] + 1 for openpyxl (1-based).
    # So Excel row in output = item['row_index'] + 1.
    # The GT file should have the same row numbers.

    def normalize_code(code_str):
        """Normalize a price code to a canonical spaced format."""
        if not code_str:
            return ""
        code_str = str(code_str).strip().upper()
        # Already spaced format: "C 31 13 CGA"
        parts = code_str.split()
        if len(parts) >= 3:
            return " ".join(parts)
        return code_str

    def code_family(code_str, depth=3):
        """Extract the first `depth` segments of a spaced code."""
        parts = normalize_code(code_str).split()
        return " ".join(parts[:depth]) if len(parts) >= depth else normalize_code(code_str)

    missed_items = []
    found_items = []

    for item, candidates in items_candidates:
        excel_row = item['row_index'] + 1  # 1-based
        if excel_row not in gt:
            continue
        total_with_gt += 1
        gt_code = normalize_code(gt[excel_row])
        gt_fam3 = code_family(gt_code, 3)  # e.g. "C 31 13"
        gt_fam2 = code_family(gt_code, 2)  # e.g. "C 31"

        found_exact = False
        found_family = False
        found_super = False

        for k_cutoff in [1, 3, 5, 10, 20]:
            for cand in candidates[:k_cutoff]:
                cand_code = normalize_code(cand.get('price_code', ''))
                cand_fam3 = code_family(cand_code, 3)
                cand_fam2 = code_family(cand_code, 2)

                if cand_code == gt_code:
                    recall_at[k_cutoff] += 1
                    found_exact = True
                    break
                    
            for cand in candidates[:k_cutoff]:
                cand_code = normalize_code(cand.get('price_code', ''))
                cand_fam3 = code_family(cand_code, 3)
                if cand_fam3 == gt_fam3:
                    family_recall_at[k_cutoff] += 1
                    if k_cutoff == 5:
                        found_family = True
                    break

            for cand in candidates[:k_cutoff]:
                cand_code = normalize_code(cand.get('price_code', ''))
                cand_fam2 = code_family(cand_code, 2)
                if cand_fam2 == gt_fam2:
                    super_family_recall_at_k = True
                    super_family_recall[k_cutoff] += 1
                    if k_cutoff == 5:
                        found_super = True
                    break

        if not found_exact:
            top_codes = [normalize_code(c.get('price_code', '')) for c in candidates[:5]]
            missed_items.append({
                "row": excel_row,
                "desc": item.get('description', '')[:60],
                "gt": gt_code,
                "top5": top_codes,
                "family_match": found_family,
            })
        else:
            found_items.append({
                "row": excel_row,
                "gt": gt_code,
            })

    # ─── Print report ──────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print("LEXICAL SEARCH CANDIDATE RECALL STATS")
    print("=" * 70)
    print(f"Total items with ground truth: {total_with_gt}")
    print()

    for k in [1, 3, 5, 10, 20]:
        exact_pct = (recall_at[k] / total_with_gt * 100) if total_with_gt else 0
        fam_pct = (family_recall_at[k] / total_with_gt * 100) if total_with_gt else 0
        sup_pct = (super_family_recall[k] / total_with_gt * 100) if total_with_gt else 0
        print(f"Recall@{k}:")
        print(f"  Exact match:       {recall_at[k]:3d}/{total_with_gt} = {exact_pct:5.1f}%")
        print(f"  Family match (3):  {family_recall_at[k]:3d}/{total_with_gt} = {fam_pct:5.1f}%")
        print(f"  Super-family (2):  {super_family_recall[k]:3d}/{total_with_gt} = {sup_pct:5.1f}%")
        print()

    # ─── Error analysis: group missed items by GT family ───────────────
    if missed_items:
        print(f"\n--- Missed items (exact match not in top 5): {len(missed_items)} ---")
        family_groups = defaultdict(list)
        for m in missed_items:
            fam = code_family(m['gt'], 3)
            family_groups[fam].append(m)

        for fam in sorted(family_groups, key=lambda f: -len(family_groups[f])):
            items_in_fam = family_groups[fam]
            print(f"\n  GT family {fam} ({len(items_in_fam)} missed):")
            for m in items_in_fam[:3]:  # show first 3
                top_fams = [code_family(c, 3) for c in m['top5']]
                print(f"    Row {m['row']}: {m['desc']}")
                print(f"      GT={m['gt']}  top5_families={top_fams}")

    if found_items:
        print(f"\n--- Exact matches found in top 5: {len(found_items)} ---")
        for m in found_items[:10]:
            print(f"    Row {m['row']}: GT={m['gt']}")

    print("\n" + "=" * 70)


# ─── Main ──────────────────────────────────────────────────────────────
async def main():
    download_db()
    matcher = await load_matcher()
    items, sheet_name, df, header_row_idx, columns = parse_input()
    results = search_all(matcher, items)
    gt = load_ground_truth()
    write_output(results, sheet_name, header_row_idx)
    compute_stats(results, gt)


if __name__ == "__main__":
    asyncio.run(main())
