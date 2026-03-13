"""
Price Code Vector Pipeline – allocate price codes using S3 Vectors.

Reads a BOQ that needs price codes, embeds each item's description,
queries S3 Vectors with optional metadata filtering, and writes the
top-matched price code back into the Excel.
"""

import asyncio
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from almabani.core.embeddings import EmbeddingsService
from almabani.core.vector_store import VectorStoreService
from almabani.parsers.excel_parser import ExcelParser
from almabani.parsers.hierarchy_processor import HierarchyProcessor
from almabani.core.models import ItemType
from almabani.pricecode_vector.matcher import PriceCodeVectorMatcher

logger = logging.getLogger(__name__)

# Colour coding (same palette as existing pricecode service)
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
RED_FILL = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")


class PriceCodeVectorPipeline:
    """Allocate price codes to a BOQ using vector similarity search."""

    def __init__(
        self,
        embeddings_service: EmbeddingsService,
        vector_store: VectorStoreService,
        async_openai_client=None,
        model: str = None,
        top_k: int = 5,
        similarity_threshold: float = 0.40,
    ):
        self.embeddings = embeddings_service
        self.vector_store = vector_store
        self.top_k = top_k
        self.similarity_threshold = similarity_threshold
        self.excel_parser = ExcelParser()
        self.hierarchy_processor = HierarchyProcessor()

        # LLM matcher – used when an OpenAI client is provided
        self.llm_matcher: Optional[PriceCodeVectorMatcher] = None
        if async_openai_client is not None:
            self.llm_matcher = PriceCodeVectorMatcher(
                async_openai_client=async_openai_client,
                model=model,
            )

    # ── public API ──────────────────────────────────────────────────────

    async def process_file(
        self,
        input_file: Path,
        output_file: Optional[Path] = None,
        source_files: Optional[List[str]] = None,
        max_concurrent: int = 100,
    ) -> Dict[str, Any]:
        """
        Allocate price codes to *input_file* and write *output_file*.

        Parameters
        ----------
        source_files : list[str] | None
            Restrict the search to vectors whose ``source_file`` metadata
            matches one of these names.  ``None`` → search everything.
        max_concurrent : int
            Concurrent embedding + search tasks.
        """
        start_time = datetime.now()

        if output_file is None:
            output_file = input_file.parent / f"{input_file.stem}_pricecode_vector.xlsx"

        # ── 1. Parse Excel ──────────────────────────────────────────────
        logger.info(f"Reading {input_file} …")
        sheets_data = await asyncio.to_thread(
            self.excel_parser.excel_io.read_excel, str(input_file)
        )

        sheet_name = next(iter(sheets_data.keys()))
        df, header_row_idx = sheets_data[sheet_name]

        columns = self.excel_parser.excel_io.detect_columns(df)
        logger.info(f"Detected columns: {columns}")

        # Detect price-code output column
        code_col = self._detect_code_column(df, header_row_idx, columns)
        if code_col:
            columns["code"] = code_col
            columns["code_col_position"] = list(df.columns).index(code_col)
            logger.info(f"Price-code column: {code_col} (pos {columns['code_col_position']})")

        # ── 2. Build hierarchy & extract items ──────────────────────────
        parent_map = await asyncio.to_thread(
            self._build_parent_map, df, header_row_idx, columns
        )
        items = await asyncio.to_thread(
            self._extract_items_for_allocation, df, header_row_idx, columns, parent_map
        )
        logger.info(f"{len(items)} items to allocate")

        if not items:
            return {
                "total_items": 0,
                "matched": 0,
                "not_matched": 0,
                "output_file": str(output_file),
            }

        # ── 3. Build filter dict ────────────────────────────────────────
        filter_dict: Optional[Dict[str, Any]] = None
        if source_files:
            filter_dict = {"source_file": {"$in": source_files}}
            logger.info(f"Source-file filter: {source_files}")

        # ── 4. Concurrent embed + search ────────────────────────────────
        sem = asyncio.Semaphore(max_concurrent)
        _completed = 0

        async def process_item(item: Dict[str, Any]) -> Tuple[Dict[str, Any], Dict[str, Any]]:
            nonlocal _completed
            async with sem:
                text = self._build_embedding_text(item)
                embedding = await self.embeddings.generate_embedding_async(text)
                matches = await self.vector_store.search(
                    query_embedding=embedding,
                    top_k=self.top_k,
                    filter_dict=filter_dict,
                )

                result: Dict[str, Any]
                if matches and matches[0]["score"] >= self.similarity_threshold:
                    if self.llm_matcher is not None:
                        # ── LLM judge step ──────────────────────────────
                        result = await self.llm_matcher.match(item, matches)
                    else:
                        # ── Fallback: score-only (no LLM client) ────────
                        best = matches[0]
                        meta = best.get("metadata", {})
                        result = {
                            "matched": True,
                            "price_code": meta.get("price_code", ""),
                            "price_description": meta.get("description", ""),
                            "source_file": meta.get("source_file", ""),
                            "reference_sheet": meta.get("sheet_name", ""),
                            "score": best["score"],
                            "confidence_level": (
                                "EXACT" if best["score"] >= 0.92 else "HIGH"
                            ),
                        }
                else:
                    result = {
                        "matched": False,
                        "reason": (
                            f"Below threshold ({matches[0]['score']:.2f} < {self.similarity_threshold})"
                            if matches
                            else "No candidates found"
                        ),
                    }

                _completed += 1
                if _completed % 20 == 0 or _completed == len(items):
                    elapsed = (datetime.now() - start_time).total_seconds()
                    logger.info(
                        f"[progress] {_completed}/{len(items)} done ({elapsed:.0f}s)"
                    )
                return item, result

        tasks = [process_item(it) for it in items]
        results: List[Tuple[Dict[str, Any], Dict[str, Any]]] = await asyncio.gather(*tasks)

        elapsed = (datetime.now() - start_time).total_seconds()
        matched_results = [r for _, r in results if r["matched"]]
        matched_exact = sum(
            1 for r in matched_results if r.get("confidence_level") == "EXACT"
        )
        matched_high = sum(
            1 for r in matched_results if r.get("confidence_level") != "EXACT"
        )
        not_matched = sum(1 for _, r in results if not r["matched"])

        report = {
            "total_items": len(items),
            "matched": len(matched_results),
            "matched_exact": matched_exact,
            "matched_high": matched_high,
            "not_matched": not_matched,
            "match_rate": len(matched_results) / len(items) if items else 0,
            "output_file": str(output_file),
            "elapsed_seconds": elapsed,
            "items_per_second": len(items) / elapsed if elapsed > 0 else 0,
            "filters_used": source_files,
        }

        # ── 5. Write output Excel ──────────────────────────────────────
        await asyncio.to_thread(
            self._write_results,
            input_file,
            output_file,
            sheet_name,
            results,
            columns,
            header_row_idx,
            report,
        )

        # ── 6. Summary text file ────────────────────────────────────────
        summary_file = output_file.with_suffix(".txt")
        summary_text = self._generate_summary(input_file, output_file, sheet_name, report)
        summary_file.write_text(summary_text)
        report["summary_file"] = str(summary_file)

        logger.info(f"Allocation complete: {report}")
        return report

    # ── Parsing helpers (same as pricecode pipeline) ────────────────────

    def _build_parent_map(self, df, header_row_idx, columns):
        raw_items = self.excel_parser._extract_raw_items(df, columns, header_row_idx)
        tree = self.hierarchy_processor._build_hierarchy(raw_items)
        parent_map: Dict[int, Dict[str, Optional[str]]] = {}

        def walk(nodes, parent_desc, grandparent_desc, path):
            for node in nodes:
                np, ngp, added = parent_desc, grandparent_desc, False
                if node.item_type in (ItemType.NUMERIC_LEVEL, ItemType.SUBCATEGORY):
                    ngp = parent_desc
                    np = node.description
                    if node.description:
                        path.append(str(node.description))
                        added = True
                if node.item_type == ItemType.ITEM and node.row_number is not None:
                    parent_map[node.row_number] = {
                        "parent": np,
                        "grandparent": ngp,
                        "category_path": " > ".join(path) if path else None,
                    }
                if node.children:
                    walk(node.children, np, ngp, path)
                if added:
                    path.pop()

        walk(tree, None, None, [])
        return parent_map

    def _extract_items_for_allocation(self, df, header_row_idx, columns, parent_map):
        items: List[Dict[str, Any]] = []
        level_col = columns.get("level")
        item_col = columns.get("item")
        desc_col = columns.get("description")
        unit_col = columns.get("unit")
        code_col = columns.get("code")

        for idx in range(header_row_idx + 1, len(df)):
            row = df.iloc[idx]
            level_val = row.get(level_col) if level_col else None
            item_val = row.get(item_col) if item_col else None
            desc_val = row.get(desc_col) if desc_col else None
            unit_val = row.get(unit_col) if unit_col else None
            code_val = row.get(code_col) if code_col else None

            level_val = None if pd.isna(level_val) else level_val
            item_val = None if pd.isna(item_val) else item_val
            desc_val = None if pd.isna(desc_val) else desc_val
            unit_val = None if pd.isna(unit_val) else unit_val
            code_val = None if pd.isna(code_val) else code_val

            has_level = level_val is not None and str(level_val).strip() != ""
            has_item = item_val is not None and str(item_val).strip() != ""
            has_desc = desc_val is not None and str(desc_val).strip() != ""
            has_code = code_val is not None and str(code_val).strip() not in ("", "nan", "None")

            if (not has_level) and (has_item or has_desc) and not has_code:
                mk = idx + 1
                p = parent_map.get(mk, {})
                items.append({
                    "row_index": idx,
                    "item_code": str(item_val).strip() if item_val else "",
                    "description": str(desc_val).strip() if desc_val else "",
                    "unit": str(unit_val).strip() if unit_val else "",
                    "parent": p.get("parent"),
                    "grandparent": p.get("grandparent"),
                    "category_path": p.get("category_path"),
                })
        return items

    def _detect_code_column(self, df, header_row_idx, columns):
        for col in df.columns:
            if "price code" in str(col).lower():
                return col
        if header_row_idx > 0:
            group_row = df.iloc[header_row_idx - 1]
            for col_idx, col in enumerate(df.columns):
                cl = str(col).lower()
                if "code" in cl and "description" not in cl and "item" not in cl:
                    if columns.get("item") and columns["item"] == col:
                        continue
                    gv = group_row.iloc[col_idx] if col_idx < len(group_row) else None
                    if gv is not None and not pd.isna(gv):
                        gs = str(gv).strip().lower()
                        if "pricing" in gs or "price" in gs:
                            return col
                    for s in range(col_idx - 1, max(col_idx - 6, -1), -1):
                        if 0 <= s < len(group_row):
                            sv = group_row.iloc[s]
                            if sv is not None and not pd.isna(sv):
                                ss = str(sv).strip().lower()
                                if "pricing" in ss or "price" in ss:
                                    return col
                                break
        for col in df.columns:
            cl = str(col).lower()
            if "code" in cl and "description" not in cl and "item" not in cl:
                if columns.get("item") and columns["item"] == col:
                    continue
                return col
        return None

    @staticmethod
    def _build_embedding_text(item: Dict[str, Any]) -> str:
        parts: List[str] = []
        if item.get("category_path"):
            parts.append(f"[{item['category_path']}]")
        if item.get("description"):
            parts.append(item["description"])
        if item.get("unit"):
            parts.append(f"(Unit: {item['unit']})")
        return " ".join(parts)

    # ── Excel output ────────────────────────────────────────────────────

    def _write_results(
        self,
        input_file: Path,
        output_file: Path,
        sheet_name: str,
        results: List[Tuple[Dict[str, Any], Dict[str, Any]]],
        columns: Dict[str, str],
        header_row_idx: int,
        report: Dict[str, Any],
    ) -> None:
        logger.info(f"Writing output to {output_file}")
        wb = load_workbook(input_file)
        ws = wb[sheet_name]

        def _col_idx(name, pos_key=None):
            if pos_key and columns.get(pos_key) is not None:
                return columns[pos_key] + 1
            if not name:
                return None
            row = header_row_idx + 1
            for c in range(1, ws.max_column + 1):
                cv = ws.cell(row=row, column=c).value
                if cv and str(cv).strip() == str(name).strip():
                    return c
            return None

        code_col_idx = _col_idx(columns.get("code"), "code_col_position")

        # Add reference + reason columns
        ref_col = ws.max_column + 1
        reason_col = ref_col + 1
        score_col = reason_col + 1
        ws.cell(row=header_row_idx + 1, column=ref_col).value = "Reference"
        ws.cell(row=header_row_idx + 1, column=reason_col).value = "Reason"
        ws.cell(row=header_row_idx + 1, column=score_col).value = "Score"

        for item, result in results:
            row_idx = item["row_index"] + 1  # 1-based
            fill = RED_FILL
            if result["matched"]:
                fill = GREEN_FILL if result.get("confidence_level") == "EXACT" else YELLOW_FILL

            if code_col_idx and result.get("price_code"):
                cell = ws.cell(row=row_idx, column=code_col_idx)
                cell.value = result["price_code"]
                cell.fill = fill

            if result["matched"]:
                parts = []
                if result.get("source_file"):
                    parts.append(result["source_file"])
                if result.get("reference_sheet"):
                    parts.append(result["reference_sheet"])
                if result.get("price_description"):
                    parts.append(result["price_description"])
                ws.cell(row=row_idx, column=ref_col).value = " - ".join(parts)
                ws.cell(row=row_idx, column=reason_col).value = result.get("reason", "")
                ws.cell(row=row_idx, column=score_col).value = round(result.get("score", 0), 3)
            else:
                ws.cell(row=row_idx, column=reason_col).value = result.get("reason", "")
                if code_col_idx:
                    ws.cell(row=row_idx, column=code_col_idx).fill = fill

        wb.save(output_file)
        logger.info(f"Saved {output_file}")

    def _generate_summary(self, input_file, output_file, sheet_name, report):
        lines = [
            "=" * 60,
            "PRICE CODE VECTOR ALLOCATION SUMMARY",
            "=" * 60,
            f"  Input:      {input_file.name}",
            f"  Output:     {output_file.name}",
            f"  Sheet:      {sheet_name}",
            f"  Time:       {report['elapsed_seconds']:.1f}s",
            "",
            "RESULTS",
            f"  Total Items:     {report['total_items']}",
            f"  Matched:         {report['matched']}  ({report['match_rate']:.0%})",
            f"    Exact (green):   {report['matched_exact']}",
            f"    High  (yellow):  {report['matched_high']}",
            f"  Not Matched:     {report['not_matched']}",
            "",
        ]
        if report.get("filters_used"):
            lines.append(f"FILTERS: {', '.join(report['filters_used'])}")
        lines.append("=" * 60)
        return "\n".join(lines)
