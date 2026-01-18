"""
Price Code Pipeline - Process Excel files to allocate price codes.

Uses the same hierarchy parsing logic as RateFillerPipeline.
"""

import logging
import pandas as pd
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple
import asyncio
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .matcher import PriceCodeMatcher

# Reuse existing parsing components
from almabani.parsers.excel_parser import ExcelParser
from almabani.parsers.hierarchy_processor import HierarchyProcessor
from almabani.core.models import ItemType

logger = logging.getLogger(__name__)


class PriceCodePipeline:
    """
    Process Excel BOQ files to allocate price codes.
    
    Uses the same hierarchy extraction as RateFillerPipeline for consistency.
    """
    
    # Color coding
    GREEN_FILL = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Exact Match
    YELLOW_FILL = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid') # High Match
    RED_FILL = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # No match
    
    def __init__(self, matcher: PriceCodeMatcher):
        self.matcher = matcher
        self.excel_parser = ExcelParser()
        self.hierarchy_processor = HierarchyProcessor()
        
    def write_results(
        self,
        input_file: Path,
        output_file: Path,
        sheet_name: str,
        results: List[Tuple[Dict[str, Any], Dict[str, Any]]],
        columns: Dict[str, str],
        header_row_idx: int,
        report: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Write results to Excel with color coding and reference columns.
        """
        logger.info(f"Writing filled Excel: {output_file}")
        
        # Load workbook
        wb = load_workbook(input_file)
        ws = wb[sheet_name]
        
        # Determine column indices (1-based)
        def get_col_idx(name: str):
            if not name: return None
            # Find column by header value
            row = header_row_idx + 1
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val and str(cell_val).strip() == str(name).strip():
                    return col
            return None
            
        code_col_idx = get_col_idx(columns.get('code'))
        desc_col_idx = get_col_idx(columns.get('code_description'))
        
        # Create output columns if they don't exist
        start_col = ws.max_column + 1
        
        # Reference Column (Consolidated)
        ref_col_idx = start_col
        ws.cell(row=header_row_idx + 1, column=ref_col_idx).value = "Reference"
        
        # Reason Column
        reason_col_idx = start_col + 1
        ws.cell(row=header_row_idx + 1, column=reason_col_idx).value = "Reason"
        
        logger.info(f"Created Reference column at {ref_col_idx}, Reason column at {reason_col_idx}")
        
        # Write results
        for item, result in results:
            row_idx = item['row_index'] + 1  # 1-based
            
            # Determine fill color based on Match + Confidence
            fill = self.RED_FILL
            if result['matched']:
                confidence = result.get('confidence_level', 'HIGH') # Default to HIGH/Yellow if missing
                if str(confidence).upper() == 'EXACT':
                    fill = self.GREEN_FILL
                else:
                    fill = self.YELLOW_FILL
            
            # 1. Price Code
            if code_col_idx:
                cell = ws.cell(row=row_idx, column=code_col_idx)
                if result.get('price_code'):
                    cell.value = result['price_code']
                cell.fill = fill
            
            # 2. Price Description - DISABLED
            # if desc_col_idx and result.get('price_description'):
            #     cell = ws.cell(row=row_idx, column=desc_col_idx)
            #     cell.value = result['price_description']
            
            # 3. Reference fields (Consolidated)
            if result['matched']:
                # Format: Source - Sheet - Description
                # Note: source_file is the filename, reference_sheet is the tab name
                parts = []
                if result.get('source_file'):
                    parts.append(str(result['source_file']))
                if result.get('reference_sheet'):
                    parts.append(str(result['reference_sheet']))
                if result.get('price_description'):
                    parts.append(str(result['price_description']))
                
                ref_text = " - ".join(parts)
                ws.cell(row=row_idx, column=ref_col_idx).value = ref_text
            
            # 4. Reason Column (Only write if matched)
            if result['matched'] and result.get('reason'):
                ws.cell(row=row_idx, column=reason_col_idx).value = result['reason']
            
            # 5. If not matched, color the Item Code cell red if exists
            if not result['matched'] and columns.get('item'):
                input_code_idx = get_col_idx(columns.get('item'))
                if input_code_idx:
                     ws.cell(row=row_idx, column=input_code_idx).fill = self.RED_FILL
        
        if report:
             logger.info(f"Writing results, match rate: {report.get('match_rate', 0):.1%}")
            
        # Save
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        logger.info(f"Saved filled Excel to: {output_file}")

    def _write_summary_file(self, file_path: Path, content: str) -> None:
        """Write summary content to a text file."""
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
        logger.info(f"Summary written to: {file_path}")
    def _generate_summary(
        self,
        input_file: Path,
        output_file: Path,
        sheet_name: str,
        report: Dict[str, Any]
    ) -> str:
        """Generate a text summary of the processing results."""
        lines = []
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Header
        lines.append("=" * 70)
        lines.append("ALMABANI PRICE CODE ALLOCATION - PROCESSING SUMMARY")
        lines.append("=" * 70)
        lines.append("")
        
        # File info
        lines.append("FILE INFORMATION")
        lines.append("-" * 40)
        lines.append(f"  Input File:    {input_file.name}")
        lines.append(f"  Output File:   {output_file.name}")
        lines.append(f"  Sheet:         {sheet_name}")
        lines.append(f"  Generated:     {timestamp}")
        lines.append(f"  Processing Time: {report['elapsed_seconds']:.1f}s")
        lines.append("")

        # Filters Used
        filters = report.get('filters_used')
        if filters:
            lines.append("FILTERS USED")
            lines.append("-" * 40)
            lines.append(f"  Source Files: {', '.join(filters)}")
            lines.append("")
        elif report.get('filters_used') is None: # Explicitly check for None vs Empty
            pass # No filter logic invoked
        
        # Statistics
        lines.append("PROCESSING STATISTICS")
        lines.append("-" * 40)
        lines.append(f"  Total Items:     {report['total_items']}")
        lines.append(f"  Total Matched:   {report['matched']}")
        lines.append(f"    - Exact Match:   {report.get('matched_exact', 0)} (Green)")
        lines.append(f"    - High Conf:     {report.get('matched_high', 0)} (Yellow)")
        lines.append(f"  Not Matched:     {report['not_matched']} (Red)")
        if report.get('errors', 0) > 0:
            lines.append(f"  Errors:          {report['errors']}")
        lines.append("")
        
        # Success rate
        if report['total_items'] > 0:
            lines.append(f"  Match Rate:      {report['match_rate']:.1%}")
        
        lines.append("")
        
        # Error Details
        failed_items = report.get('failed_items')
        if failed_items:
            lines.append("ERROR DETAILS")
            lines.append("-" * 40)
            for item in failed_items[:50]: # Limit to 50 errors in summary
                 lines.append(f"  Row {item['row']}: {item['reason']}")
            if len(failed_items) > 50:
                 lines.append(f"  ... and {len(failed_items) - 50} more errors.")
            lines.append("")

        lines.append("=" * 70)
        lines.append("END OF SUMMARY")
        lines.append("=" * 70)
        
        return "\n".join(lines)
    
    def _build_parent_map(
        self,
        df,
        header_row_idx: int,
        columns: Dict[str, str]
    ) -> Dict[int, Dict[str, Optional[str]]]:
        """
        Build a mapping of row_index -> {'parent': ..., 'grandparent': ..., 'category_path': ...}
        using the same hierarchy logic as RateFillerPipeline.
        """
        raw_items = self.excel_parser._extract_raw_items(df, columns, header_row_idx)
        tree = self.hierarchy_processor._build_hierarchy(raw_items)
        
        parent_map: Dict[int, Dict[str, Optional[str]]] = {}
        
        def walk(nodes: List, parent_desc: Optional[str], grandparent_desc: Optional[str], path: List[str]):
            for node in nodes:
                node_parent = parent_desc
                node_grandparent = grandparent_desc
                path_added = False
                
                # Update lineage if this node is a parent type
                if node.item_type in (ItemType.NUMERIC_LEVEL, ItemType.SUBCATEGORY):
                    node_grandparent = parent_desc
                    node_parent = node.description
                    if node.description:
                        path.append(str(node.description))
                        path_added = True
                
                if node.item_type == ItemType.ITEM and node.row_number is not None:
                    parent_map[node.row_number] = {
                        'parent': node_parent,
                        'grandparent': node_grandparent,
                        'category_path': ' > '.join(path) if path else None
                    }
                
                # Recurse into children
                if node.children:
                    walk(node.children, node_parent, node_grandparent, path)
                
                # Pop path if we added for this node
                if path_added:
                    path.pop()
        
        walk(tree, None, None, [])
        return parent_map
    
    def _extract_items_for_allocation(
        self,
        df,
        header_row_idx: int,
        columns: Dict[str, str],
        parent_map: Dict[int, Dict[str, Optional[str]]]
    ) -> List[Dict[str, Any]]:
        """
        Extract items that need price code allocation.
        
        Logic: If Level is EMPTY and row contains an Item -> Process for allocation
        Same logic as RateFillerPipeline._extract_items_for_filling
        """
        items = []
        
        level_col = columns.get('level')
        item_col = columns.get('item')
        desc_col = columns.get('description')
        unit_col = columns.get('unit')
        code_col = columns.get('code')  # Price code column (output)
        
        for idx in range(header_row_idx + 1, len(df)):
            row = df.iloc[idx]
            
            level_val = row.get(level_col) if level_col else None
            item_val = row.get(item_col) if item_col else None
            desc_val = row.get(desc_col) if desc_col else None
            unit_val = row.get(unit_col) if unit_col else None
            code_val = row.get(code_col) if code_col else None
            
            # Clean None/NaN values
            level_val = None if pd.isna(level_val) else level_val
            item_val = None if pd.isna(item_val) else item_val
            desc_val = None if pd.isna(desc_val) else desc_val
            unit_val = None if pd.isna(unit_val) else unit_val
            code_val = None if pd.isna(code_val) else code_val
            
            # Extract items needing filling (Level is empty, has Item OR Description)
            has_level = level_val is not None and str(level_val).strip() != ''
            has_item_code = item_val is not None and str(item_val).strip() != ''
            has_desc = desc_val is not None and str(desc_val).strip() != ''
            
            # Skip if already has a price code
            has_existing_code = code_val is not None and str(code_val).strip() not in ['', 'nan', 'None']
            
            if (not has_level) and (has_item_code or has_desc) and not has_existing_code:
                parent = None
                grandparent = None
                category_path = None
                map_key = idx + 1  # parent_map uses 1-based row_number
                if map_key in parent_map:
                    parent = parent_map[map_key].get('parent')
                    grandparent = parent_map[map_key].get('grandparent')
                    category_path = parent_map[map_key].get('category_path')
                
                items.append({
                    'row_index': idx,
                    'item_code': str(item_val).strip() if item_val else '',
                    'description': str(desc_val).strip() if desc_val else '',
                    'unit': str(unit_val) if unit_val else '',
                    'parent': parent,
                    'grandparent': grandparent,
                    'category_path': category_path
                })
        
        return items
    
    def build_search_text(self, item: Dict[str, Any]) -> str:
        """
        Build the search text for vector search and LLM matching.
        
        Combines: description + parent context + unit
        This is what gets embedded and passed to the LLM.
        """
        parts = []
        
        # Add category path for context (e.g., "SITE DEMOLITION > EARTHWORK")
        if item.get('category_path'):
            parts.append(f"[{item['category_path']}]")
        
        # Add main description
        if item.get('description'):
            parts.append(item['description'])
        
        # Add unit for specification matching
        if item.get('unit'):
            parts.append(f"(Unit: {item['unit']})")
        
        return ' '.join(parts)
    
    async def process_file(
        self,
        input_file: Path,
        output_file: Optional[Path] = None,
        namespace: str = "",
        source_files: Optional[List[str]] = None,
        max_concurrent: int = None
    ) -> Dict[str, Any]:
        """
        Process an Excel file to allocate price codes.
        
        Args:
            source_files: Optional list of source files to filter by, e.g., ["AI Codes - Civil"]
                          If None, searches all indexed price codes.
            max_concurrent: Max concurrent matching tasks (from settings if None)
        """
        # Load max_concurrent from settings if not provided
        if max_concurrent is None:
            from almabani.config.settings import get_settings
            max_concurrent = get_settings().pricecode_max_concurrent
        
        start_time = datetime.now()
        
        if output_file is None:
            output_file = input_file.parent / f"{input_file.stem}_pricecode.xlsx"
        
        # Read Excel using excel_io (the I/O helper in ExcelParser)
        logger.info(f"Reading {input_file}...")
        sheets_data = await asyncio.to_thread(
            self.excel_parser.excel_io.read_excel, 
            str(input_file)
        )
        
        # Process first sheet
        sheet_name = next(iter(sheets_data.keys()))
        df, header_row_idx = sheets_data[sheet_name]
        
        # Detect columns using excel_io
        columns = self.excel_parser.excel_io.detect_columns(df)
        logger.info(f"Detected columns: {columns}")
        
        # Detect Code column for output
        # Hierarchy of checks:
        # 1. "Price Code" (Explicit)
        # 2. "Code" (Generic) - but verified to NOT be the Item column
        
        detected_code_col = None
        
        # 1. Strict textual match first
        for col in df.columns:
            col_lower = str(col).lower()
            if 'price code' in col_lower:
                detected_code_col = col
                break
        
        # 2. Fallback to generic "Code" if no explicit "Price Code" found
        if not detected_code_col:
            for col in df.columns:
                col_lower = str(col).lower()
                # Must contain 'code' but NOT 'description' or 'item' (unless it's just 'code')
                if 'code' in col_lower and 'description' not in col_lower and 'item' not in col_lower:
                    # HEURISTIC CHECK:
                    # Check first few non-null values. If they look like BOQ Item IDs (e.g. matches the Item column pattern), 
                    # we risk skipping specific rows. 
                    # However, user explicitly requested to consider "Code".
                    # We will assume "Code" is the target unless it's obviously the Item Numbering column.
                    
                    # If we already identified an 'item' column, ensure this 'code' column isn't practically identical
                    item_col_name = columns.get('item')
                    if item_col_name and item_col_name != col:
                         detected_code_col = col
                         break
                    elif not item_col_name:
                         # If no item column found yet, this might be it, but we are looking for Price Code destination.
                         # We'll take it as a candidate.
                         detected_code_col = col
                         break
        
        if detected_code_col:
            columns['code'] = detected_code_col
        
        # Detect output Description column (different from input description)
        # DISABLED: User requested not to fill Price Description
        # code_col_idx = list(df.columns).index(columns['code']) if columns.get('code') else None
        # for col in df.columns:
        #     col_lower = str(col).lower()
        #     if 'description' in col_lower and col != columns.get('description'):
        #         col_idx = list(df.columns).index(col)
        #         if code_col_idx is not None and col_idx > code_col_idx:
        #             columns['code_description'] = col
        #             break
        
        # Build parent map (same hierarchy logic as rate filler)
        parent_map = await asyncio.to_thread(
            self._build_parent_map, df, header_row_idx, columns
        )
        
        # Extract items for allocation
        items = await asyncio.to_thread(
            self._extract_items_for_allocation, df, header_row_idx, columns, parent_map
        )
        
        logger.info(f"Found {len(items)} items needing price code allocation")
        
        if not items:
            logger.warning("No items found for allocation")
            return {
                "total_items": 0,
                "matched": 0,
                "not_matched": 0,
                "output_file": str(output_file)
            }
        
        # Match items
        matched_count = 0
        not_matched_count = 0
        
        semaphore = asyncio.Semaphore(max_concurrent)
        
        # Build filter if source_files specified
        filter_dict = None
        if source_files:
            filter_dict = {"source_file": {"$in": source_files}}
            logger.info(f"Filtering by source files: {source_files}")
        
        from almabani.core.async_vector_store import get_async_vector_store
        
        async def process_item(item, vs):
            async with semaphore:
                # Pass structured context to matcher
                result = await self.matcher.match(
                    description=item.get('description', ''),
                    namespace=namespace,
                    filter_dict=filter_dict,
                    vector_store=vs,
                    # New context fields
                    parent=item.get('parent'),
                    grandparent=item.get('grandparent'),
                    unit=item.get('unit'),
                    item_code=item.get('item_code'),
                    category_path=item.get('category_path')
                )
                return item, result
        
        # Use shared vector store connection for all items
        results = []
        async with get_async_vector_store() as vector_store:
            tasks = [process_item(item, vector_store) for item in items]
            results = await asyncio.gather(*tasks)
        
        elapsed = (datetime.now() - start_time).total_seconds()
        
        # Build report dictionary first so it can be passed to write_results
        error_count = sum(1 for _, res in results if str(res.get('reason', '')).startswith('LLM error'))
        matched_results = [res for _, res in results if res['matched']]
        matched_exact = sum(1 for res in matched_results if str(res.get('confidence_level', '')).upper() == 'EXACT')
        matched_high = sum(1 for res in matched_results if str(res.get('confidence_level', '')).upper() != 'EXACT')
        
        report = {
            "total_items": len(items),
            "matched": len(matched_results),
            "matched_exact": matched_exact,
            "matched_high": matched_high,
            "not_matched": sum(1 for _, res in results if not res['matched']),
            "errors": error_count,
            "match_rate": len(matched_results) / len(items) if items else 0,
            "output_file": str(output_file),
            "elapsed_seconds": elapsed,
            "items_per_second": len(items) / elapsed if elapsed > 0 else 0,
            "filters_used": source_files,
            "failed_items": [
                {
                    "row": item['row_index'] + 1,
                    "reason": res.get('reason', 'Unknown error')
                }
                for item, res in results
                if not res.get('matched') and str(res.get('reason', '')).startswith('LLM error')
            ]
        }
        
        # Save output using rich Excel writer, passing the report
        await asyncio.to_thread(
            self.write_results,
            input_file,
            output_file,
            sheet_name,
            results,
            columns,
            header_row_idx,
            report  # Pass report for summary sheet 
        )
        
        # Generate and save summary file
        summary_file = output_file.with_suffix('.txt')
        summary_content = self._generate_summary(
            input_file=input_file,
            output_file=output_file,
            sheet_name=sheet_name,
            report=report
        )
        await asyncio.to_thread(self._write_summary_file, summary_file, summary_content)
        
        report["summary_file"] = str(summary_file)
        
        logger.info(f"Report: {report}")
        return report
