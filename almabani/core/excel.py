"""
Shared Excel I/O module - Reading and writing Excel files with BOQ data.
Consolidates functionality from all pipelines.
"""
import pandas as pd
import numpy as np
import logging
from pathlib import Path
from typing import Dict, Any, Optional, Tuple, List
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelIO:
    """Unified Excel reading and writing with BOQ-specific features."""
    
    # Color coding for filled rates
    GREEN_FILL = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Exact match
    YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Close match
    ORANGE_FILL = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Approximation
    RED_FILL = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Not filled
    
    def __init__(self):
        self.logger = logger
    
    # ==================== Reading ====================
    
    def read_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None
    ) -> Dict[str, Tuple[pd.DataFrame, int]]:
        """
        Read Excel file and return DataFrames with header row index.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet to read (None = all sheets)
            
        Returns:
            Dictionary of {sheet_name: (DataFrame, header_row_index)}
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        logger.info(f"Reading Excel file: {path.name}")
        
        # Read all sheets without assuming header location
        if sheet_name:
            raw_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            df, header_idx = self._process_sheet(raw_df)
            sheets = {sheet_name: (df, header_idx)}
            logger.info(f"Read sheet '{sheet_name}': {len(df)} rows, header at row {header_idx}")
        else:
            all_sheets = pd.read_excel(file_path, sheet_name=None, header=None)
            sheets = {}
            for name, raw_df in all_sheets.items():
                df, header_idx = self._process_sheet(raw_df)
                sheets[name] = (df, header_idx)
                logger.info(f"Read sheet '{name}': {len(df)} rows, header at row {header_idx}")
        
        return sheets
    
    def _process_sheet(self, raw_df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
        """
        Process raw sheet to detect header and preserve all rows.
        
        Args:
            raw_df: Raw DataFrame with no header
            
        Returns:
            Tuple of (processed DataFrame, header row index)
        """
        # Find header row
        header_idx = self._find_header_row(raw_df)
        
        if header_idx is None:
            logger.warning("No header row detected, using first row")
            header_idx = 0
        
        # Create DataFrame with proper header
        df = raw_df.copy()
        df.columns = raw_df.iloc[header_idx].values
        
        # Handle duplicate column names by renaming them
        cols = pd.Series(df.columns)
        for dup in cols[cols.duplicated()].unique():
            # Find indices of duplicate columns
            dup_indices = cols[cols == dup].index.tolist()
            # Rename duplicates with suffix
            for i, idx in enumerate(dup_indices[1:], start=2):
                cols.iloc[idx] = f"{dup}_{i}"
        df.columns = cols
        
        # Keep ALL rows including header (we'll need original structure)
        # But mark the data start
        df['_header_row_index'] = header_idx
        
        return df, header_idx
    
    def _find_header_row(self, raw_df: pd.DataFrame) -> Optional[int]:
        """
        Find the row containing column headers.
        Looks for keywords: Level, Item, Bill description, Unit, Rate
        """
        keywords = ['level', 'item', 'description', 'unit', 'rate']
        
        for idx in range(min(10, len(raw_df))):
            row_values = [str(val).strip().lower() for val in raw_df.iloc[idx].values 
                         if pd.notna(val) and str(val).strip()]
            
            # Count matches
            matches = sum(1 for keyword in keywords 
                         if any(keyword in val for val in row_values))
            
            if matches >= 3:
                logger.info(f"Detected header row at index {idx}")
                return idx
        
        return None
    
    def detect_columns(self, df: pd.DataFrame) -> Dict[str, str]:
        """
        Detect required column names.
        
        Returns:
            Dictionary mapping canonical names to actual column names
        """
        columns = {}
        col_names = [str(col).strip().lower() for col in df.columns]
        
        # Level column
        level_candidates = ['level', 'lvl']
        for candidate in level_candidates:
            for i, col in enumerate(col_names):
                if candidate in col:
                    columns['level'] = df.columns[i]
                    break
            if 'level' in columns:
                break
        
        # Item column
        item_candidates = ['item', 'item code']
        for candidate in item_candidates:
            for i, col in enumerate(col_names):
                if candidate in col and 'description' not in col:
                    columns['item'] = df.columns[i]
                    break
            if 'item' in columns:
                break
        
        # Description column
        desc_candidates = ['description', 'bill description', 'desc']
        for candidate in desc_candidates:
            for i, col in enumerate(col_names):
                if candidate in col:
                    columns['description'] = df.columns[i]
                    break
            if 'description' in columns:
                break
        
        # Unit column
        unit_candidates = ['unit', 'units', 'uom']
        for candidate in unit_candidates:
            for i, col in enumerate(col_names):
                if candidate in col:
                    columns['unit'] = df.columns[i]
                    break
            if 'unit' in columns:
                break
        
        # Rate column
        rate_candidates = ['rate', 'unit rate', 'price']
        for candidate in rate_candidates:
            for i, col in enumerate(col_names):
                if candidate in col:
                    columns['rate'] = df.columns[i]
                    break
            if 'rate' in columns:
                break
        
        # Optional columns
        for col in df.columns:
            col_lower = str(col).strip().lower()
            if 'trade' in col_lower:
                columns['trade'] = col
            elif 'code' in col_lower and 'item' not in col_lower:
                columns['code'] = col
            elif 'full description' in col_lower:
                columns['full_description'] = col
            elif 'reference' in col_lower:
                columns['reference'] = col
            elif 'reasoning' in col_lower:
                columns['reasoning'] = col
        
        return columns
    
    # ==================== Writing ====================
    
    def write_filled_excel(
        self,
        input_file: str,
        output_file: str,
        sheet_results: Dict[str, Dict[str, Any]]
    ) -> str:
        """
        Write filled Excel file with color coding.
        
        Args:
            input_file: Path to original Excel file
            output_file: Path to output Excel file
            sheet_results: Results from rate filling process
                {
                    'sheet_name': {
                        'header_row_index': int,
                        'columns': dict,
                        'filled_items': [
                            {
                                'row_index': int,
                                'item_code': str,
                                'description': str,
                                'filled_unit': str or None,
                                'filled_rate': float or None,
                                'status': 'filled' or 'not_filled',
                                'match_type': 'exact' or 'close' or 'approximation',
                                'reference': str,
                                'reasoning': str,
                                'confidence': float
                            }
                        ]
                    }
                }
        
        Returns:
            Path to output file
        """
        logger.info(f"Writing filled Excel: {output_file}")
        
        # Load the original workbook to preserve formatting
        wb = load_workbook(input_file)
        
        for sheet_name, result in sheet_results.items():
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                continue
            
            ws = wb[sheet_name]
            header_row_idx = result['header_row_index']
            columns = result['columns']
            filled_items = result.get('filled_items', [])
            
            # Get column indices (1-based for openpyxl)
            unit_col_idx = self._get_column_index_in_ws(ws, columns.get('unit'), header_row_idx)
            rate_col_idx = self._get_column_index_in_ws(ws, columns.get('rate'), header_row_idx)
            reference_col_idx = self._get_column_index_in_ws(ws, columns.get('reference'), header_row_idx)
            reasoning_col_idx = self._get_column_index_in_ws(ws, columns.get('reasoning'), header_row_idx)
            
            # Auto-create "AutoRate Reference" column if it doesn't exist
            if not reference_col_idx:
                reference_col_idx = ws.max_column + 1
                ws.cell(row=header_row_idx + 1, column=reference_col_idx).value = "AutoRate Reference"
                logger.info(f"  Created 'AutoRate Reference' column at index {reference_col_idx}")
            
            # Auto-create "AutoRate Reasoning" column if it doesn't exist
            if not reasoning_col_idx:
                reasoning_col_idx = max(ws.max_column, reference_col_idx) + 1
                ws.cell(row=header_row_idx + 1, column=reasoning_col_idx).value = "AutoRate Reasoning"
                logger.info(f"  Created 'AutoRate Reasoning' column at index {reasoning_col_idx}")
            
            # Process each filled item
            for item in filled_items:
                excel_row = item['row_index'] + 1  # +1 for 1-based indexing
                
                # Determine fill color
                match_type = item.get('match_type', 'exact')
                fill_color = {
                    'exact': self.GREEN_FILL,
                    'close': self.YELLOW_FILL,
                    'approximation': self.ORANGE_FILL
                }.get(match_type, self.GREEN_FILL)
                
                if item['status'] == 'filled':
                    # Fill unit if needed
                    if item.get('filled_unit') and unit_col_idx:
                        cell = ws.cell(row=excel_row, column=unit_col_idx)
                        cell.value = item['filled_unit']
                        cell.fill = fill_color
                    
                    # Fill rate if needed
                    if item.get('filled_rate') is not None and rate_col_idx:
                        cell = ws.cell(row=excel_row, column=rate_col_idx)
                        cell.value = item['filled_rate']
                        cell.fill = fill_color
                    
                    # Fill reference
                    if reference_col_idx:
                        cell = ws.cell(row=excel_row, column=reference_col_idx)
                        ref_value = item.get('reference', '')
                        if match_type in ['close', 'approximation'] and item.get('confidence'):
                            ref_value = f"{ref_value} [{item.get('confidence')}%]"
                        cell.value = ref_value
                    
                    # Fill reasoning
                    if reasoning_col_idx:
                        cell = ws.cell(row=excel_row, column=reasoning_col_idx)
                        cell.value = item.get('reasoning', '')
                
                else:  # not filled
                    if unit_col_idx:
                        ws.cell(row=excel_row, column=unit_col_idx).fill = self.RED_FILL
                    if rate_col_idx:
                        ws.cell(row=excel_row, column=rate_col_idx).fill = self.RED_FILL
        
        # Save workbook
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        logger.info(f"Saved filled Excel to: {output_file}")
        
        return str(output_path)
    
    def _get_column_index_in_ws(
        self,
        ws,
        column_name: Optional[str],
        header_row_idx: int
    ) -> Optional[int]:
        """Get 1-based column index in worksheet."""
        if not column_name:
            return None
        
        header_row = header_row_idx + 1  # 1-based
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value and str(cell_value).strip() == str(column_name).strip():
                return col
        
        return None
