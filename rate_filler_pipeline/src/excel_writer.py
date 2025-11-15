"""
Excel Writer - Write filled Excel files with color coding.
Copies original Excel and fills missing unit/rate values.
"""
import pandas as pd
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Write Excel files with filled rates and color coding."""
    
    def __init__(self):
        self.logger = logger
        self.green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Exact match
        self.yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Close match
        self.blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Approximation
        self.red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Not filled
    
    def write_filled_excel(
        self,
        input_file: str,
        output_file: str,
        sheet_results: Dict[str, Any]
    ) -> str:
        """
        Write filled Excel file with color coding.
        
        Args:
            input_file: Path to original Excel file
            output_file: Path to output Excel file
            sheet_results: Results from rate filling process
                {
                    'sheet_name': {
                        'dataframe': pd.DataFrame,
                        'header_row_index': int,
                        'columns': dict,
                        'filled_items': [
                            {
                                'row_index': int,
                                'item_code': str,
                                'description': str,
                                'filled_unit': str or None,
                                'filled_rate': float or None,
                                'status': 'filled' or 'not_filled'
                            }
                        ]
                    }
                }
        
        Returns:
            Path to output file
        """
        logger.info(f"Writing filled Excel: {output_file}")
        
        # Load the original workbook to preserve formatting
        wb = load_workbook(input_file)
        
        for sheet_name, result in sheet_results.items():
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                continue
            
            ws = wb[sheet_name]
            header_row_idx = result['header_row_index']
            columns = result['columns']
            filled_items = result['filled_items']
            
            # Get column indices (1-based for openpyxl)
            unit_col_idx = self._get_column_index(ws, columns['unit'], header_row_idx)
            rate_col_idx = self._get_column_index(ws, columns['rate'], header_row_idx)
            reference_col_idx = self._get_column_index(ws, columns.get('reference'), header_row_idx)
            reasoning_col_idx = self._get_column_index(ws, columns.get('reasoning'), header_row_idx)
            
            # Auto-create "AutoRate Reference" column if it doesn't exist
            if not reference_col_idx:
                # Find the last used column
                last_col = ws.max_column
                reference_col_idx = last_col + 1
                
                # Add header for new column
                header_cell = ws.cell(row=header_row_idx, column=reference_col_idx)
                header_cell.value = "AutoRate Reference"
                
                logger.info(f"  Created new 'AutoRate Reference' column at index {reference_col_idx}")
            
            # Auto-create "AutoRate Reasoning" column if it doesn't exist
            if not reasoning_col_idx:
                # Find the last used column (after reference)
                last_col = max(ws.max_column, reference_col_idx)
                reasoning_col_idx = last_col + 1
                
                # Add header for new column
                header_cell = ws.cell(row=header_row_idx, column=reasoning_col_idx)
                header_cell.value = "AutoRate Reasoning"
                
                logger.info(f"  Created new 'AutoRate Reasoning' column at index {reasoning_col_idx}")
            
            logger.info(f"Processing sheet '{sheet_name}':")
            logger.info(f"  Unit column: {columns['unit']} (index {unit_col_idx})")
            logger.info(f"  Rate column: {columns['rate']} (index {rate_col_idx})")
            logger.info(f"  Reference column: index {reference_col_idx}")
            logger.info(f"  Reasoning column: index {reasoning_col_idx}")
            
            # Process each filled item
            for item in filled_items:
                # openpyxl uses 1-based indexing, and we need to account for header
                excel_row = item['row_index'] + 1  # +1 for 1-based indexing
                
                # Determine fill color based on match type
                match_type = item.get('match_type', 'exact')
                if match_type == 'exact':
                    fill_color = self.green_fill
                elif match_type == 'close':
                    fill_color = self.yellow_fill
                elif match_type == 'approximation':
                    fill_color = self.blue_fill
                else:
                    fill_color = self.green_fill  # Default
                
                if item['status'] == 'filled':
                    # Fill unit if needed
                    if item.get('filled_unit') and unit_col_idx:
                        cell = ws.cell(row=excel_row, column=unit_col_idx)
                        cell.value = item['filled_unit']
                        cell.fill = fill_color
                        logger.debug(f"Row {excel_row}: Filled unit = {item['filled_unit']} ({match_type})")
                    
                    # Fill rate if needed
                    if item.get('filled_rate') is not None and rate_col_idx:
                        cell = ws.cell(row=excel_row, column=rate_col_idx)
                        cell.value = item['filled_rate']
                        cell.fill = fill_color
                        logger.debug(f"Row {excel_row}: Filled rate = {item['filled_rate']} ({match_type})")
                    
                    # Always fill reference for filled items
                    if reference_col_idx:
                        cell = ws.cell(row=excel_row, column=reference_col_idx)
                        ref_value = item.get('reference', '')
                        # Add confidence to reference for similar matches
                        if match_type == 'similar' and item.get('confidence'):
                            ref_value = f"{ref_value} [Confidence: {item.get('confidence')}%]"
                        cell.value = ref_value
                        cell.fill = fill_color
                        logger.debug(f"Row {excel_row}: Filled reference = {item.get('reference', '')}")
                    
                    # Always fill reasoning for filled items
                    if reasoning_col_idx:
                        cell = ws.cell(row=excel_row, column=reasoning_col_idx)
                        cell.value = item.get('reasoning', '')
                        cell.fill = fill_color
                        logger.debug(f"Row {excel_row}: Filled reasoning = {item.get('reasoning', '')}")
                
                else:  # not_filled
                    # Mark cells as red (unfilled)
                    if item.get('needs_unit') and unit_col_idx:
                        cell = ws.cell(row=excel_row, column=unit_col_idx)
                        cell.fill = self.red_fill
                        logger.debug(f"Row {excel_row}: Unit not filled (marked red)")
                    
                    if item.get('needs_rate') and rate_col_idx:
                        cell = ws.cell(row=excel_row, column=rate_col_idx)
                        cell.fill = self.red_fill
                        logger.debug(f"Row {excel_row}: Rate not filled (marked red)")
                    
                    # Fill reasoning for not_filled items if available
                    if reasoning_col_idx and item.get('reasoning'):
                        cell = ws.cell(row=excel_row, column=reasoning_col_idx)
                        cell.value = item.get('reasoning', '')
                        # No fill color for reasoning when not filled
        
        # Save workbook
        wb.save(output_file)
        logger.info(f"✓ Filled Excel saved: {output_file}")
        
        return output_file
    
    def _get_column_index(self, ws, column_name: str, header_row_idx: int) -> int:
        """
        Get column index for a given column name.
        
        Args:
            ws: Worksheet
            column_name: Name of the column
            header_row_idx: Index of header row (0-based in pandas, need to convert)
            
        Returns:
            Column index (1-based for openpyxl), or None if not found
        """
        if not column_name:
            return None
        
        # Header row in openpyxl is 1-based
        header_row = header_row_idx + 1
        
        # Search for column name in header row
        for cell in ws[header_row]:
            if cell.value and str(cell.value).strip() == column_name:
                return cell.column
        
        logger.warning(f"Column '{column_name}' not found in header row {header_row}")
        return None
    
    def write_report(
        self,
        output_file: str,
        sheet_results: Dict[str, Any],
        summary: Dict[str, int]
    ) -> str:
        """
        Write detailed text report of filling results.
        
        Args:
            output_file: Path to report file
            sheet_results: Results from rate filling
            summary: Summary statistics
                {
                    'total_items': int,
                    'filled': int,
                    'not_filled': int
                }
        
        Returns:
            Path to report file
        """
        logger.info(f"Writing report: {output_file}")
        
        with open(output_file, 'w', encoding='utf-8') as f:
            # Header
            f.write("=" * 80 + "\n")
            f.write("RATE FILLING REPORT\n")
            f.write("=" * 80 + "\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            # Summary - compute method breakdown across all sheets
            total_items = summary.get('total_items', 0)
            total_filled = summary.get('filled', 0)
            total_not_filled = summary.get('not_filled', 0)

            # Count by match type
            exact_total = 0
            close_total = 0
            approx_total = 0
            for res in sheet_results.values():
                for it in res.get('filled_items', []):
                    if it.get('status') != 'filled':
                        continue
                    mt = it.get('match_type', 'exact')
                    if mt == 'exact':
                        exact_total += 1
                    elif mt == 'close':
                        close_total += 1
                    elif mt == 'approximation':
                        approx_total += 1

            # Defensive division
            def pct(part, whole):
                return (part / whole * 100) if whole else 0.0

            f.write("SUMMARY\n")
            f.write("-" * 80 + "\n")
            f.write(f"Total items needing filling: {total_items}\n")
            f.write(f"Items filled: {total_filled} ({pct(total_filled, total_items):.1f}%)\n")
            f.write(f"  - Exact matches: {exact_total} ({pct(exact_total, total_items):.1f}%)\n")
            f.write(f"  - Close matches: {close_total} ({pct(close_total, total_items):.1f}%)\n")
            f.write(f"  - Approximations: {approx_total} ({pct(approx_total, total_items):.1f}%)\n")
            f.write(f"Items not filled: {total_not_filled} ({pct(total_not_filled, total_items):.1f}%)\n\n")
            
            # Details per sheet
            for sheet_name, result in sheet_results.items():
                f.write("\n" + "=" * 80 + "\n")
                f.write(f"SHEET: {sheet_name}\n")
                f.write("=" * 80 + "\n\n")
                
                filled_items = result['filled_items']

                # Per-sheet breakdown
                sheet_total = len(filled_items)
                sheet_filled = sum(1 for i in filled_items if i.get('status') == 'filled')
                sheet_not_filled = sum(1 for i in filled_items if i.get('status') == 'not_filled')
                sheet_exact = sum(1 for i in filled_items if i.get('status') == 'filled' and i.get('match_type') == 'exact')
                sheet_close = sum(1 for i in filled_items if i.get('status') == 'filled' and i.get('match_type') == 'close')
                sheet_approx = sum(1 for i in filled_items if i.get('status') == 'filled' and i.get('match_type') == 'approximation')

                f.write(f"Sheet summary: Total={sheet_total}, Filled={sheet_filled} ({pct(sheet_filled, sheet_total):.1f}%), Not filled={sheet_not_filled} ({pct(sheet_not_filled, sheet_total):.1f}%)\n")
                f.write(f"  Exact: {sheet_exact} ({pct(sheet_exact, sheet_total):.1f}%), Close: {sheet_close} ({pct(sheet_close, sheet_total):.1f}%), Approx: {sheet_approx} ({pct(sheet_approx, sheet_total):.1f}%)\n\n")
                
                for item in filled_items:
                    row_num = item['row_index'] + 1  # Excel row number
                    
                    if item['status'] == 'filled':
                        match_type = item.get('match_type', 'exact')
                        stage = item.get('stage', 'matcher')
                        
                        # Determine symbol and label based on match type
                        if match_type == 'exact':
                            match_symbol = "✓"
                            match_label = "EXACT MATCH"
                        elif match_type == 'close':
                            match_symbol = "≈"
                            match_label = "CLOSE MATCH"
                        elif match_type == 'approximation':
                            match_symbol = "~"
                            match_label = "APPROXIMATION"
                        else:
                            match_symbol = "✓"
                            match_label = "MATCH"
                        
                        f.write(f"Row {row_num}: {match_symbol} {match_label} (Stage: {stage.upper()})\n")
                        f.write(f"  Item: {item['item_code']}\n")
                        f.write(f"  Description: {item['description']}\n")
                        
                        if item.get('filled_unit'):
                            f.write(f"  Unit: {item['filled_unit']}\n")
                        
                        if item.get('filled_rate') is not None:
                            f.write(f"  Rate: {item['filled_rate']:.2f}\n")
                        
                        if match_type in ['close', 'approximation'] and item.get('confidence'):
                            f.write(f"  Confidence: {item.get('confidence')}%\n")
                        
                        if match_type == 'approximation' and item.get('adjustment'):
                            f.write(f"  Adjustment: {item.get('adjustment')}\n")
                        
                        if item.get('reference'):
                            f.write(f"  Reference: {item['reference']}\n")
                        
                        if item.get('reasoning'):
                            f.write(f"  Reasoning: {item['reasoning']}\n")
                        
                        if item.get('match_info'):
                            f.write(f"  Source: {item['match_info'].get('source', 'Unknown')}\n")
                        
                    else:  # not_filled
                        f.write(f"Row {row_num}: ✗ NOT FILLED\n")
                        f.write(f"  Item: {item['item_code']}\n")
                        f.write(f"  Description: {item['description']}\n")
                        
                        if item.get('reasoning'):
                            f.write(f"  Reasoning: {item['reasoning']}\n")
                        elif item.get('reason'):
                            f.write(f"  Reason: {item['reason']}\n")
                    
                    f.write("\n")
        
        logger.info(f"✓ Report saved: {output_file}")
        
        return output_file
    
    def generate_output_filename(self, input_file: str, suffix: str = 'filled') -> str:
        """
        Generate output filename with timestamp.
        
        Args:
            input_file: Original input filename
            suffix: Suffix to add before extension
            
        Returns:
            Output filename
        """
        input_path = Path(input_file)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_name = f"{input_path.stem}_{suffix}_{timestamp}{input_path.suffix}"
        
        return output_name
